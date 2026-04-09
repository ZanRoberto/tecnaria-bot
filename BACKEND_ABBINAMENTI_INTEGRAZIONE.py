"""
BACKEND ABBINAMENTI — Da aggiungere a app.py di Oracolo Covolo

ISTRUZIONI:
1. Copia tutto il codice sotto
2. Incollalo in app.py PRIMA di app.run()
3. Riavvia l'app
4. Testa: GET /api/abbina/GSS-316-001
"""

# ============================================================================
# 1. AGGIUNGI QUESTE RIGHE A init_db() — DOPO le altre CREATE TABLE
# ============================================================================

def init_db():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    
    # ... tabelle esistenti ...
    
    # AGGIUNGI QUESTE:
    
    # Tabella categorie accessori
    c.execute('''CREATE TABLE IF NOT EXISTS categories_accessori (
        id INTEGER PRIMARY KEY,
        categoria_id TEXT UNIQUE NOT NULL,
        categoria_nome TEXT NOT NULL,
        descrizione TEXT,
        icona TEXT,
        created_at TEXT
    )''')
    
    # Tabella abbinamenti
    c.execute('''CREATE TABLE IF NOT EXISTS product_accessories (
        id INTEGER PRIMARY KEY,
        prodotto_padre TEXT NOT NULL,
        accessorio_id TEXT NOT NULL,
        accessorio_nome TEXT,
        brand_accessorio TEXT,
        categoria_accessorio TEXT,
        tipo_relazione TEXT,
        priority INTEGER DEFAULT 99,
        note TEXT,
        created_at TEXT,
        FOREIGN KEY (categoria_accessorio) REFERENCES categories_accessori(categoria_id),
        UNIQUE(prodotto_padre, accessorio_id)
    )''')
    
    # Tabella vincoli
    c.execute('''CREATE TABLE IF NOT EXISTS product_accessories_vincoli (
        id INTEGER PRIMARY KEY,
        relazione_id INTEGER NOT NULL,
        campo_vincolo TEXT,
        valore_vincolo TEXT,
        severity TEXT,
        messaggio TEXT,
        created_at TEXT,
        FOREIGN KEY (relazione_id) REFERENCES product_accessories(id)
    )''')
    
    # Tabella regole matching
    c.execute('''CREATE TABLE IF NOT EXISTS matching_rules (
        id INTEGER PRIMARY KEY,
        categoria_prodotto TEXT,
        categoria_accessorio TEXT,
        soglia_compatibilita TEXT,
        nota TEXT,
        created_at TEXT,
        FOREIGN KEY (categoria_accessorio) REFERENCES categories_accessori(categoria_id)
    )''')
    
    conn.commit()
    conn.close()

# ============================================================================
# 2. FUNZIONE PER CARICARE ACCESSORI DA EXCEL
# ============================================================================

def load_accessories_from_excel(file_content, brand):
    """Carica categorie, abbinamenti, vincoli da Excel"""
    import base64 as b64
    import io as iomod
    import openpyxl as oxl
    
    try:
        # Decodifica base64
        if ',' in file_content:
            file_content = file_content.split(',', 1)[1]
        
        raw = b64.b64decode(file_content)
        wb = oxl.load_workbook(iomod.BytesIO(raw), data_only=True)
        
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        
        # ---- CARICA CATEGORIE ----
        if "CATEGORIE_ACCESSORI" in wb.sheetnames:
            ws = wb["CATEGORIE_ACCESSORI"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue
                categoria_id = row[0]
                categoria_nome = row[1]
                descrizione = row[2] if len(row) > 2 else ""
                icona = row[3] if len(row) > 3 else ""
                
                try:
                    c.execute("""
                        INSERT OR REPLACE INTO categories_accessori 
                        (categoria_id, categoria_nome, descrizione, icona, created_at)
                        VALUES (?, ?, ?, ?, ?)
                    """, (categoria_id, categoria_nome, descrizione, icona, datetime.now().isoformat()))
                except:
                    pass
        
        # ---- CARICA ABBINAMENTI ----
        if "ABBINAMENTI" in wb.sheetnames:
            ws = wb["ABBINAMENTI"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue
                
                prodotto_padre = row[0]
                accessorio_id = row[1]
                accessorio_nome = row[2] if len(row) > 2 else ""
                brand_accessorio = row[3] if len(row) > 3 else brand
                categoria_accessorio = row[4] if len(row) > 4 else ""
                tipo_relazione = row[5] if len(row) > 5 else "ufficiale"
                priority = row[6] if len(row) > 6 else 99
                
                vincoli_campo = row[7] if len(row) > 7 else None
                vincoli_valore = row[8] if len(row) > 8 else None
                vincoli_severity = row[9] if len(row) > 9 else None
                vincoli_messaggio = row[10] if len(row) > 10 else None
                
                note = row[11] if len(row) > 11 else ""
                
                try:
                    c.execute("""
                        INSERT OR REPLACE INTO product_accessories
                        (prodotto_padre, accessorio_id, accessorio_nome, brand_accessorio,
                         categoria_accessorio, tipo_relazione, priority, note, created_at)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, (prodotto_padre, accessorio_id, accessorio_nome, brand_accessorio,
                          categoria_accessorio, tipo_relazione, int(priority) if priority else 99,
                          note, datetime.now().isoformat()))
                    
                    rel_id = c.lastrowid
                    
                    if vincoli_campo:
                        c.execute("""
                            INSERT INTO product_accessories_vincoli
                            (relazione_id, campo_vincolo, valore_vincolo, severity, messaggio, created_at)
                            VALUES (?, ?, ?, ?, ?, ?)
                        """, (rel_id, vincoli_campo, vincoli_valore, vincoli_severity,
                              vincoli_messaggio, datetime.now().isoformat()))
                except Exception as e:
                    print(f"[ABBINAMENTI] Errore: {e}")
        
        # ---- CARICA REGOLE ----
        if "REGOLE_MATCHING" in wb.sheetnames:
            ws = wb["REGOLE_MATCHING"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue
                
                categoria_prodotto = row[0]
                categoria_accessorio = row[1]
                soglia = row[2] if len(row) > 2 else "media"
                nota = row[3] if len(row) > 3 else ""
                
                try:
                    c.execute("""
                        INSERT INTO matching_rules
                        (categoria_prodotto, categoria_accessorio, soglia_compatibilita, nota, created_at)
                        VALUES (?, ?, ?, ?, ?)
                    """, (categoria_prodotto, categoria_accessorio, soglia, nota,
                          datetime.now().isoformat()))
                except:
                    pass
        
        conn.commit()
        conn.close()
        return True, "Accessori caricati con successo"
    except Exception as e:
        return False, f"Errore caricamento: {str(e)}"

# ============================================================================
# 3. ENDPOINT: POST /api/upload-accessories
# ============================================================================

@app.route('/api/upload-accessories', methods=['POST'])
def upload_accessories():
    """Upload Excel con accessori per un brand"""
    if 'file' not in request.files:
        return jsonify({"error": "File non fornito"}), 400
    
    file = request.files['file']
    brand = request.form.get('brand', 'Gessi')
    
    if file.filename == '':
        return jsonify({"error": "Filename vuoto"}), 400
    
    try:
        file_bytes = file.read()
        import base64
        b64_content = base64.b64encode(file_bytes).decode()
        
        success, msg = load_accessories_from_excel(b64_content, brand)
        
        if success:
            return jsonify({"ok": True, "message": msg})
        else:
            return jsonify({"error": msg}), 400
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ============================================================================
# 4. ENDPOINT: GET /api/abbina/<prodotto_id>
# ============================================================================

@app.route('/api/abbina/<prodotto_id>', methods=['GET'])
def abbina(prodotto_id):
    """Restituisce abbinamenti per un prodotto con filtraggio per vincoli"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    
    # Prendi il prodotto principale da Excel Gessi
    c.execute("""
        SELECT d.content FROM documents d
        JOIN aziende a ON d.azienda_id = a.id
        WHERE LOWER(a.nome) LIKE '%gessi%' AND d.filename LIKE '%[EXCEL]%'
        ORDER BY d.upload_date DESC LIMIT 1
    """)
    row = c.fetchone()
    
    prodotto_info = {}
    if row:
        try:
            import base64 as b64m, io as iom, openpyxl as oxl2
            content = row[0]
            if ',' in content:
                content = content.split(',', 1)[1]
            raw = b64m.b64decode(content)
            wb = oxl2.load_workbook(iom.BytesIO(raw), data_only=True)
            ws = wb.active
            
            for r in ws.iter_rows(min_row=3, values_only=True):
                if r[1] == prodotto_id:
                    prodotto_info = {
                        'codice': r[1],
                        'nome': r[2],
                        'collezione': r[3],
                        'categoria': r[4],
                        'prezzo_cliente': r[5],
                        'prezzo_riv': r[6]
                    }
                    break
        except:
            pass
    
    # Prendi gli abbinamenti
    c.execute("""
        SELECT pa.*, pav.campo_vincolo, pav.valore_vincolo, pav.severity, pav.messaggio
        FROM product_accessories pa
        LEFT JOIN product_accessories_vincoli pav ON pa.id = pav.relazione_id
        WHERE pa.prodotto_padre = ?
        ORDER BY pa.tipo_relazione DESC, pa.priority ASC
    """, (prodotto_id,))
    
    abbinamenti = c.fetchall()
    conn.close()
    
    risultati = {
        "prodotto": prodotto_info,
        "ufficiali": [],
        "alternative": [],
        "esclusi": []
    }
    
    for acc in abbinamenti:
        acc_id = acc[2]
        acc_nome = acc[3]
        brand = acc[4]
        categoria = acc[5]
        tipo_rel = acc[6]
        priority = acc[7]
        note = acc[8]
        campo_vincolo = acc[10]
        valore_vincolo = acc[11]
        severity = acc[12]
        messaggio_vincolo = acc[13]
        
        compatibile = True
        if campo_vincolo:
            campo_prodotto = prodotto_info.get(campo_vincolo)
            if campo_prodotto != valore_vincolo:
                compatibile = False
        
        item = {
            "id": acc_id,
            "nome": acc_nome,
            "brand": brand,
            "categoria": categoria,
            "priority": priority,
            "nota_prodotto": note,
            "compatibile": compatibile,
            "vincolo_messaggio": messaggio_vincolo,
            "severity": severity or "soft"
        }
        
        if severity == 'hard' and not compatibile:
            risultati["esclusi"].append(item)
        elif tipo_rel == 'ufficiale':
            risultati["ufficiali"].append(item)
        else:
            risultati["alternative"].append(item)
    
    return jsonify(risultati)

