"""
ORACOLO COVOLO - SISTEMA COMPLETO V2
+ Login 3 livelli: superadmin (Tecnaria) / admin cliente / commerciale
+ Moduli ON/OFF configurabili da superadmin per cliente
+ Pannello destra: Cantieri, Carrello, BI
+ Tutto il precedente invariato
"""
import os, json, sqlite3, re, hashlib, secrets, base64, io
from datetime import datetime
from flask import Flask, render_template_string, request, jsonify, session
import httpx
try:
    import openpyxl
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, "data")
DB_PATH = os.path.join(DATA_DIR, "oracolo_covolo.db")
os.makedirs(DATA_DIR, exist_ok=True)

OPENAI_API_KEY = os.getenv("OPENAI_API_KEY", "").strip()
OPENAI_MODEL = os.getenv("OPENAI_MODEL", "gpt-4o")
SUPERADMIN_PASSWORD = os.getenv("SUPERADMIN_PASSWORD", "ZANNA1959?")
GOOGLE_API_KEY = os.getenv("GOOGLE_API_KEY", "").strip()
GOOGLE_CSE_ID = os.getenv("GOOGLE_CSE_ID", "").strip()

BRANDS_LIST = [
    "Acquabella", "Altamarea", "Anem", "Antoniolupi", "Aparici", "Apavisa",
    "Ariostea", "Artesia", "Austroflamm", "BGP", "Brera", "Bisazza",
    "Blue Design", "Baufloor", "Bauwerk", "Caros", "Caesar", "Casalgrande Padana",
    "Cerasarda", "Cerasa", "Cielo", "Colombo", "Cottodeste", "CP Parquet",
    "CSA", "Decor Walther", "Demm", "DoorAmeda", "Duscholux", "Duravit",
    "Edimax Astor", "FAP Ceramiche", "FMG", "Floorim", "Gerflor", "Gessi",
    "Gigacer", "Glamm Fire", "GOman", "Gridiron", "Gruppo Bardelli", "Gruppo Geromin",
    "Ier Hurne", "Inklostro Bianco", "Iniziativa Legno", "Iris", "Italgraniti",
    "Kaldewei", "Linki", "Madegan", "Marca Corona", "Mirage", "Milldue",
    "Murexin", "Noorth", "Omegius", "Piastrelle d Arredo", "Profiletec", "Remer",
    "Sichenia", "Simas", "Schluter Systems", "SDR", "Sterneldesign", "Stuv",
    "Sunshower", "Sunshower Wellness", "Tonalite", "Tresse", "Trimline Fires",
    "Tubes", "Valdama", "Vismara Vetro", "Wedi"
]

MODULI_DISPONIBILI = ["cantieri", "carrello", "bi", "commerciali"]

def init_db():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    # Tabelle esistenti
    c.execute('''CREATE TABLE IF NOT EXISTS aziende (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nome TEXT UNIQUE NOT NULL,
        admin_password TEXT,
        admin_required BOOLEAN DEFAULT 0
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS documents (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        filename TEXT, content TEXT, azienda_id INTEGER,
        visibility TEXT DEFAULT 'public', access_code TEXT, upload_date TEXT,
        FOREIGN KEY (azienda_id) REFERENCES aziende(id)
    )''')
    # Tabelle nuove
    c.execute('''CREATE TABLE IF NOT EXISTS clienti (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nome TEXT UNIQUE NOT NULL,
        slug TEXT UNIQUE NOT NULL
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS utenti (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nome TEXT NOT NULL,
        username TEXT UNIQUE NOT NULL,
        password_hash TEXT NOT NULL,
        ruolo TEXT NOT NULL,
        cliente_id INTEGER,
        attivo INTEGER DEFAULT 1,
        FOREIGN KEY (cliente_id) REFERENCES clienti(id)
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS moduli_cliente (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        cliente_id INTEGER NOT NULL,
        modulo TEXT NOT NULL,
        attivo INTEGER DEFAULT 0,
        UNIQUE(cliente_id, modulo),
        FOREIGN KEY (cliente_id) REFERENCES clienti(id)
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS cantieri (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        cliente_id INTEGER NOT NULL,
        commerciale_id INTEGER,
        nome TEXT NOT NULL,
        configurazione_piani JSON,
        stato TEXT DEFAULT 'bozza',
        modalita TEXT DEFAULT 'semplice',
        note TEXT,
        data_creazione TEXT,
        data_aggiornamento TEXT,
        FOREIGN KEY (cliente_id) REFERENCES clienti(id),
        FOREIGN KEY (commerciale_id) REFERENCES utenti(id)
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS cantiere_righe (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        cantiere_id INTEGER NOT NULL,
        prodotto_codice TEXT,
        piano TEXT,
        ambiente TEXT,
        quantita INTEGER DEFAULT 1,
        prezzo_unitario REAL,
        subtotale REAL,
        brand TEXT, categoria TEXT, descrizione TEXT, note TEXT,
        importo REAL DEFAULT 0,
        FOREIGN KEY (cantiere_id) REFERENCES cantieri(id),
        FOREIGN KEY (prodotto_codice) REFERENCES products(codice)
    )''')
    
    # TABELLE PIANI/STANZE/VOCI (MODALITA PIANI)
    c.execute('''CREATE TABLE IF NOT EXISTS piani (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        cantiere_id INTEGER NOT NULL,
        numero INTEGER,
        nome TEXT,
        totale_piano REAL DEFAULT 0,
        created_at TEXT,
        updated_at TEXT,
        FOREIGN KEY (cantiere_id) REFERENCES cantieri(id)
    )''')
    
    c.execute('''CREATE TABLE IF NOT EXISTS stanze (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        piano_id INTEGER NOT NULL,
        nome TEXT,
        descrizione TEXT,
        totale_stanza REAL DEFAULT 0,
        created_at TEXT,
        updated_at TEXT,
        FOREIGN KEY (piano_id) REFERENCES piani(id)
    )''')
    
    c.execute('''CREATE TABLE IF NOT EXISTS stanza_voci (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        stanza_id INTEGER NOT NULL,
        tipo TEXT,
        codice TEXT,
        brand TEXT,
        descrizione TEXT,
        quantita REAL DEFAULT 1,
        udm TEXT,
        prezzo_unitario REAL DEFAULT 0,
        sconto_percentuale REAL DEFAULT 0,
        sconto_fisso REAL DEFAULT 0,
        subtotale REAL DEFAULT 0,
        colore TEXT DEFAULT 'verde',
        note TEXT,
        immagine_b64 TEXT,
        abbinamenti TEXT,
        created_at TEXT,
        updated_at TEXT,
        FOREIGN KEY (stanza_id) REFERENCES stanze(id)
    )''')
    
    # CARRELLO PER STANZE (voci aggiunte per ogni stanza)
    c.execute('''CREATE TABLE IF NOT EXISTS carrello_stanze (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        stanza_id INTEGER NOT NULL,
        prodotto_codice TEXT,
        prodotto_nome TEXT,
        brand TEXT,
        quantita REAL DEFAULT 1,
        prezzo_unitario REAL DEFAULT 0,
        sconto_percentuale REAL DEFAULT 0,
        subtotale REAL DEFAULT 0,
        colore TEXT DEFAULT 'verde',
        created_at TEXT,
        updated_at TEXT,
        FOREIGN KEY (stanza_id) REFERENCES stanze(id)
    )''')
    
    # CONFIGURAZIONE SISTEMA (sconto mode, etc)
    c.execute('''CREATE TABLE IF NOT EXISTS config_sistema (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        chiave TEXT UNIQUE NOT NULL,
        valore TEXT,
        descrizione TEXT,
        updated_at TEXT
    )''')
    

    # TABELLE LAZY LOADING ABBINAMENTI
    c.execute('''CREATE TABLE IF NOT EXISTS products (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        codice TEXT UNIQUE NOT NULL,
        nome TEXT,
        collezione TEXT,
        categoria TEXT,
        prezzo REAL,
        prezzo_rivenditore REAL,
        prezzo_scontato REAL,
        disponibilita TEXT,
        descrizione TEXT,
        finiture TEXT,
        fonte TEXT,
        brand TEXT,
        image_url TEXT,
        created_at TEXT
    )''')
    
    c.execute('''CREATE TABLE IF NOT EXISTS categories_accessori (
        id INTEGER PRIMARY KEY,
        categoria_id TEXT UNIQUE NOT NULL,
        categoria_nome TEXT NOT NULL,
        descrizione TEXT,
        icona TEXT,
        created_at TEXT
    )''')
    
    c.execute('''CREATE TABLE IF NOT EXISTS product_accessories (
        id INTEGER PRIMARY KEY,
        prodotto_padre TEXT NOT NULL,
        accessorio_id TEXT NOT NULL,
        accessorio_nome TEXT,
        brand_accessorio TEXT,
        categoria_accessorio TEXT,
        tipo_relazione TEXT,
        priority INTEGER DEFAULT 99,
        note TEXT,
        image_url TEXT,
        created_at TEXT,
        FOREIGN KEY (categoria_accessorio) REFERENCES categories_accessori(categoria_id),
        UNIQUE(prodotto_padre, accessorio_id)
    )''')
    
    c.execute('''CREATE TABLE IF NOT EXISTS product_accessories_vincoli (
        id INTEGER PRIMARY KEY,
        relazione_id INTEGER NOT NULL,
        campo_vincolo TEXT,
        valore_vincolo TEXT,
        severity TEXT,
        messaggio TEXT,
        created_at TEXT,
        FOREIGN KEY (relazione_id) REFERENCES product_accessories(id)
    )''')
    
    c.execute('''CREATE TABLE IF NOT EXISTS matching_rules (
        id INTEGER PRIMARY KEY,
        categoria_prodotto TEXT,
        categoria_accessorio TEXT,
        soglia_compatibilita TEXT,
        nota TEXT,
        created_at TEXT,
        FOREIGN KEY (categoria_accessorio) REFERENCES categories_accessori(categoria_id)
    )''')
    
    # TABELLE PIANI/STANZE/VOCI (MODALITA PIANI)
    c.execute('''CREATE TABLE IF NOT EXISTS piani (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        cantiere_id INTEGER NOT NULL,
        numero INTEGER,
        nome TEXT,
        totale_piano REAL DEFAULT 0,
        created_at TEXT,
        updated_at TEXT,
        FOREIGN KEY (cantiere_id) REFERENCES cantieri(id)
    )''')
    
    c.execute('''CREATE TABLE IF NOT EXISTS stanze (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        piano_id INTEGER NOT NULL,
        nome TEXT,
        descrizione TEXT,
        totale_stanza REAL DEFAULT 0,
        created_at TEXT,
        updated_at TEXT,
        FOREIGN KEY (piano_id) REFERENCES piani(id)
    )''')
    
    c.execute('''CREATE TABLE IF NOT EXISTS stanza_voci (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        stanza_id INTEGER NOT NULL,
        tipo TEXT,
        codice TEXT,
        brand TEXT,
        descrizione TEXT,
        quantita REAL DEFAULT 1,
        udm TEXT,
        prezzo_unitario REAL DEFAULT 0,
        sconto_percentuale REAL DEFAULT 0,
        sconto_fisso REAL DEFAULT 0,
        subtotale REAL DEFAULT 0,
        colore TEXT DEFAULT 'verde',
        note TEXT,
        immagine_b64 TEXT,
        created_at TEXT,
        updated_at TEXT,
        FOREIGN KEY (stanza_id) REFERENCES stanze(id)
    )''')
    
    # CARRELLO PER STANZE
    c.execute('''CREATE TABLE IF NOT EXISTS carrello_stanze (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        stanza_id INTEGER NOT NULL,
        prodotto_codice TEXT,
        prodotto_nome TEXT,
        brand TEXT,
        quantita REAL DEFAULT 1,
        prezzo_unitario REAL DEFAULT 0,
        sconto_percentuale REAL DEFAULT 0,
        subtotale REAL DEFAULT 0,
        colore TEXT DEFAULT 'verde',
        created_at TEXT,
        updated_at TEXT,
        FOREIGN KEY (stanza_id) REFERENCES stanze(id)
    )''')
    
    conn.commit()
    # Inizializza moduli per clienti esistenti
    c.execute("SELECT id FROM clienti")
    for (cid,) in c.fetchall():
        for m in MODULI_DISPONIBILI:
            c.execute("INSERT OR IGNORE INTO moduli_cliente (cliente_id, modulo, attivo) VALUES (?,?,0)", (cid, m))
    conn.commit()
    # Brand default
    for brand in BRANDS_LIST:
        c.execute('INSERT OR IGNORE INTO aziende (nome) VALUES (?)', (brand,))
    conn.commit()
    conn.close()

app = Flask(__name__)
app.secret_key = os.getenv("SECRET_KEY", secrets.token_hex(32))
init_db()

def dedup_brands_on_start():
    """Unifica brand duplicati case-insensitive all'avvio"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT LOWER(nome), GROUP_CONCAT(id), GROUP_CONCAT(nome) FROM aziende GROUP BY LOWER(nome) HAVING COUNT(*) > 1")
    dups = c.fetchall()
    for lower_name, ids_str, names_str in dups:
        ids = [int(x) for x in ids_str.split(',')]
        names = names_str.split(',')
        canonical_name = next((n for n in names if n in BRANDS_LIST), names[0])
        canonical_id = ids[names.index(canonical_name)]
        for i, bid in enumerate(ids):
            if bid != canonical_id:
                c.execute("UPDATE documents SET azienda_id=? WHERE azienda_id=?", (canonical_id, bid))
                c.execute("DELETE FROM aziende WHERE id=?", (bid,))
    conn.commit()
    conn.close()

dedup_brands_on_start()

def load_gessi_abbinamenti_on_start():
    """Placeholder — abbinamenti caricheranno da Excel quando l'utente clicca il bottone"""
    pass

def hash_pwd(pwd):
    return hashlib.sha256(pwd.encode()).hexdigest()

def get_session_user():
    return session.get('user')

def is_superadmin():
    """Verifica se l'utente corrente è superadmin"""
    u = get_session_user()
    return u and u.get('ruolo') == 'superadmin'

def require_login(ruoli=None):
    u = get_session_user()
    if not u:
        return None
    if ruoli and u['ruolo'] not in ruoli:
        return None
    return u

# ---------------------------------------------------------------------------
# HELPER FUNCTIONS — PIANI/STANZE/VOCI
# ---------------------------------------------------------------------------

def calcola_subtotale(prezzo, qty, sconto_perc=0):
    """Calcola subtotale: (prezzo - (prezzo * sconto%)) * qty"""
    return max(0, (prezzo - (prezzo * sconto_perc / 100)) * qty)

def ricalcola_totali_stanza(stanza_id):
    """Ricalcola totale stanza da voci"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    
    # Somma subtotali voci
    c.execute("SELECT COALESCE(SUM(subtotale), 0) FROM stanza_voci WHERE stanza_id = ?", (stanza_id,))
    total = c.fetchone()[0]
    
    # Aggiorna stanza
    c.execute("UPDATE stanze SET totale_stanza = ?, updated_at = ? WHERE id = ?", 
              (total, datetime.now().isoformat(), stanza_id))
    
    # Trova piano e ricalcola anche lui
    c.execute("SELECT piano_id FROM stanze WHERE id = ?", (stanza_id,))
    piano_id = c.fetchone()[0]
    
    c.execute("SELECT COALESCE(SUM(totale_stanza), 0) FROM stanze WHERE piano_id = ?", (piano_id,))
    piano_total = c.fetchone()[0]
    
    c.execute("UPDATE piani SET totale_piano = ?, updated_at = ? WHERE id = ?", 
              (piano_total, datetime.now().isoformat(), piano_id))
    
    # Trova cantiere e ricalcola
    c.execute("SELECT cantiere_id FROM piani WHERE id = ?", (piano_id,))
    cantiere_id = c.fetchone()[0]
    
    c.execute("SELECT COALESCE(SUM(totale_piano), 0) FROM piani WHERE cantiere_id = ?", (cantiere_id,))
    cant_total = c.fetchone()[0]
    
    c.execute("UPDATE cantieri SET data_aggiornamento = ? WHERE id = ?", 
              (datetime.now().isoformat(), cantiere_id))
    
    conn.commit()
    conn.close()

def ricalcola_totali_stanza(stanza_id):
    """Ricalcola totale stanza da voci"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    
    # Somma subtotali voci
    c.execute("SELECT COALESCE(SUM(subtotale), 0) FROM stanza_voci WHERE stanza_id = ?", (stanza_id,))
    total = c.fetchone()[0]
    
    # Aggiorna stanza
    c.execute("UPDATE stanze SET totale_stanza = ?, updated_at = ? WHERE id = ?", 
              (total, datetime.now().isoformat(), stanza_id))
    
    # Trova piano e ricalcola anche lui
    c.execute("SELECT piano_id FROM stanze WHERE id = ?", (stanza_id,))
    piano_row = c.fetchone()
    if piano_row:
        piano_id = piano_row[0]
        c.execute("SELECT COALESCE(SUM(totale_stanza), 0) FROM stanze WHERE piano_id = ?", (piano_id,))
        piano_total = c.fetchone()[0]
        c.execute("UPDATE piani SET totale_piano = ?, updated_at = ? WHERE id = ?", 
                  (piano_total, datetime.now().isoformat(), piano_id))
        
        # Trova cantiere e ricalcola
        c.execute("SELECT cantiere_id FROM piani WHERE id = ?", (piano_id,))
        cant_row = c.fetchone()
        if cant_row:
            cantiere_id = cant_row[0]
            c.execute("UPDATE cantieri SET data_aggiornamento = ? WHERE id = ?", 
                      (datetime.now().isoformat(), cantiere_id))
    
    conn.commit()
    conn.close()

# ---------------------------------------------------------------------------
# API AUTH
# ---------------------------------------------------------------------------

@app.route('/api/login', methods=['POST'])
def login():
    data = request.get_json()
    username = data.get('username', '').strip()
    password = data.get('password', '').strip()
    # Superadmin
    if username == 'superadmin' and password == SUPERADMIN_PASSWORD:
        session['user'] = {'id': 0, 'nome': 'Tecnaria', 'ruolo': 'superadmin', 'cliente_id': None}
        return jsonify({"ok": True, "ruolo": "superadmin", "nome": "Tecnaria"})
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT id, nome, ruolo, cliente_id, password_hash, attivo FROM utenti WHERE username=?", (username,))
    row = c.fetchone()
    conn.close()
    if not row:
        return jsonify({"ok": False, "error": "Utente non trovato"})
    uid, nome, ruolo, cliente_id, pwd_hash, attivo = row
    if not attivo:
        return jsonify({"ok": False, "error": "Utente disabilitato"})
    if hash_pwd(password) != pwd_hash:
        return jsonify({"ok": False, "error": "Password errata"})
    session['user'] = {'id': uid, 'nome': nome, 'ruolo': ruolo, 'cliente_id': cliente_id}
    return jsonify({"ok": True, "ruolo": ruolo, "nome": nome})

@app.route('/api/logout', methods=['POST'])
def logout():
    session.clear()
    return jsonify({"ok": True})

@app.route('/api/me', methods=['GET'])
def me():
    u = get_session_user()
    if not u:
        return jsonify({"logged": False})
    # Carica moduli attivi per il cliente
    moduli = []
    cid = u.get('cliente_id')
    if cid:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute("SELECT modulo FROM moduli_cliente WHERE cliente_id=? AND attivo=1", (cid,))
        moduli = [r[0] for r in c.fetchall()]
        conn.close()
    elif u['ruolo'] == 'superadmin':
        moduli = MODULI_DISPONIBILI
    return jsonify({"logged": True, "user": u, "moduli": moduli})

# ---------------------------------------------------------------------------
# API SUPERADMIN — gestione clienti e moduli
# ---------------------------------------------------------------------------

@app.route('/api/sa/clienti', methods=['GET'])
def sa_get_clienti():
    if not require_login(['superadmin']):
        return jsonify({"error": "Non autorizzato"}), 403
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT id, nome, slug FROM clienti ORDER BY nome")
    clienti = [{"id": r[0], "nome": r[1], "slug": r[2]} for r in c.fetchall()]
    conn.close()
    return jsonify({"clienti": clienti})

@app.route('/api/sa/clienti', methods=['POST'])
def sa_add_cliente():
    if not require_login(['superadmin']):
        return jsonify({"error": "Non autorizzato"}), 403
    data = request.get_json()
    nome = data.get('nome', '').strip()
    slug = data.get('slug', '').strip().lower().replace(' ', '-')
    if not nome or not slug:
        return jsonify({"error": "Nome e slug richiesti"}), 400
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    try:
        c.execute("INSERT INTO clienti (nome, slug) VALUES (?,?)", (nome, slug))
        cid = c.lastrowid
        for m in MODULI_DISPONIBILI:
            c.execute("INSERT OR IGNORE INTO moduli_cliente (cliente_id, modulo, attivo) VALUES (?,?,0)", (cid, m))
        conn.commit()
        conn.close()
        return jsonify({"ok": True, "id": cid})
    except Exception as e:
        conn.close()
        return jsonify({"error": str(e)}), 400

@app.route('/api/sa/moduli/<int:cliente_id>', methods=['GET'])
def sa_get_moduli(cliente_id):
    if not require_login(['superadmin']):
        return jsonify({"error": "Non autorizzato"}), 403
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT modulo, attivo FROM moduli_cliente WHERE cliente_id=?", (cliente_id,))
    moduli = {r[0]: bool(r[1]) for r in c.fetchall()}
    conn.close()
    return jsonify({"moduli": moduli})

@app.route('/api/sa/moduli/<int:cliente_id>', methods=['POST'])
def sa_set_moduli(cliente_id):
    if not require_login(['superadmin']):
        return jsonify({"error": "Non autorizzato"}), 403
    data = request.get_json()
    modulo = data.get('modulo')
    attivo = 1 if data.get('attivo') else 0
    if modulo not in MODULI_DISPONIBILI:
        return jsonify({"error": "Modulo non valido"}), 400
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("INSERT OR REPLACE INTO moduli_cliente (cliente_id, modulo, attivo) VALUES (?,?,?)", (cliente_id, modulo, attivo))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})

@app.route('/api/sa/utenti', methods=['GET'])
def sa_get_utenti():
    if not require_login(['superadmin']):
        return jsonify({"error": "Non autorizzato"}), 403
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("""SELECT u.id, u.nome, u.username, u.ruolo, u.attivo, cl.nome as cliente
                 FROM utenti u LEFT JOIN clienti cl ON u.cliente_id=cl.id ORDER BY u.nome""")
    utenti = [{"id": r[0], "nome": r[1], "username": r[2], "ruolo": r[3], "attivo": r[4], "cliente": r[5]} for r in c.fetchall()]
    conn.close()
    return jsonify({"utenti": utenti})

@app.route('/api/sa/utenti', methods=['POST'])
def sa_add_utente():
    if not require_login(['superadmin']):
        return jsonify({"error": "Non autorizzato"}), 403
    data = request.get_json()
    nome = data.get('nome', '').strip()
    username = data.get('username', '').strip()
    password = data.get('password', '').strip()
    ruolo = data.get('ruolo', 'commerciale')
    cliente_id = data.get('cliente_id')
    if not nome or not username or not password:
        return jsonify({"error": "Dati incompleti"}), 400
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    try:
        c.execute("INSERT INTO utenti (nome, username, password_hash, ruolo, cliente_id) VALUES (?,?,?,?,?)",
                  (nome, username, hash_pwd(password), ruolo, cliente_id))
        conn.commit()
        conn.close()
        return jsonify({"ok": True})
    except Exception as e:
        conn.close()
        return jsonify({"error": str(e)}), 400

@app.route('/api/sa/utenti/<int:uid>', methods=['DELETE'])
def sa_delete_utente(uid):
    if not require_login(['superadmin']):
        return jsonify({"error": "Non autorizzato"}), 403
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("UPDATE utenti SET attivo=0 WHERE id=?", (uid,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})

# ---------------------------------------------------------------------------
# MODALITA SEMPLICE/PIANI - Switch intelligente
# ---------------------------------------------------------------------------

@app.route('/api/cantieri/<int:cid>/modalita', methods=['GET'])
def get_modalita_cantiere(cid):
    """Legge modalita cantiere: 'semplice' o 'piani'"""
    try:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute("SELECT modalita FROM cantieri WHERE id = ?", (cid,))
        row = c.fetchone()
        conn.close()
        modalita = row[0] if row else 'semplice'
        return jsonify({'modalita': modalita}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/cantieri/<int:cid>/modalita', methods=['PUT'])
def set_modalita_cantiere(cid):
    """Cambia modalita cantiere"""
    try:
        data = request.get_json()
        nuova_modalita = data.get('modalita', 'semplice')
        
        if nuova_modalita not in ['semplice', 'piani']:
            return jsonify({'error': 'Modalita non valida'}), 400
        
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute("UPDATE cantieri SET modalita = ?, data_aggiornamento = ? WHERE id = ?",
                 (nuova_modalita, datetime.now().isoformat(), cid))
        conn.commit()
        conn.close()
        
        return jsonify({
            'ok': True,
            'cantiere_id': cid,
            'modalita': nuova_modalita,
            'message': f'Convertito a modalita {nuova_modalita}'
        }), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/cantieri/<int:cid>/struttura', methods=['GET'])
def get_struttura_piani(cid):
    """Legge struttura piani/stanze/voci"""
    try:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute("SELECT id, numero, nome, totale_piano FROM piani WHERE cantiere_id = ? ORDER BY numero", (cid,))
        piani_rows = c.fetchall()
        
        piani = []
        for pid, num, nome, tot_piano in piani_rows:
            c.execute("SELECT id, nome, descrizione, totale_stanza FROM stanze WHERE piano_id = ? ORDER BY nome", (pid,))
            stanze_rows = c.fetchall()
            
            stanze = []
            for sid, snome, sdescrizione, tot_stanza in stanze_rows:
                c.execute("""SELECT id, codice, brand, descrizione, quantita, udm, prezzo_unitario, 
                                    sconto_percentuale, sconto_fisso, subtotale, abbinamenti
                             FROM stanza_voci WHERE stanza_id = ? ORDER BY created_at""", (sid,))
                voci_rows = c.fetchall()
                
                voci = []
                for vid, codice, brand, descrizione, qty, udm, prezzo, sconto_perc, sconto_fisso, subtotale, abbinamenti in voci_rows:
                    voci.append({
                        'id': vid,
                        'codice': codice,
                        'brand': brand,
                        'descrizione': descrizione,
                        'quantita': qty,
                        'udm': udm,
                        'prezzo_unitario': prezzo or 0,
                        'sconto_percentuale': sconto_perc or 0,
                        'sconto_fisso': sconto_fisso or 0,
                        'subtotale': subtotale or 0,
                        'abbinamenti': abbinamenti or None
                    })
                
                stanze.append({
                    'id': sid,
                    'nome': snome,
                    'descrizione': sdescrizione,
                    'totale_stanza': tot_stanza or 0,
                    'voci': voci
                })
            
            piani.append({
                'id': pid,
                'numero': num,
                'nome': nome,
                'totale_piano': tot_piano or 0,
                'stanze': stanze
            })
        
        conn.close()
        return jsonify({'ok': True, 'piani': piani}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/stanze/<int:sid>/voci', methods=['POST'])
def add_voce(sid):
    """POST aggiunge voce a stanza"""
    try:
        data = request.get_json()
        
        # Calcola subtotale
        qty = float(data.get('quantita', 1))
        prezzo = float(data.get('prezzo_unitario', 0))
        sconto = float(data.get('sconto_percentuale', 0))
        sub = calcola_subtotale(prezzo, qty, sconto)
        
        # Abbinamenti da salvare
        abbinamenti = data.get('abbinamenti_selezionati', [])
        abbinamenti_json = json.dumps(abbinamenti)
        
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        
        c.execute("""INSERT INTO stanza_voci 
                    (stanza_id, codice, brand, descrizione, quantita, prezzo_unitario, sconto_percentuale, subtotale, colore, abbinamenti, created_at, updated_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                  (sid, data.get('codice', ''), data.get('brand', ''), data.get('descrizione', ''),
                   qty, prezzo, sconto, sub, data.get('colore', 'verde'),
                   abbinamenti_json,
                   datetime.now().isoformat(), datetime.now().isoformat()))
        
        voce_id = c.lastrowid
        conn.commit()
        conn.close()
        
        ricalcola_totali_stanza(sid)
        
        # ✅ RITORNA ANCHE GLI ABBINAMENTI SALVATI
        return jsonify({
            'ok': True,
            'voce_id': voce_id,
            'abbinamenti': abbinamenti
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/stanza_voci/<int:vid>', methods=['PUT'])
def edit_voce(vid):
    """PUT modifica voce"""
    try:
        data = request.get_json()
        
        qty = float(data.get('quantita', 1))
        prezzo = float(data.get('prezzo_unitario', 0))
        sconto = float(data.get('sconto_percentuale', 0))
        sub = calcola_subtotale(prezzo, qty, sconto)
        
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        
        # Trova stanza
        c.execute("SELECT stanza_id FROM stanza_voci WHERE id = ?", (vid,))
        sid = c.fetchone()[0]
        
        # Aggiorna voce
        c.execute("""UPDATE stanza_voci SET quantita = ?, prezzo_unitario = ?, sconto_percentuale = ?, subtotale = ?, updated_at = ? WHERE id = ?""",
                  (qty, prezzo, sconto, sub, datetime.now().isoformat(), vid))
        
        conn.commit()
        conn.close()
        
        ricalcola_totali_stanza(sid)
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/stanza_voci/<int:vid>', methods=['DELETE'])
def delete_voce(vid):
    """DELETE elimina voce"""
    try:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        
        # Trova stanza
        c.execute("SELECT stanza_id FROM stanza_voci WHERE id = ?", (vid,))
        sid = c.fetchone()[0]
        
        # Elimina voce
        c.execute("DELETE FROM stanza_voci WHERE id = ?", (vid,))
        
        conn.commit()
        conn.close()
        
        ricalcola_totali_stanza(sid)
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# =============================================================================
# API PLANIMETRIA - VISION + AUTO-CREAZIONE PIANI/STANZE
# =============================================================================

@app.route('/api/analizza-planimetria', methods=['POST'])
def analizza_planimetria():
    """
    Riceve un'immagine (PNG/JPG) di una planimetria in base64.
    DeepSeek Vision analizza e crea automaticamente piani + stanze nel cantiere.
    """
    try:
        data = request.get_json()
        cantiere_id = data.get('cantiere_id')
        immagine_base64 = data.get('immagine_base64')
        
        if not cantiere_id or not immagine_base64:
            return jsonify({'ok': False, 'error': 'cantiere_id e immagine_base64 richiesti'}), 400
        
        if not OPENAI_API_KEY:
            return jsonify({'ok': False, 'error': 'OPENAI_API_KEY non configurata'}), 400
        
        # Estrai il base64 puro (rimuovi data:image/jpeg;base64, ecc.)
        if ',' in immagine_base64:
            immagine_base64 = immagine_base64.split(',', 1)[1]
        
        # Prompt per OpenAI Vision
        prompt = """Analizza questa planimetria architetturale e estrai la struttura.

Rispondi SOLO in questo formato JSON, nient'altro:
{
  "piani": [
    {
      "numero": 1,
      "nome": "Piano Terra",
      "stanze": [
        {"nome": "Bagno principale", "mq": 8},
        {"nome": "Cucina", "mq": 15}
      ]
    },
    {
      "numero": 2,
      "nome": "Primo Piano",
      "stanze": [
        {"nome": "Camera da letto", "mq": 20}
      ]
    }
  ]
}

Se la planimetria non è leggibile, rispondi con un JSON valido ma vuoto:
{"piani": []}

IMPORTANTE: Restituisci SOLO il JSON, senza markdown o spiegazioni."""

        # ✅ Chiama OpenAI GPT-4 Vision (gpt-4o è più economico)
        if not OPENAI_API_KEY:
            return jsonify({'ok': False, 'error': 'OPENAI_API_KEY non configurata'}), 400
        
        headers = {
            'Authorization': f'Bearer {OPENAI_API_KEY}',
            'Content-Type': 'application/json'
        }
        
        payload = {
            'model': 'gpt-4o',  # ✅ Miglior rapporto prezzo/performance
            'messages': [
                {
                    'role': 'user',
                    'content': [
                        {'type': 'text', 'text': prompt},
                        {
                            'type': 'image_url',
                            'image_url': {
                                'url': f'data:image/jpeg;base64,{immagine_base64}'
                            }
                        }
                    ]
                }
            ],
            'max_tokens': 2000,
            'temperature': 0.3
        }
        
        resp = httpx.post(
            'https://api.openai.com/v1/chat/completions',
            json=payload,
            headers=headers,
            timeout=60
        )
        
        if resp.status_code != 200:
            return jsonify({
                'ok': False,
                'error': f'DeepSeek error {resp.status_code}: {resp.text}'
            }), 500
        
        result = resp.json()
        if 'choices' not in result or len(result['choices']) == 0:
            return jsonify({'ok': False, 'error': 'No response from DeepSeek'}), 500
        
        risposta_text = result['choices'][0]['message']['content'].strip()
        
        # Estrai JSON dalla risposta (potrebbe avere markdown)
        if '```json' in risposta_text:
            risposta_text = risposta_text.split('```json')[1].split('```')[0].strip()
        elif '```' in risposta_text:
            risposta_text = risposta_text.split('```')[1].split('```')[0].strip()
        
        # Parse JSON
        try:
            struttura = json.loads(risposta_text)
        except json.JSONDecodeError as je:
            return jsonify({
                'ok': False,
                'error': f'JSON parsing failed: {str(je)}',
                'raw': risposta_text
            }), 400
        
        piani_data = struttura.get('piani', [])
        
        if not piani_data:
            return jsonify({
                'ok': True,
                'piani_creati': 0,
                'stanze_create': 0,
                'message': 'Nessun piano rilevato dalla planimetria'
            })
        
        # CREAZIONE nel DB
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        
        piani_creati = 0
        stanze_create = 0
        
        for piano_data in piani_data:
            numero = piano_data.get('numero', piani_creati + 1)
            nome_piano = piano_data.get('nome', f'Piano {numero}')
            
            # Crea piano
            c.execute(
                'INSERT INTO piani (cantiere_id, numero, nome, totale_piano, created_at, updated_at) VALUES (?, ?, ?, ?, ?, ?)',
                (
                    cantiere_id,
                    numero,
                    nome_piano,
                    0,  # totale iniziale
                    datetime.now().isoformat(),
                    datetime.now().isoformat()
                )
            )
            piano_id = c.lastrowid
            piani_creati += 1
            
            # Crea stanze
            stanze_data = piano_data.get('stanze', [])
            for stanza_data in stanze_data:
                nome_stanza = stanza_data.get('nome', 'Stanza senza nome')
                mq = stanza_data.get('mq', 0)
                
                c.execute(
                    'INSERT INTO stanze (piano_id, nome, descrizione, totale_stanza, created_at, updated_at) VALUES (?, ?, ?, ?, ?, ?)',
                    (
                        piano_id,
                        nome_stanza,
                        f'{mq}mq' if mq > 0 else '',  # Metratura come descrizione
                        0,  # totale iniziale
                        datetime.now().isoformat(),
                        datetime.now().isoformat()
                    )
                )
                stanze_create += 1
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'ok': True,
            'piani_creati': piani_creati,
            'stanze_create': stanze_create,
            'message': f'✅ Creati {piani_creati} piani e {stanze_create} stanze'
        })
        
    except Exception as e:
        print(f'[PLANIMETRIA ERROR] {str(e)}')
        import traceback
        traceback.print_exc()
        return jsonify({'ok': False, 'error': str(e)}), 500

# API CANTIERI
# ---------------------------------------------------------------------------

@app.route('/api/cantieri', methods=['GET'])
def get_cantieri():
    u = require_login(['superadmin', 'admin', 'commerciale'])
    if not u:
        return jsonify({"error": "Non autorizzato"}), 403
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    if u['ruolo'] == 'superadmin':
        c.execute("""SELECT ca.id, ca.nome, ca.stato, ca.data_creazione, u.nome as comm, cl.nome as cliente
                     FROM cantieri ca
                     LEFT JOIN utenti u ON ca.commerciale_id=u.id
                     LEFT JOIN clienti cl ON ca.cliente_id=cl.id
                     ORDER BY ca.data_creazione DESC LIMIT 100""")
    elif u['ruolo'] == 'admin':
        c.execute("""SELECT ca.id, ca.nome, ca.stato, ca.data_creazione, u.nome as comm, cl.nome as cliente
                     FROM cantieri ca
                     LEFT JOIN utenti u ON ca.commerciale_id=u.id
                     LEFT JOIN clienti cl ON ca.cliente_id=cl.id
                     WHERE ca.cliente_id=?
                     ORDER BY ca.data_creazione DESC""", (u['cliente_id'],))
    else:
        c.execute("""SELECT ca.id, ca.nome, ca.stato, ca.data_creazione, u.nome as comm, cl.nome as cliente
                     FROM cantieri ca
                     LEFT JOIN utenti u ON ca.commerciale_id=u.id
                     LEFT JOIN clienti cl ON ca.cliente_id=cl.id
                     WHERE ca.commerciale_id=?
                     ORDER BY ca.data_creazione DESC""", (u['id'],))
    rows = c.fetchall()
    conn.close()
    return jsonify({"cantieri": [{"id": r[0], "nome": r[1], "stato": r[2], "data": r[3], "commerciale": r[4], "cliente": r[5]} for r in rows]})

@app.route('/api/cantieri', methods=['POST'])
def add_cantiere():
    u = require_login(['superadmin', 'admin', 'commerciale'])
    if not u:
        return jsonify({"error": "Non autorizzato"}), 403
    data = request.get_json()
    nome = data.get('nome', '').strip()
    if not nome:
        return jsonify({"error": "Nome richiesto"}), 400
    cliente_id = u.get('cliente_id') or 1
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("INSERT INTO cantieri (cliente_id, commerciale_id, nome, stato, data_creazione, data_aggiornamento) VALUES (?,?,?,?,?,?)",
              (cliente_id, u['id'] if u['ruolo'] != 'superadmin' else None, nome, 'bozza', datetime.now().isoformat(), datetime.now().isoformat()))
    cid = c.lastrowid
    conn.commit()
    conn.close()
    return jsonify({"ok": True, "id": cid})

@app.route('/api/cantieri/<int:cid>', methods=['PUT'])
def update_cantiere(cid):
    u = require_login(['superadmin', 'admin', 'commerciale'])
    if not u:
        return jsonify({"error": "Non autorizzato"}), 403
    data = request.get_json()
    stato = data.get('stato')
    note = data.get('note')
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    if stato:
        c.execute("UPDATE cantieri SET stato=?, data_aggiornamento=? WHERE id=?", (stato, datetime.now().isoformat(), cid))
    if note is not None:
        c.execute("UPDATE cantieri SET note=?, data_aggiornamento=? WHERE id=?", (note, datetime.now().isoformat(), cid))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})

@app.route('/api/cantieri/<int:cid>/righe', methods=['GET'])
def get_righe(cid):
    u = require_login(['superadmin', 'admin', 'commerciale'])
    if not u:
        return jsonify({"error": "Non autorizzato"}), 403
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT id, brand, categoria, descrizione, note, importo FROM cantiere_righe WHERE cantiere_id=?", (cid,))
    righe = [{"id": r[0], "brand": r[1], "categoria": r[2], "descrizione": r[3], "note": r[4], "importo": r[5]} for r in c.fetchall()]
    conn.close()
    return jsonify({"righe": righe})

@app.route('/api/cantieri/<int:cid>/righe', methods=['POST'])
def add_riga(cid):
    u = require_login(['superadmin', 'admin', 'commerciale'])
    if not u:
        return jsonify({"error": "Non autorizzato"}), 403
    data = request.get_json()
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("INSERT INTO cantiere_righe (cantiere_id, brand, categoria, descrizione, note, importo) VALUES (?,?,?,?,?,?)",
              (cid, data.get('brand',''), data.get('categoria',''), data.get('descrizione',''), data.get('note',''), data.get('importo',0)))
    rid = c.lastrowid
    c.execute("UPDATE cantieri SET data_aggiornamento=? WHERE id=?", (datetime.now().isoformat(), cid))
    conn.commit()
    conn.close()
    return jsonify({"ok": True, "id": rid})

@app.route('/api/cantieri/righe/<int:rid>', methods=['DELETE'])
def delete_riga(rid):
    u = require_login(['superadmin', 'admin', 'commerciale'])
    if not u:
        return jsonify({"error": "Non autorizzato"}), 403
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("DELETE FROM cantiere_righe WHERE id=?", (rid,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})

@app.route('/api/cantieri/<int:cid>', methods=['DELETE'])
def delete_cantiere(cid):
    u = require_login(['superadmin', 'admin'])
    if not u:
        return jsonify({"error": "Non autorizzato"}), 403
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("DELETE FROM cantiere_righe WHERE cantiere_id=?", (cid,))
    c.execute("DELETE FROM cantieri WHERE id=?", (cid,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})

# ---------------------------------------------------------------------------
# API PIANI E STANZE — endpoint mancanti
# ---------------------------------------------------------------------------

@app.route('/api/cantieri/<int:cid>/piani', methods=['POST'])
def add_piano(cid):
    """Crea un nuovo piano nel cantiere"""
    try:
        data = request.get_json()
        numero = data.get('numero', 1)
        nome = data.get('nome', f'Piano {numero}').strip()
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        now = datetime.now().isoformat()
        c.execute("INSERT INTO piani (cantiere_id, numero, nome, totale_piano, created_at, updated_at) VALUES (?,?,?,0,?,?)",
                  (cid, numero, nome, now, now))
        piano_id = c.lastrowid
        conn.commit()
        conn.close()
        return jsonify({'ok': True, 'id': piano_id})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/piani/<int:pid>/stanze', methods=['POST'])
def add_stanza(pid):
    """Crea una nuova stanza nel piano"""
    try:
        data = request.get_json()
        nome = data.get('nome', 'Stanza').strip()
        descrizione = data.get('descrizione', '')
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        now = datetime.now().isoformat()
        c.execute("INSERT INTO stanze (piano_id, nome, descrizione, totale_stanza, created_at, updated_at) VALUES (?,?,?,0,?,?)",
                  (pid, nome, descrizione, now, now))
        stanza_id = c.lastrowid
        conn.commit()
        conn.close()
        return jsonify({'ok': True, 'id': stanza_id})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ---------------------------------------------------------------------------
# API BI
# ---------------------------------------------------------------------------

@app.route('/api/bi/stats', methods=['GET'])
def bi_stats():
    u = require_login(['superadmin', 'admin'])
    if not u:
        return jsonify({"error": "Non autorizzato"}), 403
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    filter_cid = "" if u['ruolo'] == 'superadmin' else f" AND ca.cliente_id={u['cliente_id']}"
    c.execute("SELECT stato, COUNT(*) FROM cantieri ca WHERE 1=1" + filter_cid + " GROUP BY stato")
    per_stato = {r[0]: r[1] for r in c.fetchall()}
    c.execute("SELECT COALESCE(SUM(cr.importo),0) FROM cantiere_righe cr JOIN cantieri ca ON cr.cantiere_id=ca.id WHERE ca.stato='vinta'" + filter_cid)
    valore_vinto = c.fetchone()[0] or 0
    c.execute("SELECT COALESCE(SUM(cr.importo),0) FROM cantiere_righe cr JOIN cantieri ca ON cr.cantiere_id=ca.id WHERE ca.stato='bozza' OR ca.stato='inviata'" + filter_cid)
    valore_aperto = c.fetchone()[0] or 0
    c.execute("""SELECT u.nome, COUNT(ca.id), COALESCE(SUM(cr.importo),0)
                 FROM cantieri ca
                 LEFT JOIN utenti u ON ca.commerciale_id=u.id
                 LEFT JOIN cantiere_righe cr ON cr.cantiere_id=ca.id
                 WHERE 1=1""" + filter_cid + " GROUP BY ca.commerciale_id")
    per_comm = [{"nome": r[0] or "N/D", "cantieri": r[1], "valore": round(r[2], 2)} for r in c.fetchall()]
    conn.close()
    return jsonify({"per_stato": per_stato, "valore_vinto": round(valore_vinto, 2), "valore_aperto": round(valore_aperto, 2), "per_commerciale": per_comm})

@app.route('/api/bi/cancella', methods=['POST'])
def bi_cancella():
    u = require_login(['superadmin', 'admin'])
    if not u:
        return jsonify({"error": "Non autorizzato"}), 403
    data = request.get_json()
    da = data.get('da', '')
    a = data.get('a', '')
    stati = data.get('stati', [])
    if not da or not a or not stati:
        return jsonify({"error": "Parametri mancanti"}), 400
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    placeholders = ','.join('?' * len(stati))
    filter_cid = () if u['ruolo'] == 'superadmin' else (u['cliente_id'],)
    extra = "" if u['ruolo'] == 'superadmin' else " AND cliente_id=?"
    c.execute("SELECT id FROM cantieri WHERE data_creazione>=? AND data_creazione<=? AND stato IN (" + placeholders + ")" + extra,
              [da, a + "T23:59:59"] + stati + list(filter_cid))
    ids = [r[0] for r in c.fetchall()]
    for cid in ids:
        c.execute("DELETE FROM cantiere_righe WHERE cantiere_id=?", (cid,))
        c.execute("DELETE FROM cantieri WHERE id=?", (cid,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True, "eliminati": len(ids)})

# ---------------------------------------------------------------------------
# API ESISTENTI (invariate)
# ---------------------------------------------------------------------------

@app.route('/api/get-brands', methods=['GET'])
def get_brands():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('SELECT nome FROM aziende ORDER BY nome')
    brands = [row[0] for row in c.fetchall()]
    conn.close()
    return jsonify({"brands": brands})

@app.route('/api/add-azienda', methods=['POST'])
def add_azienda():
    data = request.get_json()
    nome = data.get('nome', '').strip()
    if not nome:
        return jsonify({"error": "Nome richiesto"}), 400
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    try:
        # Controlla se esiste già (case-insensitive)
        c.execute('SELECT id, nome FROM aziende WHERE LOWER(nome) = LOWER(?)', (nome,))
        existing = c.fetchone()
        if existing:
            conn.close()
            return jsonify({"ok": True, "nome": existing[1], "existed": True})
        c.execute('INSERT INTO aziende (nome) VALUES (?)', (nome,))
        conn.commit()
        conn.close()
        return jsonify({"ok": True, "nome": nome})
    except Exception as e:
        conn.close()
        return jsonify({"error": str(e)}), 400

@app.route('/api/sa/dedup-brands', methods=['POST'])
def dedup_brands():
    """Unifica brand duplicati case-insensitive — mantiene quello con più documenti"""
    if not require_login(['superadmin']):
        return jsonify({"error": "Non autorizzato"}), 403
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT LOWER(nome), GROUP_CONCAT(id), GROUP_CONCAT(nome) FROM aziende GROUP BY LOWER(nome) HAVING COUNT(*) > 1")
    dups = c.fetchall()
    merged = 0
    for lower_name, ids_str, names_str in dups:
        ids = [int(x) for x in ids_str.split(',')]
        names = names_str.split(',')
        # Tieni quello con il nome corretto (prima maiuscola) o il primo della BRANDS_LIST
        canonical_name = next((n for n in names if n in BRANDS_LIST), names[0])
        canonical_id = ids[names.index(canonical_name)]
        for i, bid in enumerate(ids):
            if bid != canonical_id:
                # Sposta documenti al brand canonico
                c.execute("UPDATE documents SET azienda_id=? WHERE azienda_id=?", (canonical_id, bid))
                c.execute("UPDATE cantiere_righe SET brand=? WHERE brand=?", (canonical_name, names[i]))
                c.execute("DELETE FROM aziende WHERE id=?", (bid,))
                merged += 1
    conn.commit()
    conn.close()
    return jsonify({"ok": True, "merged": merged})

@app.route('/api/set-admin-password', methods=['POST'])
def set_admin_password():
    data = request.get_json()
    brand = data.get('brand', '')
    admin_password = data.get('admin_password', '').strip()
    if not brand or not admin_password:
        return jsonify({"error": "Brand e password richiesti"}), 400
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    try:
        pwd_hash = hashlib.sha256(admin_password.encode()).hexdigest()
        c.execute('UPDATE aziende SET admin_password=?, admin_required=1 WHERE nome=?', (pwd_hash, brand))
        conn.commit()
        conn.close()
        return jsonify({"ok": True, "message": "Password admin impostata per " + brand})
    except Exception as e:
        conn.close()
        return jsonify({"error": str(e)}), 400

@app.route('/api/verify-admin', methods=['POST'])
def verify_admin():
    data = request.get_json()
    brand = data.get('brand', '')
    admin_password = data.get('admin_password', '')
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    try:
        c.execute('SELECT admin_password, admin_required FROM aziende WHERE nome=?', (brand,))
        result = c.fetchone()
        conn.close()
        if not result:
            return jsonify({"ok": False, "error": "Brand non trovato"})
        pwd_hash, admin_required = result
        if not admin_required:
            return jsonify({"ok": True, "message": "Cassetto non protetto"})
        provided_hash = hashlib.sha256(admin_password.encode()).hexdigest()
        if provided_hash == pwd_hash:
            return jsonify({"ok": True, "message": "Password corretta"})
        else:
            return jsonify({"ok": False, "error": "Password errata"})
    except Exception as e:
        conn.close()
        return jsonify({"error": str(e)}), 400

@app.route('/api/upload-document', methods=['POST'])
def upload_document():
    data = request.get_json()
    filename = data.get('filename', '')
    content = data.get('content', '')
    brand = data.get('brand', '')
    visibility = data.get('visibility', 'public')
    access_code = data.get('access_code', '')
    admin_password = data.get('admin_password', '')
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    try:
        c.execute('SELECT id, admin_required, admin_password FROM aziende WHERE nome = ?', (brand,))
        result = c.fetchone()
        if not result:
            conn.close()
            return jsonify({"error": "Brand non trovato"}), 400
        azienda_id, admin_required, pwd_hash = result
        if admin_required:
            if not admin_password:
                conn.close()
                return jsonify({"error": "Password admin richiesta"}), 403
            provided_hash = hashlib.sha256(admin_password.encode()).hexdigest()
            if provided_hash != pwd_hash:
                conn.close()
                return jsonify({"error": "Password admin errata"}), 403
        c.execute('INSERT INTO documents (filename, content, azienda_id, visibility, access_code, upload_date) VALUES (?, ?, ?, ?, ?, ?)',
                  (filename, content, azienda_id, visibility, access_code, datetime.now().isoformat()))
        conn.commit()
        conn.close()
        return jsonify({"ok": True})
    except Exception as e:
        conn.close()
        return jsonify({"error": str(e)}), 400

@app.route('/api/list-documents', methods=['GET'])
def list_documents():
    brand = request.args.get('brand', '').strip()
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    if brand:
        c.execute('''SELECT d.id, d.filename, d.upload_date, d.visibility, a.nome
                     FROM documents d JOIN aziende a ON d.azienda_id = a.id
                     WHERE LOWER(a.nome) = LOWER(?) ORDER BY d.upload_date DESC''', (brand,))
    else:
        c.execute('''SELECT d.id, d.filename, d.upload_date, d.visibility, a.nome
                     FROM documents d JOIN aziende a ON d.azienda_id = a.id
                     ORDER BY a.nome, d.upload_date DESC LIMIT 100''')
    docs = c.fetchall()
    conn.close()
    return jsonify({"documents": [{"id": d[0], "filename": d[1], "date": d[2], "visibility": d[3], "brand": d[4]} for d in docs]})

@app.route('/api/delete-document/<int:doc_id>', methods=['DELETE'])
def delete_document(doc_id):
    admin_password = request.args.get('admin_password', '')
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    try:
        c.execute('''SELECT a.nome, a.admin_required, a.admin_password
                     FROM documents d JOIN aziende a ON d.azienda_id = a.id WHERE d.id = ?''', (doc_id,))
        result = c.fetchone()
        if not result:
            conn.close()
            return jsonify({"error": "Documento non trovato"}), 404
        brand, admin_required, pwd_hash = result
        if admin_required:
            if not admin_password:
                conn.close()
                return jsonify({"error": "Password admin richiesta"}), 403
            provided_hash = hashlib.sha256(admin_password.encode()).hexdigest()
            if provided_hash != pwd_hash:
                conn.close()
                return jsonify({"error": "Password admin errata"}), 403
        c.execute('DELETE FROM documents WHERE id = ?', (doc_id,))
        conn.commit()
        conn.close()
        return jsonify({"ok": True})
    except Exception as e:
        conn.close()
        return jsonify({"error": str(e)}), 400

@app.route('/api/search-documents', methods=['POST'])
def search_documents():
    data = request.get_json()
    brands = data.get('brands', [])
    question = data.get('question', '')
    access_code = data.get('access_code')
    docs = _search_docs_internal(brands, question, access_code)
    if docs:
        return jsonify({"found": True, "docs": docs})
    return jsonify({"found": False})

def _search_docs_internal(brands, question, access_code):
    if not brands:
        return []
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    placeholders = ','.join('?' * len(brands))
    if access_code:
        query = ("SELECT filename, content FROM documents "
                 "WHERE (azienda_id IN (SELECT id FROM aziende WHERE nome IN (" + placeholders + ")) AND visibility='public') "
                 "OR (visibility='private' AND access_code=?) LIMIT 10")
        c.execute(query, brands + [access_code])
    else:
        query = ("SELECT filename, content FROM documents "
                 "WHERE azienda_id IN (SELECT id FROM aziende WHERE nome IN (" + placeholders + ")) "
                 "AND visibility='public' LIMIT 10")
        c.execute(query, brands)
    docs = c.fetchall()
    conn.close()
    return docs

def search_web(question, brands):
    return None

@app.route('/api/debug-images', methods=['GET'])
def debug_images():
    q = request.args.get('q', 'Gessi rubinetto')
    if not GOOGLE_API_KEY or not GOOGLE_CSE_ID:
        return jsonify({"error": "Chiavi non configurate", "api_key": bool(GOOGLE_API_KEY), "cse_id": bool(GOOGLE_CSE_ID)})
    try:
        resp = httpx.get(
            "https://www.googleapis.com/customsearch/v1",
            params={"key": GOOGLE_API_KEY, "cx": GOOGLE_CSE_ID, "q": q, "searchType": "image", "num": 3},
            timeout=8
        )
        return jsonify({"status": resp.status_code, "body": resp.json()})
    except Exception as e:
        return jsonify({"error": str(e)})

def search_images(query, brands):
    if not GOOGLE_API_KEY or not GOOGLE_CSE_ID:
        return []
    try:
        search_query = " ".join(brands) + " " + query + " product"
        resp = httpx.get(
            "https://www.googleapis.com/customsearch/v1",
            params={
                "key": GOOGLE_API_KEY,
                "cx": GOOGLE_CSE_ID,
                "q": search_query,
                "searchType": "image",
                "num": 6,
                "imgSize": "medium",
                "safe": "active"
            },
            timeout=8
        )
        if resp.status_code == 200:
            data = resp.json()
            items = data.get("items", [])
            return [item["link"] for item in items if "link" in item]
        print("[GOOGLE CSE] Errore: " + str(resp.status_code))
        return []
    except Exception as e:
        print("[GOOGLE CSE ERROR] " + str(e))
        return []

def scarica_immagini_gessi():
    """Estrae URL immagini da Gessi.it e le salva nel DB"""
    try:
        import urllib.request
        import urllib.error
        
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        
        # PARTE 1: PRODOTTI GESSI
        print("[IMMAGINI] === PRODOTTI GESSI ===")
        c.execute("SELECT id, codice, nome FROM products WHERE LOWER(brand)='gessi' AND (image_url IS NULL OR image_url='')")
        prodotti = c.fetchall()
        
        print(f"[IMMAGINI] Trovati {len(prodotti)} prodotti senza immagine")
        
        aggiornati_prodotti = 0
        for pid, codice, nome in prodotti:
            try:
                # Cerca su Gessi.it
                gessi_url = f"https://www.gessi.it/it/cerca?search={codice}"
                req = urllib.request.Request(gessi_url, headers={'User-Agent': 'Mozilla/5.0'})
                
                with urllib.request.urlopen(req, timeout=10) as resp:
                    html = resp.read().decode('utf-8', errors='ignore')
                    
                    # Estrai URL immagini con regex
                    img_urls = re.findall(r'https://[^"\'<>\s]+\.(?:jpg|jpeg|png|webp)', html)
                    
                    if img_urls:
                        # Prendi la prima URL valida (non placeholder)
                        for img_url in img_urls:
                            if 'placeholder' not in img_url.lower() and len(img_url) < 300:
                                # SALVA SOLO L'URL, NON BASE64
                                c.execute("UPDATE products SET image_url=? WHERE id=?", (img_url, pid))
                                aggiornati_prodotti += 1
                                print(f"  ✅ {codice}: {img_url[:80]}")
                                break
                        
            except Exception as e:
                print(f"  ⚠️ {codice}: {str(e)[:50]}")
                continue
        
        # PARTE 2: ACCESSORI
        print("[IMMAGINI] === ACCESSORI GESSI ===")
        c.execute("SELECT id, accessorio_id, accessorio_nome FROM product_accessories WHERE LOWER(brand_accessorio)='gessi' AND (image_url IS NULL OR image_url='')")
        accessori = c.fetchall()
        
        print(f"[IMMAGINI] Trovati {len(accessori)} accessori")
        
        aggiornati_accessori = 0
        for aid, acc_id, acc_nome in accessori:
            try:
                # Cerca su Gessi.it
                gessi_url = f"https://www.gessi.it/it/cerca?search={acc_nome}"
                req = urllib.request.Request(gessi_url, headers={'User-Agent': 'Mozilla/5.0'})
                
                with urllib.request.urlopen(req, timeout=10) as resp:
                    html = resp.read().decode('utf-8', errors='ignore')
                    
                    # Estrai URL immagini
                    img_urls = re.findall(r'https://[^"\'<>\s]+\.(?:jpg|jpeg|png|webp)', html)
                    
                    if img_urls:
                        for img_url in img_urls:
                            if 'placeholder' not in img_url.lower() and len(img_url) < 300:
                                # SALVA SOLO L'URL
                                c.execute("UPDATE product_accessories SET image_url=? WHERE id=?", (img_url, aid))
                                aggiornati_accessori += 1
                                print(f"  ✅ {acc_id}: {img_url[:80]}")
                                break
                        
            except Exception as e:
                print(f"  ⚠️ {acc_id}: {str(e)[:50]}")
                continue
        
        conn.commit()
        conn.close()
        
        return {
            "ok": True, 
            "prodotti_aggiornati": aggiornati_prodotti, 
            "prodotti_totali": len(prodotti),
            "accessori_aggiornati": aggiornati_accessori,
            "accessori_totali": len(accessori)
        }
        
    except Exception as e:
        print(f"[IMMAGINI ERROR] {str(e)}")
        import traceback
        traceback.print_exc()
        return {"ok": False, "error": str(e)}

def openai_ask(prompt):
    if not OPENAI_API_KEY:
        return "Errore: OPENAI_API_KEY non configurata"
    for attempt in range(2):
        try:
            resp = httpx.post(
                "https://api.openai.com/v1/chat/completions",
                json={
                    "model": OPENAI_MODEL,
                    "messages": [{"role": "user", "content": prompt}],
                    "max_tokens": 800
                },
                headers={"Authorization": f"Bearer {OPENAI_API_KEY}"},
                timeout=60
            )
            if resp.status_code == 200:
                data = resp.json()
                if "choices" in data and len(data["choices"]) > 0:
                    return data["choices"][0]["message"]["content"]
            return "Errore API: " + str(resp.status_code)
        except Exception as e:
            print("[OPENAI] Tentativo " + str(attempt + 1) + " fallito: " + str(e))
            if attempt == 1:
                return "OpenAI non risponde. Riprova tra qualche secondo."
    return "Errore sconosciuto"

@app.route('/api/ask', methods=['POST'])
def ask():
    data = request.get_json()
    question = data.get('question', '')
    brands = data.get('brands', [])
    use_web = data.get('web', True)
    access_code = data.get('access_code')
    if not question or not brands:
        return jsonify({"error": "Domanda e brand richiesti"}), 400

    # 1. Cerca prima nel listino Excel dei brand selezionati
    listino_context = ""
    for brand in brands:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute("""SELECT d.content FROM documents d
                     JOIN aziende a ON d.azienda_id = a.id
                     WHERE LOWER(a.nome) = LOWER(?) AND d.filename LIKE '%[EXCEL]%'
                     ORDER BY d.upload_date DESC LIMIT 1""", (brand,))
        row = c.fetchone()
        conn.close()
        if row:
            try:
                import base64 as b64mod, io as iomod
                content = row[0]
                if ',' in content: content = content.split(',',1)[1]
                raw = b64mod.b64decode(content)
                import openpyxl as oxl
                wb = oxl.load_workbook(iomod.BytesIO(raw), data_only=True)
                ws = wb.active
                # Leggi tutte le righe prodotto
                headers = None
                prodotti_trovati = []
                q_lower = question.lower()
                for row_data in ws.iter_rows(values_only=True):
                    row_str = [str(c).lower().strip() if c else '' for c in row_data]
                    if headers is None:
                        if 'codice' in ' '.join(row_str): headers = row_str
                        continue
                    if not any(row_data): continue
                    riga_str = ' '.join([str(v).lower() if v else '' for v in row_data])
                    # Match fuzzy: cerca parole della domanda nella riga
                    parole = [p for p in q_lower.split() if len(p) > 3]
                    if parole and any(p in riga_str for p in parole):
                        vals = [str(v).strip() if v else '—' for v in row_data]
                        if headers:
                            prodotti_trovati.append(dict(zip(headers, vals)))
                        else:
                            prodotti_trovati.append({'riga': ' | '.join(vals[:6])})
                if prodotti_trovati:
                    listino_context += f"\n[LISTINO {brand.upper()} — FONTE EXCEL]\n"
                    for p in prodotti_trovati[:5]:
                        if 'codice' in p:
                            listino_context += f"Codice: {p.get('codice','—')} | Nome: {p.get('nome prodotto', p.get('nome','—'))} | Prezzo cliente: €{p.get('prezzo cliente (€)', p.get('prezzo','—'))} | Prezzo riv.: €{p.get('prezzo rivenditore (€)', p.get('prezzo_rivenditore','—'))} | Disponibilità: {p.get('disponibilità', p.get('disponibilita','—'))} | Desc: {p.get('descrizione breve', p.get('descrizione','—'))}\n"
                        else:
                            listino_context += p.get('riga','') + '\n'
            except Exception as e:
                pass

    # 2. Cerca nei documenti caricati
    doc_context = ""
    docs = _search_docs_internal(brands, question, access_code)
    if docs:
        doc_context = "\n".join(["[DOC: " + d[0] + "] " + d[1][:200] for d in docs])

    # 3. Web solo come fallback
    web_context = ""
    images = []
    if use_web:
        web_result = search_web(question, brands)
        if web_result:
            web_context = "[WEB] " + web_result
        images = search_images(question, brands)

    # Costruisci prompt con priorità: listino Excel > doc > web
    prompt = "Sei un esperto commerciale di arredo bagno per i brand: " + ", ".join(brands) + "\n\nDomanda: " + question

    if listino_context:
        prompt += "\n\n═══ DATI DAL LISTINO UFFICIALE (fonte EXCEL — massima priorità) ═══\n" + listino_context
        prompt += "IMPORTANTE: usa SEMPRE i prezzi e codici dal listino Excel sopra. Non inventare prezzi."
    if doc_context:
        prompt += "\n\nALTRI DOCUMENTI ARCHIVIO:\n" + doc_context
    if web_context and not listino_context:
        prompt += "\n\n[FONTE WEB — usa solo se non hai dati da listino]\n" + web_context

    prompt += "\n\nREGOLE:\n- Prezzi da listino Excel = VERDE (affidabili)\n- Se il prezzo viene dal web, segnalalo come 'prezzo indicativo'\n- NON rimandare mai a siti esterni\n- Se trovi il prodotto nel listino, mostra: codice, nome, prezzo cliente, disponibilità\n- Risposta max 1200 caratteri, professionale"

    answer = openai_ask(prompt)

    # Indica se i dati vengono dal listino o dal web
    fonte = "excel" if listino_context else ("doc" if doc_context else "web")
    return jsonify({"answer": answer, "images": images, "fonte": fonte})

@app.route('/api/cerca-prodotto', methods=['POST'])
def cerca_prodotto():
    """Ricerca fuzzy nel listino Excel — restituisce prodotti matching con fonte"""
    data = request.get_json()
    query = data.get('query', '').lower().strip()
    brand = data.get('brand', '')
    if not query or not brand:
        return jsonify({"prodotti": [], "fonte": None})
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("""SELECT d.content, d.filename FROM documents d
                 JOIN aziende a ON d.azienda_id = a.id
                 WHERE LOWER(a.nome) = LOWER(?) AND d.filename LIKE '%[EXCEL]%'
                 ORDER BY d.upload_date DESC LIMIT 1""", (brand,))
    row = c.fetchone()
    conn.close()
    if not row:
        return jsonify({"prodotti": [], "fonte": None, "messaggio": "Nessun listino Excel caricato per " + brand})
    try:
        import base64 as b64m, io as iom
        content = row[0]
        if ',' in content: content = content.split(',',1)[1]
        raw = b64m.b64decode(content)
        import openpyxl as oxl2
        wb = oxl2.load_workbook(iom.BytesIO(raw), data_only=True)
        ws = wb.active
        SINONIMI_HDR = {
            'codice': ['codice','cod','code','sku'],
            'nome': ['nome prodotto','nome','name','prodotto'],
            'categoria': ['categoria','category'],
            'collezione': ['collezione','collection','linea'],
            'prezzo': ['prezzo cliente (€)','prezzo cliente','prezzo (€)','prezzo'],
            'prezzo_rivenditore': ['prezzo rivenditore (€)','prezzo rivenditore','rivenditore'],
            'disponibilita': ['disponibilità','disponibilita'],
            'descrizione': ['descrizione breve','descrizione'],
            'finiture': ['colori / finiture','finiture','colori'],
        }
        header_row = None
        col_map = {}
        prodotti = []
        parole = [p for p in query.split() if len(p) > 2]
        for row_data in ws.iter_rows(values_only=True):
            if header_row is None:
                row_str = [str(v).lower().strip() if v else '' for v in row_data]
                found = {}
                for campo, syns in SINONIMI_HDR.items():
                    for j, cell in enumerate(row_str):
                        if any(s == cell or s in cell for s in syns):
                            found[campo] = j; break
                if len(found) >= 2:
                    header_row = True
                    col_map = found
                continue
            if not any(row_data): continue
            def gv(campo, rd=row_data):
                idx = col_map.get(campo)
                if idx is None or idx >= len(rd): return ''
                v = rd[idx]; return str(v).strip() if v is not None else ''
            codice = gv('codice')
            if not codice or codice.lower() == 'codice': continue
            riga_str = ' '.join([str(v).lower() if v else '' for v in row_data])
            score = sum(1 for p in parole if p in riga_str)
            if score == 0: continue
            def pp(raw):
                if not raw: return None
                try: return float(re.sub(r'[^\d.,]','',raw).replace(',','.'))
                except: return None
            prodotti.append({
                'score': score,
                'codice': codice,
                'nome': gv('nome'),
                'categoria': gv('categoria'),
                'collezione': gv('collezione'),
                'prezzo': pp(gv('prezzo')),
                'prezzo_rivenditore': pp(gv('prezzo_rivenditore')),
                'disponibilita': gv('disponibilita'),
                'descrizione': gv('descrizione'),
                'finiture': gv('finiture'),
                'fonte': 'excel'
            })
        prodotti.sort(key=lambda x: -x['score'])
        return jsonify({"prodotti": prodotti[:10], "fonte": row[1]})
    except Exception as e:
        return jsonify({"prodotti": [], "fonte": None, "errore": str(e)})

@app.route('/api/listino/<brand>', methods=['GET'])
def get_listino(brand):
    """Restituisce i prodotti dal listino Excel caricato per un brand"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    # Cerca documenti Excel per questo brand (case-insensitive)
    c.execute("""SELECT d.content, d.filename FROM documents d
                 JOIN aziende a ON d.azienda_id = a.id
                 WHERE LOWER(a.nome) = LOWER(?) AND (LOWER(d.filename) LIKE '%.xlsx' OR LOWER(d.filename) LIKE '%.xls' OR LOWER(d.filename) LIKE '%excel%')
                 ORDER BY d.upload_date DESC LIMIT 1""", (brand,))
    row = c.fetchone()
    conn.close()
    if not row:
        return jsonify({"ok": False, "prodotti": [], "fonte": None})
    content_b64, filename = row
    try:
        import base64, io
        if ',' in content_b64:
            content_b64 = content_b64.split(',', 1)[1]
        raw = base64.b64decode(content_b64)
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
        ws = wb.active
        # Trova header row
        header_row = None
        col_map = {}
        SINONIMI = {
            'codice': ['codice','cod','code','sku','art'],
            'nome': ['nome prodotto','nome','name','prodotto','denominazione'],
            'categoria': ['categoria','category','tipo'],
            'collezione': ['collezione','collection','linea'],
            'prezzo': ['prezzo cliente (€)','prezzo cliente','prezzo (€)','prezzo','price','costo','importo','listino'],
            'prezzo_rivenditore': ['prezzo rivenditore (€)','prezzo rivenditore','rivenditore','prezzo riv','costo acquisto'],
            'prezzo_scontato': ['prezzo scontato (€)','prezzo scontato','scontato','offerta'],
            'disponibilita': ['disponibilità','disponibilita','stock','disponibile'],
            'descrizione': ['descrizione breve','descrizione','description','dettaglio'],
            'finiture': ['colori / finiture','finiture','colori','finitura'],
        }
        prodotti = []
        header_row = None
        col_map = {}
        for i, row_data in enumerate(ws.iter_rows(values_only=True)):
            if header_row is None:
                row_str = [str(c).lower().strip() if c else '' for c in row_data]
                found = {}
                for campo, syns in SINONIMI.items():
                    for j, cell in enumerate(row_str):
                        if any(s == cell or s in cell for s in syns):
                            found[campo] = j
                            break
                if len(found) >= 2:
                    header_row = i
                    col_map = found
                continue
            if not any(row_data): continue
            def getv(campo, rd=row_data):
                idx = col_map.get(campo)
                if idx is None or idx >= len(rd): return ''
                v = rd[idx]
                return str(v).strip() if v is not None else ''
            codice = getv('codice')
            if not codice or codice.lower() == 'codice': continue
            def parse_price(raw):
                if not raw: return None
                try: return float(re.sub(r'[^\d.,]','',raw).replace(',','.'))
                except: return None
            
            prod_dict = {
                'codice': codice,
                'nome': getv('nome'),
                'categoria': getv('categoria'),
                'collezione': getv('collezione'),
                'prezzo': parse_price(getv('prezzo')),
                'prezzo_rivenditore': parse_price(getv('prezzo_rivenditore')),
                'prezzo_scontato': parse_price(getv('prezzo_scontato')),
                'disponibilita': getv('disponibilita'),
                'descrizione': getv('descrizione'),
                'finiture': getv('finiture'),
                'fonte': 'excel'
            }
            prodotti.append(prod_dict)
            
            # SALVA SUBITO IN DB products
            try:
                c = conn.cursor()
                c.execute("""INSERT OR REPLACE INTO products 
                            (codice, nome, collezione, categoria, prezzo, prezzo_rivenditore, 
                             prezzo_scontato, disponibilita, descrizione, finiture, fonte, brand, created_at)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                         (codice, prod_dict['nome'], prod_dict['collezione'], prod_dict['categoria'],
                          prod_dict['prezzo'], prod_dict['prezzo_rivenditore'], prod_dict['prezzo_scontato'],
                          prod_dict['disponibilita'], prod_dict['descrizione'], prod_dict['finiture'],
                          'excel', brand, datetime.now().isoformat()))
                conn.commit()
            except Exception as e:
                print(f"[DB] Errore salvataggio prodotto {codice}: {e}")
        
        return jsonify({"ok": True, "prodotti": prodotti, "fonte": filename})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e), "prodotti": []})

# ---------------------------------------------------------------------------
# API EXCEL INTELLIGENTE
# ---------------------------------------------------------------------------

@app.route('/api/parse-excel', methods=['POST'])
def parse_excel():
    """Legge un file Excel base64 e restituisce righe con codice/descrizione/prezzo"""
    if not OPENPYXL_OK:
        return jsonify({"error": "openpyxl non installato sul server"}), 500
    data = request.get_json()
    b64 = data.get('content', '')
    # Rimuovi eventuale data-URI prefix
    if ',' in b64:
        b64 = b64.split(',', 1)[1]
    try:
        raw = base64.b64decode(b64)
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
        ws = wb.active
        rows = []
        header_row = None
        col_map = {}  # nome_campo -> indice colonna (0-based)
        SINONIMI = {
            'codice': ['codice','cod','code','sku','art','articolo','ref'],
            'descrizione': ['descrizione','desc','nome','prodotto','name','denominazione'],
            'prezzo': ['prezzo','price','costo','cost','importo','€','euro','listino']
        }
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if header_row is None:
                # Cerca riga header
                row_str = [str(c).lower().strip() if c else '' for c in row]
                found = {}
                for campo, syns in SINONIMI.items():
                    for j, cell in enumerate(row_str):
                        if any(s in cell for s in syns):
                            found[campo] = j
                            break
                if len(found) >= 2:
                    header_row = i
                    col_map = found
                continue
            if not any(row):
                continue
            def get(campo):
                idx = col_map.get(campo)
                if idx is None or idx >= len(row): return ''
                v = row[idx]
                return str(v).strip() if v is not None else ''
            codice = get('codice')
            descrizione = get('descrizione')
            prezzo_raw = get('prezzo')
            if not codice and not descrizione:
                continue
            # Pulisci prezzo
            prezzo = None
            if prezzo_raw:
                try:
                    prezzo = float(re.sub(r'[^\d.,]', '', prezzo_raw).replace(',', '.'))
                except:
                    pass
            rows.append({'codice': codice, 'descrizione': descrizione, 'prezzo': prezzo, 'prezzo_src': 'excel' if prezzo is not None else None})
            if len(rows) >= 200:
                break
        return jsonify({"ok": True, "righe": rows, "totale": len(rows)})
    except Exception as e:
        return jsonify({"error": str(e)}), 400

@app.route('/api/arricchisci-prodotto', methods=['POST'])
def arricchisci_prodotto():
    """Prende codice+descrizione+prezzo e genera descrizione commerciale sintetica via DeepSeek"""
    data = request.get_json()
    codice = data.get('codice', '')
    descrizione = data.get('descrizione', '')
    prezzo = data.get('prezzo')
    brand = data.get('brand', '')
    prezzo_str = f"€{prezzo}" if prezzo else "non specificato"

    prompt = f"""Sei un esperto commerciale di arredo bagno e pavimentazioni di alto livello.

Devi creare una descrizione commerciale SINTETICA (max 3 righe, circa 150 caratteri) di questo prodotto, da inserire in un'offerta/proposta per un cliente finale.

Brand: {brand if brand else 'non specificato'}
Codice: {codice if codice else 'non specificato'}
Descrizione originale: {descrizione}
Prezzo: {prezzo_str}

La descrizione deve:
- Essere convincente e professionale
- Evidenziare il valore e la qualità
- NON inventare caratteristiche tecniche specifiche non note
- Essere massimo 150 caratteri
- Non includere il prezzo nel testo

Rispondi SOLO con la descrizione commerciale, nient'altro."""

    risposta = openai_ask(prompt)
    return jsonify({"ok": True, "descrizione_ai": risposta.strip()})

@app.route('/api/cantieri/<int:cid>/righe-da-ai', methods=['POST'])
def add_riga_da_ai(cid):
    """Aggiunge una riga al cantiere proveniente dall'arricchimento AI/Excel"""
    u = require_login(['superadmin', 'admin', 'commerciale'])
    if not u:
        return jsonify({"error": "Non autorizzato"}), 403
    data = request.get_json()
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    codice = data.get('codice', '')
    descrizione = data.get('descrizione', '')
    if codice:
        descrizione = f"[{codice}] {descrizione}" if descrizione else codice
    c.execute("INSERT INTO cantiere_righe (cantiere_id, brand, categoria, descrizione, note, importo) VALUES (?,?,?,?,?,?)",
              (cid, data.get('brand',''), data.get('categoria',''), descrizione,
               data.get('note',''), data.get('importo', 0) or 0))
    rid = c.lastrowid
    c.execute("UPDATE cantieri SET data_aggiornamento=? WHERE id=?", (datetime.now().isoformat(), cid))
    conn.commit()
    conn.close()
    return jsonify({"ok": True, "id": rid})

# ============================================================================
# LAZY LOADING ABBINAMENTI PER BRAND
# ============================================================================

def load_accessories_from_excel_lazy(file_content, brand, conn, c):
    """Carica categorie, abbinamenti, vincoli da Excel in memoria"""
    try:
        if ',' in file_content:
            file_content = file_content.split(',', 1)[1]
        
        raw = base64.b64decode(file_content)
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
        
        count_categories = 0
        count_abbinamenti = 0
        count_vincoli = 0
        
        # CARICA CATEGORIE
        if "CATEGORIE_ACCESSORI" in wb.sheetnames:
            ws = wb["CATEGORIE_ACCESSORI"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue
                categoria_id = row[0]
                categoria_nome = row[1]
                descrizione = row[2] if len(row) > 2 else ""
                icona = row[3] if len(row) > 3 else ""
                
                try:
                    c.execute("""
                        INSERT OR REPLACE INTO categories_accessori 
                        (categoria_id, categoria_nome, descrizione, icona, created_at)
                        VALUES (?, ?, ?, ?, ?)
                    """, (categoria_id, categoria_nome, descrizione, icona, datetime.now().isoformat()))
                    count_categories += 1
                except:
                    pass
        
        # CARICA ABBINAMENTI
        if "ABBINAMENTI" in wb.sheetnames:
            ws = wb["ABBINAMENTI"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue
                
                prodotto_padre = row[0]
                accessorio_id = row[1]
                accessorio_nome = row[2] if len(row) > 2 else ""
                brand_accessorio = row[3] if len(row) > 3 else brand
                categoria_accessorio = row[4] if len(row) > 4 else ""
                tipo_relazione = row[5] if len(row) > 5 else "ufficiale"
                priority = row[6] if len(row) > 6 else 99
                
                vincoli_campo = row[7] if len(row) > 7 else None
                vincoli_valore = row[8] if len(row) > 8 else None
                vincoli_severity = row[9] if len(row) > 9 else None
                vincoli_messaggio = row[10] if len(row) > 10 else None
                
                note = row[11] if len(row) > 11 else ""
                
                try:
                    c.execute("""
                        INSERT OR REPLACE INTO product_accessories
                        (prodotto_padre, accessorio_id, accessorio_nome, brand_accessorio,
                         categoria_accessorio, tipo_relazione, priority, note, created_at)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, (prodotto_padre, accessorio_id, accessorio_nome, brand_accessorio,
                          categoria_accessorio, tipo_relazione, int(priority) if priority else 99,
                          note, datetime.now().isoformat()))
                    
                    rel_id = c.lastrowid
                    count_abbinamenti += 1
                    
                    if vincoli_campo:
                        c.execute("""
                            INSERT INTO product_accessories_vincoli
                            (relazione_id, campo_vincolo, valore_vincolo, severity, messaggio, created_at)
                            VALUES (?, ?, ?, ?, ?, ?)
                        """, (rel_id, vincoli_campo, vincoli_valore, vincoli_severity,
                              vincoli_messaggio, datetime.now().isoformat()))
                        count_vincoli += 1
                except:
                    pass
        
        # CARICA REGOLE
        if "REGOLE_MATCHING" in wb.sheetnames:
            ws = wb["REGOLE_MATCHING"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue
                categoria_prodotto = row[0]
                categoria_accessorio = row[1]
                soglia = row[2] if len(row) > 2 else "media"
                nota = row[3] if len(row) > 3 else ""
                
                try:
                    c.execute("""
                        INSERT INTO matching_rules
                        (categoria_prodotto, categoria_accessorio, soglia_compatibilita, nota, created_at)
                        VALUES (?, ?, ?, ?, ?)
                    """, (categoria_prodotto, categoria_accessorio, soglia, nota,
                          datetime.now().isoformat()))
                except:
                    pass
        
        msg = f"✅ {brand}: {count_abbinamenti} abbinamenti caricati"
        total = count_categories + count_abbinamenti + count_vincoli
        return True, msg, total
        
    except Exception as e:
        return False, f"❌ Errore: {str(e)}", 0

@app.route('/api/scarica-immagini/<brand>', methods=['POST'])
def scarica_immagini_brand(brand):
    """Scrapa e salva immagini dei prodotti nel DB"""
    if not is_superadmin():
        return jsonify({"ok": False, "error": "Solo superadmin"}), 403
    
    result = scarica_immagini_gessi()
    return jsonify(result)

@app.route('/api/carica-abbinamenti-excel/<brand>', methods=['POST'])
def carica_abbinamenti_excel(brand):
    """Carica abbinamenti da file Excel nel DB"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    
    try:
        # Cerca file abbinamenti per questo brand nel DB documents
        c.execute("""SELECT content, filename FROM documents d
                     JOIN aziende a ON d.azienda_id = a.id
                     WHERE LOWER(a.nome) = LOWER(?)
                     AND (LOWER(d.filename) LIKE '%abbinamenti%' OR LOWER(d.filename) LIKE '%abbina%')
                     ORDER BY d.upload_date DESC LIMIT 1""", (brand,))
        
        row = c.fetchone()
        if not row:
            conn.close()
            return jsonify({"ok": False, "error": f"File abbinamenti non trovato per {brand}"}), 404
        
        content_b64, filename = row
        
        # Decodifica Excel
        if ',' in content_b64:
            content_b64 = content_b64.split(',', 1)[1]
        
        raw = base64.b64decode(content_b64)
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
        
        # Trova foglio abbinamenti
        sheet_name = None
        for sn in wb.sheetnames:
            if 'abbina' in sn.lower():
                sheet_name = sn
                break
        
        if not sheet_name:
            conn.close()
            return jsonify({"ok": False, "error": f"Foglio 'ABBINAMENTI' non trovato"}), 400
        
        ws = wb[sheet_name]
        
        # Pulisci vecchi abbinamenti per questo brand
        c.execute("DELETE FROM product_accessories WHERE brand_accessorio=?", (brand,))
        conn.commit()
        
        count = 0
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:  # Skip header
                continue
            
            if not row or not row[0]:  # Skip empty
                continue
            
            prodotto = str(row[0]).strip() if row[0] else None
            acc_id = str(row[1]).strip() if row[1] else None
            acc_nome = str(row[2]).strip() if row[2] else ""
            brand_acc = str(row[3]).strip() if row[3] else brand
            categoria = str(row[4]).strip() if row[4] else ""
            tipo = str(row[5]).strip() if row[5] else "ufficiale"
            priority = int(row[6]) if row[6] else 99
            note = str(row[7]).strip() if row[7] else ""
            
            if not prodotto or not acc_id:
                continue
            
            try:
                c.execute("""INSERT INTO product_accessories 
                            (prodotto_padre, accessorio_id, accessorio_nome, brand_accessorio, 
                             categoria_accessorio, tipo_relazione, priority, note, created_at)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                         (prodotto, acc_id, acc_nome, brand_acc, categoria, tipo, priority, note, datetime.now().isoformat()))
                count += 1
            except Exception as e:
                print(f"Errore riga {i}: {e}")
        
        conn.commit()
        conn.close()
        
        if count == 0:
            return jsonify({"ok": False, "error": "Nessun abbinamento trovato nel file"}), 400
        
        return jsonify({
            "ok": True,
            "message": f"✅ Caricati {count} abbinamenti",
            "brand": brand,
            "filename": filename,
            "count": count
        })
    
    except Exception as e:
        conn.close()
        return jsonify({"ok": False, "error": str(e)}), 500

@app.route('/api/load-brand-accessories/<brand>', methods=['GET'])
def load_brand_accessories(brand):
    """Carica SOLO i file Excel del brand selezionato"""
    try:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        
        # Cerca i file Excel del brand (con parola chiave abbinamenti)
        c.execute("""
            SELECT d.content, d.filename 
            FROM documents d
            JOIN aziende a ON d.azienda_id = a.id
            WHERE LOWER(a.nome) = LOWER(?)
            AND (d.filename LIKE '%ABBINAMENTI%' OR d.filename LIKE '%abbinamenti%')
            ORDER BY d.upload_date DESC
            LIMIT 1
        """, (brand,))
        
        row = c.fetchone()
        
        if not row:
            conn.close()
            return jsonify({
                "ok": False, 
                "message": f"Nessun file abbinamenti trovato per {brand}",
                "loaded": 0
            }), 404
        
        file_content, filename = row
        
        # Carica i dati dal file
        success, msg, count = load_accessories_from_excel_lazy(file_content, brand, conn, c)
        
        conn.commit()
        conn.close()
        
        if success:
            return jsonify({
                "ok": True,
                "message": msg,
                "brand": brand,
                "filename": filename,
                "loaded": count,
                "status": "✅ Caricato"
            })
        else:
            return jsonify({
                "ok": False,
                "message": msg,
                "brand": brand
            }), 400
    except Exception as e:
        return jsonify({
            "ok": False,
            "error": str(e),
            "brand": brand
        }), 500

@app.route('/api/abbina/<codice_prodotto>', methods=['GET'])
def get_abbinamenti_prodotto(codice_prodotto):
    """Restituisce accessori ufficiali + alternativi per un codice prodotto"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    
    # Cerca il prodotto (opzionale — gli abbinamenti si vedono comunque)
    c.execute("""SELECT codice, nome, collezione, categoria, prezzo FROM products 
                 WHERE codice = ? LIMIT 1""", (codice_prodotto,))
    prodotto_row = c.fetchone()
    
    prodotto = None
    if prodotto_row:
        prodotto = {
            'codice': prodotto_row[0],
            'nome': prodotto_row[1],
            'collezione': prodotto_row[2],
            'categoria': prodotto_row[3],
            'prezzo': prodotto_row[4]
        }
    
    # Cerca accessori per questo prodotto (SEMPRE, anche se prodotto non in DB)
    c.execute("""SELECT accessorio_id, accessorio_nome, brand_accessorio, tipo_relazione, priority, note
                 FROM product_accessories
                 WHERE prodotto_padre = ?
                 ORDER BY tipo_relazione DESC, priority ASC""", (codice_prodotto,))
    
    rows = c.fetchall()
    conn.close()
    
    ufficiali = []
    alternative = []
    esclusi = []
    
    for row in rows:
        acc = {
            'accessorio_id': row[0],
            'nome': row[1],
            'brand': row[2],
            'tipo': row[3],
            'priority': row[4],
            'note': row[5] if row[5] else ''
        }
        if row[3] == 'ufficiale':
            ufficiali.append(acc)
        elif row[3] == 'alternativa':
            alternative.append(acc)
        elif row[3] == 'escluso':
            esclusi.append(acc)
    
    # Se non ci sono abbinamenti, ritorna 404
    if len(ufficiali) == 0 and len(alternative) == 0:
        return jsonify({"ok": False, "error": "Nessun abbinamento trovato"}), 404
    
    return jsonify({
        "ok": True,
        "prodotto": prodotto,
        "ufficiali": ufficiali,
        "alternative": alternative,
        "esclusi": esclusi,
        "count": len(ufficiali) + len(alternative)
    })

# ---------------------------------------------------------------------------
# CERCA IMMAGINE PRODOTTO — GOOGLE CUSTOM SEARCH
# ---------------------------------------------------------------------------

@app.route('/api/cerca-immagine-prodotto', methods=['POST'])
def cerca_immagine_prodotto():
    """Ricerca immagine per codice prodotto su Google Images + scarica Base64"""
    try:
        data = request.get_json()
        codice = data.get('codice', '').strip()
        nome = data.get('nome', '').strip()
        brand = data.get('brand', '').strip()
        
        if not codice:
            return jsonify({'error': 'Codice prodotto richiesto'}), 400
        
        # 🔑 ESTRAI SOLO LA PARTE PRIMA DEL # (es: 386610#031 → 386610)
        codice_base = codice.split('#')[0].strip()
        
        print(f"🔑 Codice originale: {codice} → Codice base per ricerca: {codice_base}")
        
        # Credenziali Google Custom Search
        API_KEY = os.getenv('GOOGLE_API_KEY')
        CSE_ID = os.getenv('GOOGLE_CSE_ID')
        
        if not API_KEY or not CSE_ID:
            return jsonify({'error': 'Credenziali Google non configurate'}), 500
        
        # Query di ricerca: SOLO codice base + brand (nome troppo generico)
        query = f"{codice_base} {brand}".strip()
        
        print(f"🔍 Ricerca immagini per: {query}")
        
        import httpx
        
        # 1. Chiama Google Custom Search API
        google_url = "https://www.googleapis.com/customsearch/v1"
        google_params = {
            'q': query,
            'cx': CSE_ID,
            'key': API_KEY,
            'searchType': 'image',
            'num': 10  # Cerca 10 per averne almeno 3 scaricabili
        }
        
        print(f"📡 Chiamando Google Custom Search con query: {query}")
        
        response = httpx.get(google_url, params=google_params, timeout=10.0)
        response.raise_for_status()
        
        risultati_google = response.json()
        items = risultati_google.get('items', [])
        
        print(f"📦 Google ha ritornato {len(items)} risultati")
        
        if not items:
            return jsonify({'ok': True, 'risultati': [], 'messaggio': f'Nessuna immagine trovata per {codice_base}'}), 200
        
        # 2. Scarica e converti in Base64 le prime 3 immagini SCARICABILI
        risultati = []
        errori = []
        
        for idx, item in enumerate(items):
            if len(risultati) >= 3:  # Stop quando ne hai 3 buone
                break
                
            try:
                url_immagine = item.get('link', '')
                fonte = item.get('displayLink', 'sconosciuta')
                
                if not url_immagine:
                    print(f"⚠️ Item {idx}: URL vuoto, skip")
                    continue
                
                print(f"🖼️ Tentando di scaricare ({idx+1}): {url_immagine[:60]}... da {fonte}")
                
                # Scarica l'immagine con User-Agent per evitare blocchi
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
                }
                
                try:
                    img_response = httpx.get(
                        url_immagine, 
                        timeout=8.0, 
                        follow_redirects=True,
                        headers=headers,
                        verify=False  # Ignora verifiche SSL problematiche
                    )
                    img_response.raise_for_status()
                except httpx.ReadTimeout:
                    errori.append(f"Timeout su {fonte}")
                    print(f"⏱️ Timeout scaricamento da {fonte}")
                    continue
                except httpx.HTTPError as he:
                    errori.append(f"HTTP error {fonte}: {str(he)}")
                    print(f"❌ HTTP error da {fonte}: {str(he)}")
                    continue
                
                # Controlla content-type
                content_type = img_response.headers.get('content-type', 'image/jpeg')
                if ';' in content_type:
                    content_type = content_type.split(';')[0]
                
                if 'image' not in content_type:
                    errori.append(f"Content-type non immagine: {content_type}")
                    print(f"⚠️ {fonte}: content-type non valido: {content_type}")
                    continue
                
                # Controlla dimensione
                img_data = img_response.content
                if len(img_data) < 1000:  # Meno di 1KB = probabilmente non valida
                    errori.append(f"Immagine troppo piccola da {fonte}")
                    print(f"⚠️ {fonte}: immagine troppo piccola ({len(img_data)} bytes)")
                    continue
                
                if len(img_data) > 5000000:  # Più di 5MB = troppo grande
                    errori.append(f"Immagine troppo grande da {fonte}")
                    print(f"⚠️ {fonte}: immagine troppo grande ({len(img_data)} bytes)")
                    continue
                
                # Converti in Base64
                try:
                    b64_data = base64.b64encode(img_data).decode('utf-8')
                except Exception as be:
                    errori.append(f"Errore encoding Base64 da {fonte}")
                    print(f"❌ Errore encoding Base64 da {fonte}: {str(be)}")
                    continue
                
                # Aggiungi ai risultati SUCCESS
                risultati.append({
                    'url': url_immagine,
                    'fonte': fonte,
                    'b64': b64_data,
                    'content_type': content_type,
                    'size': len(img_data)
                })
                
                print(f"✅ {fonte}: scaricata ({len(img_data)} bytes, Base64: {len(b64_data)} chars)")
                
            except Exception as e:
                errori.append(f"Errore generale item {idx}: {str(e)}")
                print(f"❌ Errore item {idx}: {str(e)}")
                continue
        
        print(f"\n📊 RISULTATI: {len(risultati)} immagini scaricate, {len(errori)} errori")
        
        return jsonify({
            'ok': True,
            'risultati': risultati,
            'query': query,
            'codice_base': codice_base,
            'count': len(risultati),
            'errori': errori,
            'totale_cercate': len(items)
        }), 200
        
    except Exception as e:
        print(f"❌ ERRORE GRAVE cerca_immagine_prodotto: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f"Errore ricerca: {str(e)}"}), 500

# ---------------------------------------------------------------------------
# GET ABBINAMENTI PER PRODOTTO
# ---------------------------------------------------------------------------

@app.route('/api/abbinamenti/<brand>/<codice>', methods=['GET'])
def get_abbinamenti(brand, codice):
    """Ritorna abbinamenti per un prodotto (da Excel o hardcoded)"""
    try:
        # TODO: implementare logica per caricare abbinamenti dal DB o Excel
        # Per ora ritorna lista vuota
        # In futuro: leggere da tabella abbinamenti_prodotto
        
        abbinamenti = [
            {
                'codice': 'ACC001',
                'nome': 'Flessibile scarico',
                'prezzo': 25,
                'immagine_url': None
            },
            {
                'codice': 'ACC002',
                'nome': 'Rubinetteria cromata',
                'prezzo': 45,
                'immagine_url': None
            },
            {
                'codice': 'ACC003',
                'nome': 'Sigillante silicone',
                'prezzo': 8,
                'immagine_url': None
            }
        ]
        
        return jsonify({'ok': True, 'abbinamenti': abbinamenti}), 200
        
    except Exception as e:
        print(f"❌ ERRORE get_abbinamenti: {str(e)}")
        return jsonify({'error': str(e)}), 500

# ---------------------------------------------------------------------------
# FRONTEND
# ---------------------------------------------------------------------------

@app.route('/')
def index():
    return render_template_string(r'''<!DOCTYPE html>
<html>
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Oracolo Covolo</title>
<style>
* { margin: 0; padding: 0; box-sizing: border-box; }
body { font-family: -apple-system; background: linear-gradient(135deg, #0f172e 0%, #1a1f3a 100%); color: #e0e0e0; min-height: 100vh; }
.container { display: flex; height: 100vh; overflow: hidden; }
.sidebar { width: 320px; background: rgba(15,23,46,0.9); border-right: 1px solid rgba(59,130,245,0.2); padding: 16px; overflow-y: auto; flex-shrink: 0; }
.main { flex: 1; display: flex; flex-direction: column; padding: 16px; min-width: 0; min-height: 0; overflow: hidden; }
.rightpanel { width: 280px; background: rgba(15,23,46,0.9); border-left: 1px solid rgba(59,130,245,0.2); padding: 12px; overflow-y: auto; flex-shrink: 0; display: none; }
.rightpanel.visible { display: block; }
h2 { color: #3b82f6; margin-bottom: 10px; font-size: 12px; font-weight: 700; text-transform: uppercase; letter-spacing: 0.05em; }
button { padding: 8px 12px; background: #3b82f6; color: white; border: none; border-radius: 6px; cursor: pointer; font-weight: 600; margin-bottom: 6px; font-size: 11px; }
button:hover { opacity: 0.85; }
.btn-green { background: #10b981; }
.btn-red { background: #ef4444; }
.btn-gray { background: #6b7280; }
.btn-purple { background: #8b5cf6; }
.btn-sm { padding: 4px 8px; font-size: 10px; margin-bottom: 0; }
.dropdown { background: rgba(30,41,59,0.95); border: 1px solid rgba(59,130,245,0.5); border-radius: 6px; padding: 8px; max-height: 220px; overflow-y: auto; display: none; margin-bottom: 8px; }
.dropdown.show { display: block; }
.brand-item { padding: 4px; cursor: pointer; font-size: 11px; }
.brand-item input { margin-right: 4px; }
.badge { display: inline-block; background: #10b981; color: white; padding: 2px 5px; border-radius: 3px; margin: 2px; font-size: 10px; }
.chat-area { flex: 1; background: rgba(15,23,46,0.5); border: 1px solid rgba(59,130,245,0.2); border-radius: 6px; padding: 12px; overflow-y: auto; margin-bottom: 8px; font-size: 12px; min-height: 0; }
.oracolo-msg { position: relative; }
.copy-btn { position: absolute; top: 6px; right: 6px; padding: 3px 8px; font-size: 10px; background: rgba(59,130,245,0.3); color: #93c5fd; border: 1px solid rgba(59,130,245,0.4); border-radius: 4px; cursor: pointer; margin-bottom: 0; }
.copy-btn:hover { background: rgba(59,130,245,0.5); }
.message { background: rgba(59,130,245,0.1); padding: 8px; margin: 4px 0; border-radius: 4px; border-left: 3px solid #3b82f6; }
.message img { max-width: 100%; max-height: 180px; margin-top: 6px; border-radius: 4px; }
.input-area { display: flex; gap: 8px; }
input[type=text], input[type=password], input[type=number], select, textarea { padding: 8px; background: rgba(30,41,59,0.8); border: 1px solid rgba(59,130,245,0.3); color: white; border-radius: 6px; font-size: 11px; }
input[type=text]::placeholder, input[type=password]::placeholder { color: #6b7280; }
.title { color: #3b82f6; font-size: 20px; font-weight: 700; margin-bottom: 12px; }
.btn-3pulsanti { display: flex; gap: 6px; margin-bottom: 10px; }
.btn-3pulsanti button { flex: 1; padding: 7px; font-size: 10px; }
.toggle-btn { width: 100%; padding: 7px; border: none; border-radius: 6px; cursor: pointer; font-weight: 600; margin-bottom: 6px; font-size: 11px; }
.toggle-on { background: #10b981; color: white; }
.toggle-off { background: #6b7280; color: white; }
.module-box { background: rgba(30,41,59,0.6); border: 1px solid rgba(59,130,245,0.2); border-radius: 6px; margin-bottom: 8px; overflow: hidden; }
.module-header { display: flex; align-items: center; justify-content: space-between; padding: 8px 10px; cursor: pointer; }
.module-header:hover { background: rgba(59,130,245,0.1); }
.module-title { font-size: 12px; font-weight: 600; color: #93c5fd; }
.module-body { padding: 8px 10px; border-top: 1px solid rgba(59,130,245,0.2); display: none; }
.module-body.open { display: block; }
.cantiere-item { background: rgba(59,130,245,0.1); border-radius: 4px; padding: 6px 8px; margin: 4px 0; font-size: 11px; cursor: pointer; display: flex; justify-content: space-between; align-items: center; }
.cantiere-item:hover { background: rgba(59,130,245,0.2); }
.stato-badge { padding: 2px 6px; border-radius: 3px; font-size: 9px; font-weight: 700; }
.stato-bozza { background: #6b7280; color: white; }
.stato-inviata { background: #3b82f6; color: white; }
.stato-vinta { background: #10b981; color: white; }
.stato-persa { background: #ef4444; color: white; }
.riga-item { background: rgba(30,41,59,0.8); border-radius: 4px; padding: 5px 8px; margin: 3px 0; font-size: 10px; display: flex; justify-content: space-between; align-items: center; }
.login-overlay { position: fixed; top: 0; left: 0; width: 100%; height: 100%; background: rgba(0,0,0,0.85); z-index: 9999; display: flex; align-items: center; justify-content: center; }
.login-box { background: rgba(15,23,46,0.98); border: 1px solid rgba(59,130,245,0.4); border-radius: 12px; padding: 32px; width: 320px; }
.login-title { color: #3b82f6; font-size: 20px; font-weight: 700; margin-bottom: 20px; text-align: center; }
.sa-panel { background: rgba(139,92,246,0.15); border: 1px solid rgba(139,92,246,0.4); border-radius: 6px; padding: 8px; margin-bottom: 6px; font-size: 11px; }
.bi-stat { background: rgba(30,41,59,0.8); border-radius: 4px; padding: 6px 8px; margin: 3px 0; font-size: 11px; display: flex; justify-content: space-between; }
.brand-autocomplete { position: relative; width: 100%; }
.brand-dropdown-list { position: absolute; top: 100%; left: 0; right: 0; background: #1e293b; border: 1px solid rgba(59,130,245,0.5); border-top: none; border-radius: 0 0 6px 6px; max-height: 180px; overflow-y: auto; z-index: 3000; display: none; }
.brand-dropdown-list.open { display: block; }
.brand-dropdown-item { padding: 7px 10px; font-size: 11px; cursor: pointer; color: #e0e0e0; }
.brand-dropdown-item:hover { background: rgba(59,130,245,0.3); color: white; }
/* DRAWER CANTIERE */
.cantiere-drawer { display: none; position: fixed; top: 0; right: 0; width: 320px; height: 100vh; background: #0f172e; border-left: 2px solid rgba(59,130,245,0.4); z-index: 1000; flex-direction: column; box-shadow: -4px 0 24px rgba(0,0,0,0.5); overflow-y: auto; }
.cantiere-drawer.open { display: flex; }
.drawer-header { background: rgba(59,130,245,0.15); border-bottom: 1px solid rgba(59,130,245,0.3); padding: 14px 18px; display: flex; align-items: center; justify-content: space-between; flex-shrink: 0; }
.drawer-title { font-size: 15px; font-weight: 700; color: #60a5fa; }
.drawer-body { flex: 1; overflow-y: auto; padding: 0; }
.drawer-section { border-bottom: 1px solid rgba(59,130,245,0.15); padding: 10px 12px; }
.drawer-section-title { font-size: 10px; font-weight: 700; color: #6b7280; text-transform: uppercase; letter-spacing: 0.05em; margin-bottom: 8px; }
.riga-card { background: rgba(30,41,59,0.9); border: 1px solid rgba(59,130,245,0.2); border-radius: 6px; padding: 8px 12px; margin: 5px 0; display: flex; align-items: center; justify-content: space-between; }
.riga-card-info { flex: 1; }
.riga-card-brand { font-size: 12px; font-weight: 600; color: #60a5fa; }
.riga-card-cat { font-size: 11px; color: #9ca3af; }
.riga-card-importo { font-size: 13px; font-weight: 700; color: #10b981; margin: 0 12px; white-space: nowrap; }
.totale-bar { background: rgba(16,185,129,0.15); border: 1px solid rgba(16,185,129,0.3); border-radius: 6px; padding: 8px 12px; display: flex; justify-content: space-between; align-items: center; margin-top: 8px; }
.form-row { display: flex; gap: 8px; margin-bottom: 8px; }
.form-row > * { flex: 1; }
.drawer-footer { border-top: 1px solid rgba(59,130,245,0.3); padding: 12px 18px; display: flex; gap: 8px; flex-shrink: 0; background: rgba(15,23,46,0.95); }
/* LISTINO DASHBOARD */
.listino-panel { position: fixed; top: 0; left: 320px; right: 0; bottom: 0; background: rgba(10,15,35,0.97); z-index: 500; display: none; flex-direction: column; }
.listino-panel.open { display: flex; }
.listino-header { background: rgba(59,130,245,0.15); border-bottom: 1px solid rgba(59,130,245,0.3); padding: 12px 18px; display: flex; align-items: center; gap: 12px; flex-shrink: 0; }
.listino-brand-tag { background: rgba(59,130,245,0.3); border: 1px solid rgba(59,130,245,0.5); color: #93c5fd; padding: 4px 12px; border-radius: 20px; font-size: 12px; font-weight: 700; }
.listino-search { flex: 1; background: rgba(30,41,59,0.8); border: 1px solid rgba(59,130,245,0.3); color: white; border-radius: 6px; padding: 8px 12px; font-size: 12px; }
.listino-body { flex: 1; overflow-y: auto; padding: 12px 18px; }
.filtri-bar { display: flex; gap: 6px; margin-bottom: 12px; flex-wrap: wrap; }
.filtro-btn { padding: 4px 10px; font-size: 10px; border-radius: 20px; border: 1px solid rgba(59,130,245,0.3); background: transparent; color: #9ca3af; cursor: pointer; margin-bottom: 0; }
.filtro-btn.active { background: rgba(59,130,245,0.4); color: white; border-color: #3b82f6; }
.domande-bar { display: flex; gap: 6px; flex-wrap: wrap; margin-bottom: 12px; }
.domanda-chip { padding: 5px 10px; font-size: 10px; background: rgba(16,185,129,0.15); border: 1px solid rgba(16,185,129,0.3); color: #6ee7b7; border-radius: 20px; cursor: pointer; margin-bottom: 0; }
.domanda-chip:hover { background: rgba(16,185,129,0.3); }
.prodotti-grid { display: grid; grid-template-columns: repeat(auto-fill, minmax(240px, 1fr)); gap: 8px; }
.prodotto-card { background: rgba(30,41,59,0.9); border: 1px solid rgba(59,130,245,0.2); border-radius: 8px; padding: 12px; cursor: pointer; transition: border-color 0.2s; }
.prodotto-card:hover { border-color: rgba(59,130,245,0.6); }
.prodotto-card.su-ordine { border-left: 3px solid #f59e0b; }
.prodotto-card.disponibile { border-left: 3px solid #10b981; }
.prodotto-codice { font-size: 9px; color: #6b7280; font-family: monospace; margin-bottom: 3px; }
.prodotto-nome { font-size: 12px; font-weight: 600; color: #e0e0e0; margin-bottom: 4px; }
.prodotto-cat { font-size: 10px; color: #9ca3af; margin-bottom: 6px; }
.prodotto-prezzo-excel { color: #10b981; font-weight: 700; font-size: 13px; }
.prodotto-prezzo-web { color: #ef4444; font-weight: 700; font-size: 13px; }
.prodotto-prezzo-sc { color: #f59e0b; font-size: 11px; text-decoration: line-through; margin-left: 4px; }
.prodotto-disp { font-size: 9px; font-weight: 700; text-transform: uppercase; padding: 2px 6px; border-radius: 10px; }
.disp-ok { background: rgba(16,185,129,0.2); color: #10b981; }
.disp-ord { background: rgba(245,158,11,0.2); color: #f59e0b; }
.prodotto-actions { display: flex; gap: 4px; margin-top: 8px; }
.prodotto-actions button { flex: 1; padding: 4px; font-size: 9px; margin-bottom: 0; }
.doc-row { display:flex; align-items:center; gap:10px; padding:8px 10px; border-radius:6px; margin-bottom:4px; background:rgba(30,41,59,0.7); border:1px solid rgba(59,130,245,0.15); font-size:11px; }
.doc-brand-tag { background:rgba(59,130,245,0.2); color:#93c5fd; padding:2px 8px; border-radius:10px; font-size:10px; font-weight:600; white-space:nowrap; }
.doc-tipo-excel { background:rgba(16,185,129,0.2); color:#10b981; padding:2px 6px; border-radius:10px; font-size:9px; font-weight:600; }
.doc-tipo-doc { background:rgba(139,92,246,0.2); color:#a78bfa; padding:2px 6px; border-radius:10px; font-size:9px; font-weight:600; }
/* EXCEL PANEL */
.excel-row { background: rgba(30,41,59,0.9); border: 1px solid rgba(59,130,245,0.15); border-radius: 6px; padding: 8px 10px; margin: 4px 0; font-size: 11px; }
.excel-row-header { display: flex; align-items: center; gap: 8px; }
.excel-codice { color: #9ca3af; font-size: 10px; font-family: monospace; }
.excel-desc { flex: 1; color: #e0e0e0; font-size: 11px; }
.prezzo-excel { color: #10b981; font-weight: 700; font-size: 12px; }
.prezzo-web { color: #ef4444; font-weight: 700; font-size: 12px; }
.excel-ai-desc { color: #93c5fd; font-size: 11px; margin-top: 4px; padding: 4px 6px; background: rgba(59,130,245,0.08); border-radius: 4px; border-left: 2px solid #3b82f6; }
.excel-actions { display: flex; gap: 4px; margin-top: 6px; }
</style>
</head>
<body>

<!-- LOGIN OVERLAY -->
<div class="login-overlay" id="login-overlay">
  <div class="login-box">
    <div class="login-title">Oracolo Covolo</div>
    <div style="margin-bottom: 12px;">
      <input type="text" id="login-user" placeholder="Username" style="width: 100%; margin-bottom: 8px;" onkeypress="if(event.key==='Enter') doLogin()">
      <input type="password" id="login-pwd" placeholder="Password" style="width: 100%;" onkeypress="if(event.key==='Enter') doLogin()">
    </div>
    <button onclick="doLogin()" style="width: 100%; padding: 10px; font-size: 13px;">Accedi</button>
    <div id="login-error" style="color: #ef4444; font-size: 11px; margin-top: 8px; text-align: center;"></div>
    <div style="font-size: 10px; color: #4b5563; margin-top: 16px; text-align: center;">Oracolo Covolo — Tecnaria</div>
  </div>
</div>

<div class="container" id="main-app" style="display: none;">
  <!-- SIDEBAR SX -->
  <div class="sidebar">
    <div style="display: flex; align-items: center; justify-content: space-between; margin-bottom: 12px;">
      <span id="user-label" style="font-size: 11px; color: #60a5fa; font-weight: 600;"></span>
      <button onclick="doLogout()" class="btn-sm btn-gray">Esci</button>
    </div>
    <h2>Seleziona brand</h2>
    <button onclick="toggleDropdown()" style="width: 100%;">Seleziona Brand</button>
    <div id="dropdown" class="dropdown">
      <!-- TAB BAR -->
      <div style="display:flex; gap:4px; margin-bottom:8px;">
        <button id="tab-brand" onclick="switchTab('brand')" style="flex:1; padding:4px; font-size:10px; margin-bottom:0; background:#3b82f6;">Per nome</button>
        <button id="tab-cat" onclick="switchTab('cat')" style="flex:1; padding:4px; font-size:10px; margin-bottom:0; background:rgba(59,130,245,0.3);">Per categoria</button>
      </div>
      <!-- TAB BRAND -->
      <div id="tab-content-brand">
        <input type="text" id="search" placeholder="Ricerca brand..." onkeyup="filterBrands()" style="width: 100%; margin-bottom: 6px;">
        <div id="brands-list"></div>
      </div>
      <!-- TAB CATEGORIA -->
      <div id="tab-content-cat" style="display:none;">
        <input type="text" id="search-cat" placeholder="Es: rubinetteria, piastrelle, doccia..." onkeyup="filterPerCategoria()" style="width: 100%; margin-bottom: 6px;">
        <div id="cat-results"></div>
      </div>
    </div>
    <div id="selected"></div>
    <div id="brand-loading-status" style="font-size:11px; color:#10b981; margin-bottom:8px;"></div>
    <h2>Gruppi salvati</h2>
    <div style="display: flex; gap: 6px; margin-bottom: 8px;">
      <input type="text" id="group-name" placeholder="Nome gruppo..." style="flex: 1;">
      <button onclick="saveGroup()" class="btn-green btn-sm" style="margin-bottom:0;">Salva</button>
    </div>
    <div id="saved-groups" style="max-height: 120px; overflow-y: auto; font-size: 11px;"></div>
    <h2 style="margin-top: 10px;">Protezione cassetto</h2>
    <div style="display: flex; gap: 6px; margin-bottom: 8px;">
      <input type="password" id="admin-pwd" placeholder="Password admin..." style="flex: 1;">
      <button onclick="setAdminPassword()" class="btn-green btn-sm" style="margin-bottom:0;">Proteggi</button>
    </div>
    <h2>Nuovo cassetto</h2>
    <div style="display: flex; gap: 6px; margin-bottom: 8px;">
      <input type="text" id="new-cassetto" placeholder="Nome cassetto..." style="flex: 1;">
      <button onclick="addCassetto()" class="btn-green btn-sm" style="margin-bottom:0;">+</button>
    </div>
    <h2>Accesso privato</h2>
    <input type="password" id="access-code" placeholder="Codice accesso..." style="width: 100%; margin-bottom: 6px;">
    <button onclick="toggleAccess()" style="width: 100%;">Attiva</button>
    <div style="font-size: 10px; color: #9ca3af; margin-top: 4px;" id="access-status">Accesso: PUBBLICO</div>
    <h2 style="margin-top: 10px;">Web search</h2>
    <button id="web-toggle" class="toggle-btn toggle-on" onclick="toggleWeb()">ON</button>
    <h2>Upload documenti</h2>
    <div class="brand-autocomplete" style="margin-bottom:6px;">
      <input type="text" id="upload-brand-input" placeholder="Cerca brand..." autocomplete="off"
        oninput="filterAutocomplete('upload-brand-input','upload-brand-list','upload-brand-val')"
        onfocus="filterAutocomplete('upload-brand-input','upload-brand-list','upload-brand-val')"
        onblur="setTimeout(()=>closeAutocomplete('upload-brand-list'),200)"
        style="width:100%;">
      <input type="hidden" id="upload-brand-val">
      <div class="brand-dropdown-list" id="upload-brand-list"></div>
    </div>
    <label style="display:block; width:100%; background:#8b5cf6; color:white; padding:8px; border-radius:6px; cursor:pointer; font-weight:600; font-size:11px; text-align:center; margin-bottom:6px;">
      Upload Doc <input type="file" id="file-doc" style="display:none" onchange="doUpload(this, 'doc')">
    </label>
    <label style="display:block; width:100%; background:#8b5cf6; color:white; padding:8px; border-radius:6px; cursor:pointer; font-weight:600; font-size:11px; text-align:center; margin-bottom:6px;">
      Upload Excel <input type="file" id="file-excel" accept=".xlsx,.xls,.csv" style="display:none" onchange="doUpload(this, 'excel')">
    </label>
    <div id="upload-status" style="font-size:10px; color:#9ca3af; margin-top:2px;"></div>
    <button onclick="apriGestisciDoc()" style="width:100%; background:#ef4444; margin-top:6px;">Gestisci Documenti</button>
    <button onclick="caricaAbbinamentiEProdotti()" style="width:100%; background:#f59e0b; margin-top:6px; font-weight:600; font-size:11px;">📋 Carica Listino + Abbinamenti</button>
    <div id="abbinamenti-status" style="font-size:10px; color:#9ca3af; margin-top:2px;"></div>
    <button onclick="scaricaImmaginiGessi()" style="width:100%; background:#06b6d4; margin-top:6px; font-weight:600; font-size:11px;">🖼️ Scarica URL Immagini Gessi</button>
  </div>

  <!-- CENTRO -->
  <div class="main">
    <div class="title">Oracolo Covolo</div>
    <div class="btn-3pulsanti" id="btn-3pulsanti">
      <button class="btn-green" onclick="generateOfferta()">OFFERTA</button>
      <button class="btn-green" onclick="generateAnalisi()">ANALISI</button>
      <button class="btn-green" onclick="generateProposta()">PROPOSTA</button>
      <button style="background:#8b5cf6;" onclick="apriListino()">📋 LISTINO</button>
    </div>
    <div class="chat-area" id="chat"></div>
    <div class="input-area">
      <input type="text" id="question" placeholder="Domanda libera o cerca prodotto..." onkeypress="if(event.key==='Enter') ask()" oninput="cercaRapidaListino(this.value)" style="flex: 1;">
      <button onclick="ask()" style="width: 100px;">Invia</button>
    </div>
    <!-- RISULTATI RICERCA RAPIDA LISTINO -->
    <div id="quick-search-results" style="display:none; background:rgba(15,23,46,0.98); border:1px solid rgba(59,130,245,0.3); border-radius:6px; margin-top:4px; max-height:280px; overflow-y:auto; z-index:100;">
    </div>
  </div>

  <!-- PANNELLO LISTINO -->
  <div class="listino-panel" id="listino-panel">
    <div class="listino-header">
      <span id="listino-brand-tag" class="listino-brand-tag">—</span>
      <input type="text" id="listino-search" class="listino-search" placeholder="Cerca prodotto per nome, codice, collezione..." oninput="filtraListino()">
      <!-- SELETTORE LISTINO -->
      <div style="display:flex;gap:4px;flex-shrink:0;">
        <button id="btn-listino-cliente" onclick="setListinoTipo('cliente')" style="padding:5px 10px;font-size:10px;margin-bottom:0;background:#3b82f6;border-radius:4px;">👤 Cliente</button>
        <button id="btn-listino-riv" onclick="setListinoTipo('rivenditore')" style="padding:5px 10px;font-size:10px;margin-bottom:0;background:rgba(245,158,11,0.3);border-radius:4px;color:#f59e0b;">🏪 Rivenditore</button>
      </div>
      <button onclick="chiudiListino()" class="btn-gray btn-sm" style="margin-bottom:0;">✕ Chiudi</button>
    </div>

    <!-- FILTRI LINEA + CATEGORIA + RICERCA -->
    <div style="padding:8px 18px; background:rgba(15,23,46,0.8); border-bottom:1px solid rgba(59,130,245,0.15); flex-shrink:0;">
      <div id="filtri-linea" class="filtri-bar" style="margin-bottom:6px;"></div>
      <div id="filtri-cat" class="filtri-bar" style="margin-bottom:6px;"></div>
      <div style="display:flex; gap:8px; align-items:center;">
        <input type="text" id="listino-search" class="listino-search" style="flex:1;" placeholder="Cerca nome, codice..." oninput="filtraListino()">
        <button id="btn-listino-cliente" onclick="setListinoTipo('cliente')" style="padding:5px 10px;font-size:10px;margin-bottom:0;background:#3b82f6;border-radius:4px;white-space:nowrap;">👤 Cliente</button>
        <button id="btn-listino-riv" onclick="setListinoTipo('rivenditore')" style="padding:5px 10px;font-size:10px;margin-bottom:0;background:rgba(245,158,11,0.2);border-radius:4px;color:#f59e0b;white-space:nowrap;">🏪 Riv.</button>
      </div>
    </div>

    <!-- DOMANDE RAPIDE AI -->
    <div style="padding:6px 18px; border-bottom:1px solid rgba(59,130,245,0.1); flex-shrink:0;">
      <div class="domande-bar" id="domande-bar"></div>
    </div>

    <!-- GRIGLIA PRODOTTI -->
    <div class="listino-body">
      <div id="listino-count" style="font-size:10px; color:#6b7280; margin-bottom:8px;"></div>
      <div class="prodotti-grid" id="prodotti-grid"></div>
    </div>
  </div>

  <!-- PANNELLO GESTIONE DOCUMENTI -->
  <div id="gestisci-doc-panel" style="display:none; position:fixed; top:0; left:0; width:100%; height:100%; background:rgba(0,0,0,0.7); z-index:2000; align-items:center; justify-content:center;">
    <div style="background:#0f172e; border:1px solid rgba(59,130,245,0.4); border-radius:12px; width:700px; max-width:95vw; max-height:85vh; display:flex; flex-direction:column;">
      <!-- Header -->
      <div style="padding:16px 20px; border-bottom:1px solid rgba(59,130,245,0.2); display:flex; align-items:center; justify-content:space-between; flex-shrink:0;">
        <div style="font-size:14px; font-weight:700; color:#60a5fa;">📁 Gestione Documenti</div>
        <button onclick="chiudiGestisciDoc()" class="btn-gray btn-sm" style="margin-bottom:0;">✕ Chiudi</button>
      </div>
      <!-- Filtro brand -->
      <div style="padding:12px 20px; border-bottom:1px solid rgba(59,130,245,0.15); flex-shrink:0; display:flex; gap:8px;">
        <input type="text" id="filtro-doc-brand" placeholder="Filtra per brand (lascia vuoto per tutti)..." style="flex:1; font-size:11px;" oninput="filtraDocumenti()">
        <button onclick="filtraDocumenti()" style="margin-bottom:0; padding:6px 12px; font-size:11px;">🔍 Cerca</button>
        <button onclick="filtraDocumenti(true)" class="btn-gray" style="margin-bottom:0; padding:6px 12px; font-size:11px;">Tutti</button>
      </div>
      <!-- Lista documenti -->
      <div id="doc-list-panel" style="flex:1; overflow-y:auto; padding:12px 20px;"></div>
    </div>
  </div>

  <!-- PANNELLO DX -->
  <div class="rightpanel" id="rightpanel">
    <div style="font-size: 13px; font-weight: 700; color: #93c5fd; margin-bottom: 10px;">Moduli</div>

    <!-- SUPERADMIN PANEL -->
    <div id="sa-panel" style="display:none;">
      <div class="module-box">
        <div class="module-header" onclick="toggleModule('sa-config')">
          <span class="module-title">Config Superadmin</span>
          <span style="font-size:10px; color:#8b5cf6;">SA</span>
        </div>
        <div class="module-body" id="sa-config">
          <div style="margin-bottom: 8px;">
            <select id="sa-cliente-sel" style="width:100%; margin-bottom:6px;" onchange="loadSAModuli()">
              <option value="">-- Seleziona cliente --</option>
            </select>
            <div id="sa-moduli-list" style="font-size:11px;"></div>
          </div>
          <div style="border-top: 1px solid rgba(59,130,245,0.2); padding-top: 8px; margin-top: 4px;">
            <div style="font-size: 11px; font-weight: 600; color: #9ca3af; margin-bottom: 6px;">Nuovo utente</div>
            <input type="text" id="sa-u-nome" placeholder="Nome..." style="width:100%; margin-bottom:4px;">
            <input type="text" id="sa-u-username" placeholder="Username..." style="width:100%; margin-bottom:4px;">
            <input type="password" id="sa-u-pwd" placeholder="Password..." style="width:100%; margin-bottom:4px;">
            <select id="sa-u-ruolo" style="width:100%; margin-bottom:4px;">
              <option value="commerciale">Commerciale</option>
              <option value="admin">Admin cliente</option>
            </select>
            <select id="sa-u-cliente" style="width:100%; margin-bottom:6px;">
              <option value="">-- Cliente --</option>
            </select>
            <button onclick="saAddUtente()" class="btn-green" style="width:100%;">Crea utente</button>
          </div>
          <div style="border-top:1px solid rgba(59,130,245,0.2); padding-top:8px; margin-top:8px;">
            <button onclick="dedupBrands()" class="btn-red" style="width:100%; font-size:10px;">🧹 Unifica brand duplicati</button>
            <div id="dedup-result" style="font-size:10px; color:#10b981; margin-top:4px;"></div>
          </div>
          </div>
        </div>
      </div>
    </div>

    <!-- CANTIERI -->
    <div class="module-box" id="mod-cantieri" style="display:none;">
      <div class="module-header" onclick="toggleModule('cantieri-body')">
        <span class="module-title">Cantieri</span>
        <span id="cantieri-count" style="font-size:10px; color:#9ca3af;">0</span>
      </div>
      <div class="module-body" id="cantieri-body">
        <div style="display: flex; gap: 4px; margin-bottom: 8px;">
          <input type="text" id="new-cantiere" placeholder="Nome cantiere / cliente..." style="flex:1;">
          <button onclick="addCantiere()" class="btn-green btn-sm" style="margin-bottom:0;">+</button>
        </div>
        <div id="cantieri-list"></div>
      </div>
    </div>

    <!-- BI -->
    <div class="module-box" id="mod-bi" style="display:none;">
      <div class="module-header" onclick="toggleModule('bi-body'); loadBI();">
        <span class="module-title">BI / Statistiche</span>
        <span style="font-size:10px; color:#9ca3af;">admin</span>
      </div>
      <div class="module-body" id="bi-body">
        <div id="bi-stats"></div>
        <div style="border-top:1px solid rgba(59,130,245,0.2); padding-top:8px; margin-top:8px;">
          <div style="font-size:11px; font-weight:600; color:#9ca3af; margin-bottom:6px;">Pulizia archivio</div>
          <input type="date" id="bi-da" style="width:100%; margin-bottom:4px;">
          <input type="date" id="bi-a" style="width:100%; margin-bottom:4px;">
          <div style="font-size:10px; color:#9ca3af; margin-bottom:4px;">Stati da eliminare:</div>
          <label style="font-size:10px; display:block;"><input type="checkbox" value="vinta" id="del-vinta"> Vinte</label>
          <label style="font-size:10px; display:block;"><input type="checkbox" value="persa" id="del-persa"> Perse</label>
          <label style="font-size:10px; display:block; margin-bottom:6px;"><input type="checkbox" value="bozza" id="del-bozza"> Bozze</label>
          <button onclick="biCancella()" class="btn-red" style="width:100%; font-size:10px;">Cancella selezionati</button>
        </div>
      </div>
    </div>

    <!-- COMMERCIALI -->
    <div class="module-box" id="mod-commerciali" style="display:none;">
      <div class="module-header" onclick="toggleModule('comm-body')">
        <span class="module-title">Commerciali</span>
        <span style="font-size:10px; color:#9ca3af;">admin</span>
      </div>
      <div class="module-body" id="comm-body">
        <div id="comm-list" style="font-size:11px;"></div>
      </div>
    </div>

  </div>
</div>

<!-- DRAWER DETTAGLIO CANTIERE -->
<div class="cantiere-drawer" id="cantiere-drawer">
  <div class="drawer-header">
    <div>
      <div class="drawer-title" id="drawer-nome"></div>
      <div style="font-size:11px; color:#9ca3af; margin-top:2px;">Gestione offerta</div>
    </div>
    <div style="display:flex; gap:8px; align-items:center;">
      <button id="btn-switch-modalita" onclick="switchModalita()" class="btn-purple btn-sm" style="margin-bottom:0; white-space:nowrap; font-size:10px;">🔄 PIANI</button>
      <select id="cantiere-stato" style="font-size:11px; padding:5px 8px;" onchange="updateCantiere()">
        <option value="bozza">Bozza</option>
        <option value="inviata">Inviata</option>
        <option value="vinta">Vinta</option>
        <option value="persa">Persa</option>
      </select>
      <button onclick="closeCantiere()" class="btn-gray btn-sm" style="margin-bottom:0;">✕ Chiudi</button>
    </div>
  </div>

  <div class="drawer-body">
    <!-- RIGHE ESISTENTI -->
    <div class="drawer-section">
      <div class="drawer-section-title">Elementi nel carrello</div>
      <div id="righe-list"></div>
      <div class="totale-bar" id="totale-bar" style="display:none;">
        <span style="font-size:12px; color:#9ca3af;">Totale offerta</span>
        <span style="font-size:15px; font-weight:700; color:#10b981;" id="totale-valore">€0</span>
      </div>
    </div>

    <!-- ACCESSORI CONSIGLIATI -->
    <div class="drawer-section" id="accessori-section" style="display:block;">
      <div class="drawer-section-title">
        🔗 Accessori Consigliati
      </div>
      <div id="accessori-panel" style="display:block;max-height:300px;overflow-y:auto;">
        <div id="pannello-titolo" style="padding:8px;background:rgba(59,130,246,0.1);border-left:3px solid #3b82f6;margin-bottom:12px;border-radius:4px;"></div>
        <div id="sezioneUfficiali"></div>
        <div id="sezioneAlternative"></div>
      </div>
    </div>

    <!-- AGGIUNGI RIGA -->
    <div class="drawer-section">
      <div class="drawer-section-title">Aggiungi elemento</div>
      <div class="form-row">
        <div class="brand-autocomplete" style="flex:1;">
          <input type="text" id="riga-brand-input" placeholder="Cerca brand..." autocomplete="off"
            oninput="filterAutocomplete('riga-brand-input','riga-brand-list','riga-brand-val')"
            onfocus="filterAutocomplete('riga-brand-input','riga-brand-list','riga-brand-val')"
            onblur="setTimeout(()=>closeAutocomplete('riga-brand-list'),200)"
            onchange="aggiornaCampiExtra()"
            style="width:100%;">
          <input type="hidden" id="riga-brand-val" onchange="aggiornaCampiExtra()">
          <div class="brand-dropdown-list" id="riga-brand-list"></div>
        </div>
        <input type="text" id="riga-categoria" placeholder="Categoria (es. sanitari)" style="flex:1;">
      </div>
      <input type="text" id="riga-descrizione" placeholder="Descrizione prodotto..." style="width:100%; margin-bottom:8px;">

      <!-- CAMPI EXTRA PIASTRELLE -->
      <div id="extra-piastrelle" style="display:none; background:rgba(59,130,245,0.08); border:1px solid rgba(59,130,245,0.2); border-radius:6px; padding:8px; margin-bottom:8px;">
        <div style="font-size:10px; color:#60a5fa; font-weight:700; margin-bottom:6px; text-transform:uppercase;">Dettagli piastrella / rivestimento</div>
        <div class="form-row">
          <input type="text" id="extra-formato" placeholder="Formato (es. 60x120 cm)" style="flex:1;">
          <input type="text" id="extra-finitura" placeholder="Finitura (es. lappato, opaco)" style="flex:1;">
        </div>
        <input type="text" id="extra-colore" placeholder="Colore / tono (es. grigio cemento, effetto marmo)" style="width:100%;">
      </div>

      <!-- CAMPI EXTRA LEGNO / PARQUET -->
      <div id="extra-legno" style="display:none; background:rgba(16,185,129,0.08); border:1px solid rgba(16,185,129,0.2); border-radius:6px; padding:8px; margin-bottom:8px;">
        <div style="font-size:10px; color:#10b981; font-weight:700; margin-bottom:6px; text-transform:uppercase;">Dettagli legno / parquet</div>
        <div class="form-row">
          <input type="text" id="extra-essenza" placeholder="Essenza (es. rovere, noce, frassino)" style="flex:1;">
          <input type="text" id="extra-legno-finitura" placeholder="Finitura (es. oliato, laccato)" style="flex:1;">
        </div>
        <div class="form-row">
          <input type="text" id="extra-legno-formato" placeholder="Formato doga (es. 190x1900 mm)" style="flex:1;">
          <input type="text" id="extra-legno-tono" placeholder="Tono (es. sbiancato, miele, wengè)" style="flex:1;">
        </div>
      </div>

      <!-- CAMPI EXTRA VINILICO/TECNICO -->
      <div id="extra-vinilico" style="display:none; background:rgba(139,92,246,0.08); border:1px solid rgba(139,92,246,0.2); border-radius:6px; padding:8px; margin-bottom:8px;">
        <div style="font-size:10px; color:#8b5cf6; font-weight:700; margin-bottom:6px; text-transform:uppercase;">Dettagli pavimento tecnico / vinilico</div>
        <div class="form-row">
          <input type="text" id="extra-vin-formato" placeholder="Formato" style="flex:1;">
          <input type="text" id="extra-vin-spessore" placeholder="Spessore (es. 5mm)" style="flex:1;">
        </div>
        <input type="text" id="extra-vin-effetto" placeholder="Effetto (es. legno, pietra, cemento)" style="width:100%;">
      </div>

      <div class="form-row">
        <input type="number" id="riga-importo" placeholder="Importo €" style="flex:1;">
        <button onclick="addRiga()" class="btn-green" style="flex:1; margin-bottom:0;">+ Aggiungi</button>
      </div>
    </div>
    <!-- AGGIUNGI MANUALE / VOCE LIBERA -->
    <div class="drawer-section">
      <div class="drawer-section-title" style="cursor:pointer;" onclick="toggleExcelPanel()">
        ⚡ Importa da Excel / Voce libera
        <span id="excel-panel-arrow" style="float:right; color:#9ca3af;">▼</span>
      </div>
      <div id="excel-panel" style="display:none;">
        <!-- Upload Excel -->
        <div style="margin-bottom:8px;">
          <label style="display:block; width:100%; background:#8b5cf6; color:white; padding:7px; border-radius:6px; cursor:pointer; font-weight:600; font-size:11px; text-align:center; margin-bottom:6px;">
            📊 Carica Excel prodotti
            <input type="file" id="excel-listino" accept=".xlsx,.xls" style="display:none" onchange="caricaExcelListino(this)">
          </label>
          <div id="excel-status" style="font-size:10px; color:#9ca3af; margin-bottom:6px;"></div>
        </div>

        <!-- Lista righe Excel -->
        <div id="excel-righe-list" style="max-height:340px; overflow-y:auto;"></div>

        <!-- Separatore -->
        <div style="border-top:1px solid rgba(59,130,245,0.15); margin: 10px 0 8px 0;"></div>

        <!-- Voce manuale libera -->
        <div style="font-size:10px; color:#9ca3af; font-weight:700; text-transform:uppercase; letter-spacing:0.06em; margin-bottom:6px;">Voce manuale (trasporto, manodopera, ecc.)</div>
        <input type="text" id="voce-desc" placeholder="Descrizione voce..." style="width:100%; margin-bottom:6px;">
        <div class="form-row">
          <input type="number" id="voce-importo" placeholder="Importo €" style="flex:1;">
          <button onclick="addVoceManuale()" class="btn-purple" style="flex:1; margin-bottom:0; font-size:11px;">+ Aggiungi</button>
        </div>
      </div>
    </div>
  </div>

  <div class="drawer-footer">
    <button onclick="generaOffertaCantiere()" class="btn-green" style="flex:2; font-size:12px; margin-bottom:0; padding:10px;">Genera Offerta AI</button>
    <button onclick="deleteCantiere()" class="btn-red" style="flex:1; font-size:11px; margin-bottom:0;">Elimina cantiere</button>
  </div>
</div>

<!-- DRAWER PIANI/STANZE/VOCI (NASCOSTO FINCHE NON CLICCHI SWITCH) -->
<div class="cantiere-drawer" id="cantiere-drawer-piani" style="display:none;">
  <div class="drawer-header">
    <div>
      <div class="drawer-title" id="drawer-piani-nome"></div>
      <div style="font-size:11px; color:#9ca3af; margin-top:2px;">Modalità Piani/Stanze</div>
    </div>
    <div style="display:flex; gap:8px; align-items:center;">
      <button id="btn-switch-indietro" onclick="switchModalita()" class="btn-purple btn-sm" style="margin-bottom:0; white-space:nowrap; font-size:10px;">🔄 SEMPLICE</button>
      <button onclick="closeCantiere()" class="btn-gray btn-sm" style="margin-bottom:0;">✕ Chiudi</button>
    </div>
  </div>

  <div class="drawer-body" style="display:flex; gap:8px;">
    <!-- SINISTRA: PIANI/STANZE -->
    <div id="pannello-piani" style="flex:1; overflow-y:auto; padding:12px; border-right:1px solid rgba(59,130,245,0.2);">
      <div style="padding:12px; background:rgba(59,130,245,0.1); border-radius:6px; color:#93c5fd; font-size:11px;">
        ⏳ Caricamento struttura piani...
      </div>
    </div>
    
    <!-- DESTRA: LISTINO DINAMICO + CARICAMENTO PLANIMETRIA -->
    <div style="flex:1; display:flex; flex-direction:column; gap:8px;">
      <!-- CARICAMENTO PLANIMETRIA -->
      <div style="background:rgba(139,92,246,0.15); border:1px solid rgba(139,92,246,0.3); border-radius:6px; padding:8px;">
        <div style="font-size:10px; color:#8b5cf6; font-weight:700; text-transform:uppercase; margin-bottom:6px;">📁 Carica Planimetria</div>
        <label style="display:block; width:100%; background:#8b5cf6; color:white; padding:6px; border-radius:4px; cursor:pointer; font-weight:600; font-size:10px; text-align:center; margin-bottom:4px;">
          Carica disegno
          <input type="file" id="file-planimetria" accept=".png,.jpg,.jpeg,.pdf" style="display:none" onchange="caricaPlanimetria(this)">
        </label>
        <div id="planimetria-status" style="font-size:9px; color:#9ca3af;"></div>
      </div>

      <!-- LISTINO DINAMICO -->
      <div style="background:rgba(59,130,245,0.1); border:1px solid rgba(59,130,245,0.2); border-radius:6px; padding:8px; flex:1; display:flex; flex-direction:column; overflow:hidden;">
        <div style="font-size:10px; color:#60a5fa; font-weight:700; text-transform:uppercase; margin-bottom:6px;">📋 Listino Rapido</div>
        <input type="text" id="listino-piani-search" placeholder="Cerca prodotto..." oninput="filtraListinoPiani()" style="width:100%; margin-bottom:6px; font-size:10px;">
        <div id="listino-piani-grid" style="flex:1; overflow-y:auto; display:grid; grid-template-columns:1fr; gap:4px;">
          <div style="color:#6b7280; font-size:10px;">Seleziona una stanza per caricare il listino</div>
        </div>
      </div>
    </div>
  </div>

  <div class="drawer-footer">
    <button onclick="aggiungiPianoModal()" class="btn-green" style="flex:1; font-size:11px; margin-bottom:0;">➕ Piano</button>
    <button onclick="generaOffertaPiani()" class="btn-green" style="flex:2; font-size:12px; margin-bottom:0; padding:10px;">Genera Offerta AI</button>
    <button onclick="deleteCantiere()" class="btn-red" style="flex:1; font-size:11px; margin-bottom:0;">✕ Elimina</button>
  </div>
</div>

<script>
let BRANDS = [];
let selected = [];
let webEnabled = true;
let accessCode = null;
let accessLevel = "public";
let groups = JSON.parse(localStorage.getItem('oracolo_groups')) || {};
let currentUser = null;
let moduliAttivi = [];
let cantiereAttivo = null;

// ---------------------------------------------------------------------------
// AUTOCOMPLETE BRAND
// ---------------------------------------------------------------------------
function filterAutocomplete(inputId, listId, valId) {
  const input = document.getElementById(inputId);
  const list = document.getElementById(listId);
  const val = input.value.toLowerCase();
  const filtered = val.length === 0 ? BRANDS : BRANDS.filter(b => b.toLowerCase().includes(val));
  list.innerHTML = filtered.map(b =>
    '<div class="brand-dropdown-item" onmousedown="selectBrand(\'' + inputId + '\',\'' + listId + '\',\'' + valId + '\',\'' + b.replace(/'/g, "\\'") + '\')">' + b + '</div>'
  ).join('');
  list.classList.add('open');
}

function selectBrand(inputId, listId, valId, brand) {
  document.getElementById(inputId).value = brand;
  document.getElementById(valId).value = brand;
  closeAutocomplete(listId);
  if (inputId === 'riga-brand-input') aggiornaCampiExtra();
  
  // LAZY LOADING ABBINAMENTI E LISTINO quando seleziona il brand principale
  if (inputId === 'brand-input') {
    console.log('Caricando accessori e listino per:', brand);
    const statusEl = document.getElementById('brand-loading-status');
    if (statusEl) statusEl.textContent = '⏳ Caricamento ' + brand + '...';
    
    // 1. Carica abbinamenti
    fetch('/api/load-brand-accessories/' + encodeURIComponent(brand))
      .then(r => r.json())
      .then(data => {
        if (data.ok) {
          console.log('✅ Abbinamenti:', data.message);
        } else {
          console.error('❌ Abbinamenti:', data.message);
        }
      })
      .catch(e => console.error('Errore abbinamenti:', e));
    
    // 2. Carica listino
    caricaListinoBrand(brand);
  }
}

function caricaListinoBrand(brand) {
  console.log('Caricando listino per:', brand);
  
  fetch('/api/listino/' + encodeURIComponent(brand))
    .then(r => r.json())
    .then(data => {
      if (data.ok && data.prodotti) {
        console.log('✅ Listino:', data.prodotti.length, 'prodotti');
        
        window.listinoData = data.prodotti || [];
        window.listinoBrand = brand;
        window.listinoTipo = data.tipo || 'cliente';
        
        const statusEl = document.getElementById('brand-loading-status');
        if (statusEl) {
          statusEl.textContent = '✅ ' + brand + ': ' + data.prodotti.length + ' prodotti caricati';
        }
        
        if (window.renderListino) {
          renderListino(window.listinoData);
        }
      } else {
        console.error('❌ Listino non trovato');
      }
    })
    .catch(e => {
      console.error('Errore listino:', e);
    });
}

function closeAutocomplete(listId) {
  const list = document.getElementById(listId);
  if (list) list.classList.remove('open');
}

// ---------------------------------------------------------------------------
// LOGIN
// ---------------------------------------------------------------------------
function doLogin() {
  const username = document.getElementById('login-user').value.trim();
  const password = document.getElementById('login-pwd').value.trim();
  if (!username || !password) { document.getElementById('login-error').textContent = 'Inserisci username e password'; return; }
  fetch('/api/login', { method:'POST', headers:{'Content-Type':'application/json'}, body: JSON.stringify({username, password}) })
    .then(r => r.json())
    .then(d => {
      if (d.ok) {
        // Salva il ruolo nel localStorage
        localStorage.setItem('user_ruolo', d.ruolo || 'commerciale');
        localStorage.setItem('user_nome', d.nome || 'Utente');
        document.getElementById('login-overlay').style.display = 'none';
        document.getElementById('main-app').style.display = 'flex';
        initApp();
      } else {
        document.getElementById('login-error').textContent = d.error || 'Errore login';
      }
    });
}

function doLogout() {
  fetch('/api/logout', { method:'POST' }).then(() => {
    document.getElementById('login-overlay').style.display = 'flex';
    document.getElementById('main-app').style.display = 'none';
    document.getElementById('login-user').value = '';
    document.getElementById('login-pwd').value = '';
    document.getElementById('login-error').textContent = '';
    currentUser = null;
    moduliAttivi = [];
    localStorage.removeItem('user_ruolo');
    localStorage.removeItem('user_nome');
  });
}

function initApp() {
  fetch('/api/me').then(r => r.json()).then(d => {
    if (!d.logged) { doLogout(); return; }
    currentUser = d.user;
    
    // Se per qualche motivo il ruolo non è corretto, usa il localStorage
    if (!currentUser.ruolo) {
      currentUser.ruolo = localStorage.getItem('user_ruolo') || 'commerciale';
    }
    if (!currentUser.nome) {
      currentUser.nome = localStorage.getItem('user_nome') || 'Utente';
    }
    
    moduliAttivi = d.moduli || [];
    document.getElementById('user-label').textContent = currentUser.nome + ' (' + currentUser.ruolo + ')';
    setupRightPanel();
    loadBrands();
  });
}

// ---------------------------------------------------------------------------
// SETUP PANNELLO DX
// ---------------------------------------------------------------------------
function setupRightPanel() {
  const rp = document.getElementById('rightpanel');
  const hasMods = moduliAttivi.length > 0 || currentUser.ruolo === 'superadmin';
  if (hasMods) rp.classList.add('visible');

  if (currentUser.ruolo === 'superadmin') {
    document.getElementById('sa-panel').style.display = 'block';
    loadSAClienti();
    document.getElementById('mod-cantieri').style.display = 'block';
    document.getElementById('mod-bi').style.display = 'block';
    document.getElementById('mod-commerciali').style.display = 'block';
  } else {
    if (moduliAttivi.includes('cantieri')) document.getElementById('mod-cantieri').style.display = 'block';
    if (moduliAttivi.includes('bi') && (currentUser.ruolo === 'admin')) document.getElementById('mod-bi').style.display = 'block';
    if (moduliAttivi.includes('commerciali') && (currentUser.ruolo === 'admin')) document.getElementById('mod-commerciali').style.display = 'block';
  }
  loadCantieri();
}

function toggleModule(id) {
  const el = document.getElementById(id);
  if (el) el.classList.toggle('open');
}

// ---------------------------------------------------------------------------
// SUPERADMIN
// ---------------------------------------------------------------------------
function loadSAClienti() {
  fetch('/api/sa/clienti').then(r => r.json()).then(d => {
    const sel1 = document.getElementById('sa-cliente-sel');
    const sel2 = document.getElementById('sa-u-cliente');
    (d.clienti || []).forEach(cl => {
      const o1 = document.createElement('option'); o1.value = cl.id; o1.textContent = cl.nome; sel1.appendChild(o1);
      const o2 = document.createElement('option'); o2.value = cl.id; o2.textContent = cl.nome; sel2.appendChild(o2);
    });
  });
}

function loadSAModuli() {
  const cid = document.getElementById('sa-cliente-sel').value;
  if (!cid) return;
  fetch('/api/sa/moduli/' + cid).then(r => r.json()).then(d => {
    const moduli = d.moduli || {};
    const names = {'cantieri':'Cantieri','carrello':'Carrello','bi':'BI / Statistiche','commerciali':'Commerciali'};
    let html = '';
    Object.keys(names).forEach(k => {
      const on = moduli[k];
      html += '<div style="display:flex;align-items:center;justify-content:space-between;padding:4px 0;">';
      html += '<span>' + names[k] + '</span>';
      html += '<button onclick="toggleModulo(' + cid + ',\'' + k + '\',' + (on?'false':'true') + ')" class="btn-sm ' + (on?'btn-green':'btn-gray') + '" style="margin-bottom:0;">' + (on?'ON':'OFF') + '</button>';
      html += '</div>';
    });
    document.getElementById('sa-moduli-list').innerHTML = html;
  });
}

function toggleModulo(cid, modulo, attivo) {
  fetch('/api/sa/moduli/' + cid, { method:'POST', headers:{'Content-Type':'application/json'}, body: JSON.stringify({modulo, attivo}) })
    .then(r => r.json()).then(d => { if (d.ok) loadSAModuli(); });
}

function dedupBrands() {
  if (!confirm('Unificare tutti i brand duplicati (es. Gessi + GESSI → Gessi)? I documenti verranno spostati al brand corretto.')) return;
  fetch('/api/sa/dedup-brands', { method:'POST' })
    .then(r => r.json())
    .then(d => {
      if (d.ok) {
        document.getElementById('dedup-result').textContent = '✓ ' + d.merged + ' duplicati rimossi. Ricarica la pagina.';
        setTimeout(() => location.reload(), 2000);
      }
    });
}

function saAddUtente() {
  const nome = document.getElementById('sa-u-nome').value.trim();
  const username = document.getElementById('sa-u-username').value.trim();
  const password = document.getElementById('sa-u-pwd').value.trim();
  const ruolo = document.getElementById('sa-u-ruolo').value;
  const cliente_id = document.getElementById('sa-u-cliente').value || null;
  if (!nome || !username || !password) { alert('Dati incompleti'); return; }
  fetch('/api/sa/utenti', { method:'POST', headers:{'Content-Type':'application/json'}, body: JSON.stringify({nome, username, password, ruolo, cliente_id}) })
    .then(r => r.json()).then(d => {
      if (d.ok) { alert('Utente creato!'); document.getElementById('sa-u-nome').value=''; document.getElementById('sa-u-username').value=''; document.getElementById('sa-u-pwd').value=''; }
      else alert('Errore: ' + d.error);
    });
}

// ---------------------------------------------------------------------------
// ACCESSORI CONSIGLIATI
// ---------------------------------------------------------------------------
function toggleAccessoriSection() {
  const panel = document.getElementById('accessori-panel');
  const arrow = document.getElementById('accessori-arrow');
  const open = panel.style.display !== 'none';
  panel.style.display = open ? 'none' : 'block';
  arrow.textContent = open ? '▼' : '▲';
}

function caricaAccessoriProdotto(prodottoId, brand) {
  if (!prodottoId) return;
  
  fetch('/api/abbina/' + encodeURIComponent(prodottoId))
    .then(r => r.json())
    .then(d => {
      if (d.ufficiali && (d.ufficiali.length > 0 || d.alternative.length > 0 || d.esclusi.length > 0)) {
        document.getElementById('accessori-section').style.display = 'block';
        renderAccessoriHtml(d.ufficiali, d.alternative, d.esclusi);
      }
    })
    .catch(e => console.error('Errore accessori:', e));
}

function renderAccessoriHtml(ufficiali, alternative, esclusi) {
  let html = '';

  if (ufficiali && ufficiali.length > 0) {
    html += '<div style="margin-bottom:12px;">' +
      '<div style="font-size:11px; font-weight:700; color:#10b981; margin-bottom:6px;">✅ Abbinamenti Ufficiali</div>';
    ufficiali.forEach(acc => {
      html += '<div style="background:rgba(16,185,129,0.1); border:1px solid rgba(16,185,129,0.3); border-radius:6px; padding:8px; margin-bottom:4px; display:flex; align-items:center; justify-content:space-between;">' +
        '<div style="flex:1;">' +
        '<div style="font-size:11px; font-weight:600; color:#e0e0e0;">' + (acc.nome || acc.id) + '</div>' +
        '<div style="font-size:10px; color:#9ca3af;">' + (acc.id || '') + ' · ' + (acc.brand || '') + '</div>' +
        '</div>' +
        '<button onclick="aggiungiAccessorioAlCantiere(\'' + (acc.id||'').replace(/'/g,"\\'") + '\',\'' + (acc.nome||'').replace(/'/g,"\\'") + '\',\'' + (acc.brand||'').replace(/'/g,"\\'") + '\')" class="btn-sm btn-green" style="margin-bottom:0; white-space:nowrap;">✓ Aggiungi</button>' +
        '</div>';
    });
    html += '</div>';
  }

  if (alternative && alternative.length > 0) {
    html += '<div style="margin-bottom:12px;">' +
      '<div style="font-size:11px; font-weight:700; color:#f59e0b; margin-bottom:6px;">🔹 Alternative</div>';
    alternative.forEach(acc => {
      html += '<div style="background:rgba(245,158,11,0.1); border:1px solid rgba(245,158,11,0.3); border-radius:6px; padding:8px; margin-bottom:4px; display:flex; align-items:center; justify-content:space-between;">' +
        '<div style="flex:1;">' +
        '<div style="font-size:11px; font-weight:600; color:#e0e0e0;">' + (acc.nome || acc.id) + '</div>' +
        '<div style="font-size:10px; color:#9ca3af;">' + (acc.id || '') + ' · ' + (acc.brand || '') + '</div>' +
        '</div>' +
        '<button onclick="aggiungiAccessorioAlCantiere(\'' + (acc.id||'').replace(/'/g,"\\'") + '\',\'' + (acc.nome||'').replace(/'/g,"\\'") + '\',\'' + (acc.brand||'').replace(/'/g,"\\'") + '\')" class="btn-sm" style="background:rgba(245,158,11,0.2); color:#f59e0b; margin-bottom:0; white-space:nowrap;">+ Aggiungi</button>' +
        '</div>';
    });
    html += '</div>';
  }

  if (esclusi && esclusi.length > 0) {
    html += '<div style="margin-bottom:12px;">' +
      '<div style="font-size:11px; font-weight:700; color:#ef4444; margin-bottom:6px;">❌ Non Compatibili</div>';
    esclusi.forEach(acc => {
      html += '<div style="background:rgba(239,68,68,0.1); border:1px solid rgba(239,68,68,0.3); border-radius:6px; padding:8px; margin-bottom:4px; opacity:0.6;">' +
        '<div style="font-size:11px; font-weight:600; color:#e0e0e0;">' + (acc.nome || acc.id) + '</div>' +
        '<div style="font-size:10px; color:#9ca3af;">' + (acc.id || '') + ' · ' + (acc.brand || '') + '</div>' +
        '<div style="font-size:9px; color:#fca5a5; margin-top:3px;">⚠️ Non compatibile con il prodotto principale</div>' +
        '</div>';
    });
    html += '</div>';
  }

  document.getElementById('accessori-ufficiali').innerHTML = html;
}

// CARICAMENTO ACCESSORI - vedi mostraModalAbbinamenti() sotto

// ---------------------------------------------------------------------------
// CANTIERI (fine accessori)
function loadCantieri() {
  if (!document.getElementById('mod-cantieri') || document.getElementById('mod-cantieri').style.display === 'none') return;
  fetch('/api/cantieri').then(r => r.json()).then(d => {
    const list = d.cantieri || [];
    document.getElementById('cantieri-count').textContent = list.length;
    document.getElementById('cantieri-list').innerHTML = list.map(ca =>
      '<div class="cantiere-item" onclick="openCantiere(' + ca.id + ',\'' + ca.nome.replace(/'/g,'') + '\',\'' + ca.stato + '\')">' +
      '<span>' + ca.nome + '</span>' +
      '<span class="stato-badge stato-' + ca.stato + '">' + ca.stato + '</span>' +
      '</div>'
    ).join('');
  });
}

function addCantiere() {
  const nome = document.getElementById('new-cantiere').value.trim();
  if (!nome) { alert('Nome richiesto'); return; }
  fetch('/api/cantieri', { method:'POST', headers:{'Content-Type':'application/json'}, body: JSON.stringify({nome}) })
    .then(r => r.json()).then(d => {
      if (d.ok) { document.getElementById('new-cantiere').value = ''; loadCantieri(); openCantiere(d.id, nome, 'bozza'); }
      else alert('Errore: ' + d.error);
    });
}

function openCantiere(id, nome, stato) {
  cantiereAttivo = id;
  
  // Fetch modalita PRIMA di aprire il drawer
  fetch('/api/cantieri/' + id + '/modalita')
    .then(r => r.json())
    .then(d => {
      const modalita = d.modalita || 'semplice';
      
      // Chiudi entrambi i drawer prima
      document.getElementById('cantiere-drawer').classList.remove('open');
      document.getElementById('cantiere-drawer-piani').classList.remove('open');
      
      setTimeout(() => {
        if (modalita === 'semplice') {
          // DRAWER SEMPLICE
          document.getElementById('drawer-nome').textContent = nome;
          document.getElementById('cantiere-stato').value = stato;
          document.getElementById('btn-switch-modalita').textContent = '🔄 PIANI';
          document.getElementById('cantiere-drawer').style.display = 'flex';
          document.getElementById('cantiere-drawer').classList.add('open');
          precompilaBrandDrawer();
          loadRighe();
        } else {
          // DRAWER PIANI
          document.getElementById('drawer-piani-nome').textContent = nome;
          document.getElementById('btn-switch-indietro').textContent = '🔄 SEMPLICE';
          document.getElementById('cantiere-drawer-piani').style.display = 'flex';
          document.getElementById('cantiere-drawer-piani').classList.add('open');
          loadInterfacciaPiani(id);
        }
      }, 50);
    });
}

function switchModalita() {
  switchInterfaccia();
}

function switchInterfaccia(forzaNuovaModalita) {
  if (!cantiereAttivo) return;
  
  fetch('/api/cantieri/' + cantiereAttivo + '/modalita')
    .then(r => r.json())
    .then(d => {
      const modalitaAttuale = d.modalita || 'semplice';
      const nuova = forzaNuovaModalita || (modalitaAttuale === 'semplice' ? 'piani' : 'semplice');
      
      if (modalitaAttuale === nuova) return;  // Già in quella modalita
      
      fetch('/api/cantieri/' + cantiereAttivo + '/modalita', {
        method: 'PUT',
        headers: {'Content-Type': 'application/json'},
        body: JSON.stringify({ modalita: nuova })
      })
      .then(r => r.json())
      .then(d => {
        if (d.ok) {
          // Chiudi il drawer attuale
          document.getElementById('cantiere-drawer').classList.remove('open');
          document.getElementById('cantiere-drawer-piani').classList.remove('open');
          
          // Mostra il nuovo drawer dopo transizione
          setTimeout(() => {
            if (nuova === 'piani') {
              // Apri drawer PIANI
              document.getElementById('cantiere-drawer-piani').style.display = 'flex';
              document.getElementById('cantiere-drawer-piani').classList.add('open');
              loadInterfacciaPiani(cantiereAttivo);
            } else {
              // Apri drawer SEMPLICE
              document.getElementById('cantiere-drawer').style.display = 'flex';
              document.getElementById('cantiere-drawer').classList.add('open');
              loadRighe();
            }
          }, 300);
        } else {
          alert('❌ Errore: ' + d.error);
        }
      });
    });
}

function aggiungiPianoModal() {
  if (!cantiereAttivo) return;
  const nome = prompt('Nome del piano (es. "Piano 1", "Primo livello"):');
  if (!nome || !nome.trim()) return;
  
  fetch('/api/cantieri/' + cantiereAttivo + '/piani', {
    method: 'POST',
    headers: {'Content-Type': 'application/json'},
    body: JSON.stringify({ numero: 1, nome: nome.trim() })
  })
  .then(r => r.json())
  .then(d => {
    if (d.ok) {
      loadInterfacciaPiani(cantiereAttivo);
    } else {
      alert('❌ Errore: ' + (d.error || 'Errore sconosciuto'));
    }
  })
  .catch(e => alert('❌ Errore: ' + e.message));
}

function precompilaBrandDrawer() {
  if (!selected || selected.length === 0) return;
  const inp = document.getElementById('riga-brand-input');
  const val = document.getElementById('riga-brand-val');
  if (selected.length === 1) {
    if (inp) inp.value = selected[0];
    if (val) val.value = selected[0];
    aggiornaCampiExtra();
  } else {
    // Più brand: mostra quick-select buttons sopra il campo
    let container = document.getElementById('brand-quick-select');
    if (!container) {
      container = document.createElement('div');
      container.id = 'brand-quick-select';
      container.style.cssText = 'display:flex; flex-wrap:wrap; gap:4px; margin-bottom:6px;';
      const brandRow = document.querySelector('.form-row');
      if (brandRow) brandRow.parentNode.insertBefore(container, brandRow);
    }
    container.innerHTML = '<div style="font-size:10px;color:#9ca3af;width:100%;margin-bottom:2px;">Seleziona brand per questa riga:</div>' +
      selected.map(b =>
        '<button type="button" onclick="setBrandRiga(\'' + b.replace(/'/g,"\\'") + '\')" ' +
        'style="padding:3px 8px;font-size:10px;background:rgba(59,130,245,0.2);border:1px solid rgba(59,130,245,0.4);color:#93c5fd;border-radius:4px;cursor:pointer;margin-bottom:0;">' + b + '</button>'
      ).join('');
    // Pre-compila col primo
    if (inp) inp.value = selected[0];
    if (val) val.value = selected[0];
    aggiornaCampiExtra();
  }
}

function setBrandRiga(brand) {
  const inp = document.getElementById('riga-brand-input');
  const val = document.getElementById('riga-brand-val');
  if (inp) inp.value = brand;
  if (val) val.value = brand;
  // Evidenzia bottone attivo
  document.querySelectorAll('#brand-quick-select button').forEach(b => {
    b.style.background = b.textContent === brand ? 'rgba(59,130,245,0.6)' : 'rgba(59,130,245,0.2)';
    b.style.color = b.textContent === brand ? 'white' : '#93c5fd';
  });
  aggiornaCampiExtra();
}

function closeCantiere() {
  cantiereAttivo = null;
  document.getElementById('cantiere-drawer').classList.remove('open');
  document.getElementById('cantiere-drawer-piani').classList.remove('open');
  setTimeout(() => {
    document.getElementById('cantiere-drawer').style.display = 'none';
    document.getElementById('cantiere-drawer-piani').style.display = 'none';
  }, 300);
}

function updateCantiere() {
  if (!cantiereAttivo) return;
  const stato = document.getElementById('cantiere-stato').value;
  fetch('/api/cantieri/' + cantiereAttivo, { method:'PUT', headers:{'Content-Type':'application/json'}, body: JSON.stringify({stato}) })
    .then(r => r.json()).then(() => loadCantieri());
}

function deleteCantiere() {
  if (!cantiereAttivo) return;
  if (!confirm('Eliminare questo cantiere e tutte le sue righe?')) return;
  fetch('/api/cantieri/' + cantiereAttivo, { method:'DELETE' })
    .then(r => r.json()).then(() => { closeCantiere(); loadCantieri(); });
}

function loadRighe() {
  if (!cantiereAttivo) return;
  fetch('/api/cantieri/' + cantiereAttivo + '/righe').then(r => r.json()).then(d => {
    const righe = d.righe || [];
    let totale = 0;
    righe.forEach(r => { totale += (r.importo || 0); });
    document.getElementById('righe-list').innerHTML = righe.length === 0
      ? '<div style="color:#6b7280; font-size:11px; text-align:center; padding:12px 0;">Nessun elemento aggiunto</div>'
      : righe.map(r => {
          const desc = (r.descrizione || '').replace(/^Da Oracolo\s*[—-]\s*/i, '');
          const cat = (r.categoria && r.categoria !== 'Da Oracolo') ? r.categoria : '';
          return '<div class="riga-card">' +
          '<div class="riga-card-info">' +
          '<div class="riga-card-brand">' + (r.brand||'—') + '</div>' +
          '<div class="riga-card-cat">' + (cat ? cat + (desc ? ' — ' : '') : '') + desc + '</div>' +
          '</div>' +
          '<div class="riga-card-importo">' + (r.importo ? '€' + r.importo.toFixed(0) : '—') + '</div>' +
          '<button onclick="deleteRiga(' + r.id + ')" class="btn-red btn-sm" style="margin-bottom:0; padding:3px 7px;">✕</button>' +
          '</div>';
        }).join('');
    const totBar = document.getElementById('totale-bar');
    if (righe.length > 0) {
      totBar.style.display = 'flex';
      document.getElementById('totale-valore').textContent = '€' + totale.toFixed(0);
    } else {
      totBar.style.display = 'none';
    }
  });
}

// Mappa brand → categoria materiale
const BRAND_CATEGORIA = {
  piastrelle: ['Aparici','Apavisa','Ariostea','Caesar','Casalgrande Padana','Cerasarda','Cottodeste',
               'Edimax Astor','FAP Ceramiche','FMG','Floorim','Gigacer','Iris','Italgraniti',
               'Marca Corona','Mirage','Sichenia','Tonalite','Bisazza','Noorth','Tresse'],
  legno: ['Bauwerk','CP Parquet','Iniziativa Legno','Madegan'],
  vinilico: ['Gerflor']
};

function aggiornaCampiExtra() {
  const brand = document.getElementById('riga-brand-val').value || document.getElementById('riga-brand-input').value;
  document.getElementById('extra-piastrelle').style.display = 'none';
  document.getElementById('extra-legno').style.display = 'none';
  document.getElementById('extra-vinilico').style.display = 'none';
  if (BRAND_CATEGORIA.piastrelle.includes(brand)) {
    document.getElementById('extra-piastrelle').style.display = 'block';
    document.getElementById('riga-categoria').value = 'Pavimenti / Rivestimenti';
  } else if (BRAND_CATEGORIA.legno.includes(brand)) {
    document.getElementById('extra-legno').style.display = 'block';
    document.getElementById('riga-categoria').value = 'Parquet / Legno';
  } else if (BRAND_CATEGORIA.vinilico.includes(brand)) {
    document.getElementById('extra-vinilico').style.display = 'block';
    document.getElementById('riga-categoria').value = 'Pavimento Tecnico';
  }
}

function addRiga() {
  if (!cantiereAttivo) return;
  const brand = document.getElementById('riga-brand-val').value;
  const categoria = document.getElementById('riga-categoria').value.trim();
  let descrizione = document.getElementById('riga-descrizione').value.trim();
  const importo = parseFloat(document.getElementById('riga-importo').value) || 0;
  if (!brand && !categoria) { alert('Inserisci almeno brand o categoria'); return; }

  // Raccolta campi extra
  const extraParts = [];
  if (document.getElementById('extra-piastrelle').style.display !== 'none') {
    const fmt = document.getElementById('extra-formato').value.trim();
    const fin = document.getElementById('extra-finitura').value.trim();
    const col = document.getElementById('extra-colore').value.trim();
    if (fmt) extraParts.push('Formato: ' + fmt);
    if (fin) extraParts.push('Finitura: ' + fin);
    if (col) extraParts.push('Colore/Tono: ' + col);
  } else if (document.getElementById('extra-legno').style.display !== 'none') {
    const ess = document.getElementById('extra-essenza').value.trim();
    const fin = document.getElementById('extra-legno-finitura').value.trim();
    const fmt = document.getElementById('extra-legno-formato').value.trim();
    const ton = document.getElementById('extra-legno-tono').value.trim();
    if (ess) extraParts.push('Essenza: ' + ess);
    if (fin) extraParts.push('Finitura: ' + fin);
    if (fmt) extraParts.push('Formato: ' + fmt);
    if (ton) extraParts.push('Tono: ' + ton);
  } else if (document.getElementById('extra-vinilico').style.display !== 'none') {
    const fmt = document.getElementById('extra-vin-formato').value.trim();
    const spe = document.getElementById('extra-vin-spessore').value.trim();
    const eff = document.getElementById('extra-vin-effetto').value.trim();
    if (fmt) extraParts.push('Formato: ' + fmt);
    if (spe) extraParts.push('Spessore: ' + spe);
    if (eff) extraParts.push('Effetto: ' + eff);
  }
  if (extraParts.length > 0) {
    descrizione = (descrizione ? descrizione + ' — ' : '') + extraParts.join(' | ');
  }

  fetch('/api/cantieri/' + cantiereAttivo + '/righe', { method:'POST', headers:{'Content-Type':'application/json'},
    body: JSON.stringify({brand, categoria, descrizione, importo}) })
    .then(r => r.json()).then(d => {
      if (d.ok) {
        document.getElementById('riga-brand-input').value = '';
        document.getElementById('riga-brand-val').value = '';
        document.getElementById('riga-categoria').value = '';
        document.getElementById('riga-descrizione').value = '';
        document.getElementById('riga-importo').value = '';
        // Reset campi extra
        ['extra-formato','extra-finitura','extra-colore','extra-essenza','extra-legno-finitura',
         'extra-legno-formato','extra-legno-tono','extra-vin-formato','extra-vin-spessore','extra-vin-effetto']
          .forEach(id => { const el = document.getElementById(id); if (el) el.value = ''; });
        document.getElementById('extra-piastrelle').style.display = 'none';
        document.getElementById('extra-legno').style.display = 'none';
        document.getElementById('extra-vinilico').style.display = 'none';
        loadRighe();
      }
    });
}

function deleteRiga(rid) {
  fetch('/api/cantieri/righe/' + rid, { method:'DELETE' }).then(() => loadRighe());
}

function generaOffertaCantiere() {
  if (!cantiereAttivo) return;
  fetch('/api/cantieri/' + cantiereAttivo + '/righe').then(r => r.json()).then(d => {
    const righe = d.righe || [];
    if (righe.length === 0) { return; }
    const nome = document.getElementById('drawer-nome').textContent;
    let totale = 0;
    const riepilogo = righe.map(r => {
      totale += (r.importo || 0);
      const prezzo = r.importo ? ' | Prezzo: €' + r.importo.toFixed(2) : ' | Prezzo: da definire';
      return '- ' + (r.brand||'') + ' | ' + (r.categoria||'') + ' | ' + (r.descrizione||'') + prezzo;
    }).join('\n');
    const brands = [...new Set(righe.map(r => r.brand).filter(Boolean))];
    if (brands.length === 0) { alert('Aggiungi brand alle righe'); return; }
    const domanda = 'Genera una proposta commerciale professionale da presentare al cliente per il cantiere "' + nome + '".\n\n' +
      'ELEMENTI DEL PROGETTO:\n' + riepilogo + '\n\n' +
      'TOTALE OFFERTA: €' + totale.toFixed(2) + '\n\n' +
      'La proposta deve:\n' +
      '1. Avere un testo introduttivo professionale e convincente\n' +
      '2. Elencare ogni voce con descrizione commerciale e prezzo\n' +
      '3. Mostrare il totale finale in modo chiaro\n' +
      '4. Chiudersi con una call to action per il cliente\n' +
      'Usa un tono elegante, orientato al valore e alla qualità.';
    closeCantiere();
    askDirect(domanda, brands);
  });
}

// ============================================================================
// GENERA OFFERTA AI — MODALITA PIANI
// ============================================================================

function generaOffertaPiani() {
  if (!cantiereAttivo) {
    alert('❌ Nessun cantiere selezionato');
    return;
  }
  
  console.log('🚀 generaOffertaPiani avviata per cantiere:', cantiereAttivo);
  
  // Carica la struttura completa PIANI > STANZE > VOCI
  fetch('/api/cantieri/' + cantiereAttivo + '/struttura')
    .then(r => {
      console.log('📡 Response status:', r.status);
      return r.json();
    })
    .then(d => {
      console.log('📦 Dati ricevuti:', d);
      
      if (!d.ok) {
        alert('❌ Errore API: ' + (d.error || 'Sconosciuto'));
        return;
      }
      
      if (!d.piani || d.piani.length === 0) {
        alert('❌ Nessun piano trovato nel cantiere');
        return;
      }
      
      const cantiereNome = document.getElementById('drawer-piani-nome') 
        ? document.getElementById('drawer-piani-nome').textContent 
        : 'Cantiere ' + cantiereAttivo;
      const piani = d.piani || [];
      
      console.log('🎯 Cantiere:', cantiereNome);
      console.log('📐 Piani trovati:', piani.length);
      
      // Costruisci il riepilogo strutturato
      let riepilogo = '';
      let totaleGlobale = 0;
      const brands = new Set();
      
      piani.forEach(piano => {
        riepilogo += `\n📐 PIANO: ${piano.nome} (${piano.stanze.length} stanze)\n`;
        riepilogo += `${'─'.repeat(60)}\n`;
        
        const stanze = piano.stanze || [];
        let totalePiano = 0;
        
        stanze.forEach(stanza => {
          riepilogo += `  🏠 ${stanza.nome}\n`;
          const voci = stanza.voci || [];
          
          let totaleStanza = 0;
          voci.forEach(voce => {
            const subtotale = (voce.quantita || 1) * (voce.prezzo_unitario || 0);
            totaleStanza += subtotale;
            totalePiano += subtotale;
            totaleGlobale += subtotale;
            
            if (voce.brand) brands.add(voce.brand);
            
            riepilogo += `     [${voce.codice || '—'}] ${voce.brand || '—'} — ${voce.descrizione || '—'}\n`;
            riepilogo += `     Qty: ${voce.quantita} × €${(voce.prezzo_unitario || 0).toFixed(2)} = €${subtotale.toFixed(2)}\n`;
          });
          
          riepilogo += `  💰 SUBTOTALE ${stanza.nome}: €${totaleStanza.toFixed(2)}\n\n`;
        });
        
        riepilogo += `📊 TOTALE ${piano.nome}: €${totalePiano.toFixed(2)}\n`;
        riepilogo += `${'═'.repeat(60)}\n`;
      });
      
      riepilogo += `\n🎯 TOTALE CANTIERE: €${totaleGlobale.toFixed(2)}\n`;
      
      const brandsArray = Array.from(brands);
      
      console.log('✅ Riepilogo costruito, lunghezza:', riepilogo.length);
      console.log('🏷️ Brands:', brandsArray);
      
      // Costruisci il prompt per l'IA
      const domanda = `Genera una proposta commerciale professionale e convincente da presentare al cliente per il cantiere "${cantiereNome}".

STRUTTURA DEL PROGETTO:
${riepilogo}

ISTRUZIONI:
1. Crea un'introduzione elegante e professionale che evidenzi il valore della soluzione proposta
2. Organizza il contenuto per PIANO e STANZA in modo visibile e facile da seguire
3. Per ogni voce includi: codice prodotto, brand, descrizione accurata e prezzo
4. Aggiungi note su qualità, caratteristiche tecniche e vantaggi
5. Evidenzia i subtotali per ogni stanza e il totale finale
6. Chiudi con una call to action professionale per convertire il cliente
7. Tono: elegante, orientato al valore, ricco di dettagli ma leggibile

BRANDS PROPOSTI: ${brandsArray.join(', ') || 'Vari'}

Genera il testo da presentare direttamente al cliente.`;
      
      console.log('💬 Prompt creato, lunghezza:', domanda.length);
      
      // Chiudi il drawer e invia all'IA
      document.getElementById('cantiere-drawer-piani').classList.remove('open');
      setTimeout(() => {
        console.log('📤 Invio a askDirect...');
        askDirect(domanda, brandsArray);
      }, 300);
    })
    .catch(e => {
      console.error('❌ ERRORE FETCH:', e);
      alert('❌ Errore: ' + e.message);
    });
}

function askDirect(domanda, brands) {
  const chat = document.getElementById('chat');
  chat.innerHTML += '<div class="message"><strong>Tu:</strong> Genera offerta cantiere — ' + document.getElementById('drawer-nome') ? '' : brands.join(', ') + '</div>';
  const loadingId = 'loading_' + Date.now();
  chat.innerHTML += '<div class="message" id="' + loadingId + '" style="opacity:0.6;font-style:italic">Oracolo sta elaborando l\'offerta...</div>';
  chat.scrollTop = chat.scrollHeight;
  fetch('/api/ask', { method:'POST', headers:{'Content-Type':'application/json'},
    body: JSON.stringify({ question: domanda, brands: brands, web: webEnabled, access_code: accessCode }) })
    .then(r => r.json())
    .then(d => {
      const loading = document.getElementById(loadingId);
      if (loading) loading.remove();
      const formatted = parseMarkdown(d.answer || 'Nessuna risposta');
      const msgId = 'msg_' + Date.now();
      const query = encodeURIComponent(brands.join(' '));
      let html = '<div class="message oracolo-msg" id="' + msgId + '">';
      html += '<button class="copy-btn" onclick="copyRisposta(\'' + msgId + '\')">Copia</button>';
      html += '<div style="margin-top:6px;line-height:1.6">' + formatted + '</div>';
      if (d.images && d.images.length > 0) {
        html += '<div style="margin-top:12px; border-top:1px solid rgba(59,130,245,0.2); padding-top:10px;">';
        html += '<div style="font-size:10px; color:#6b7280; font-weight:700; text-transform:uppercase; letter-spacing:0.05em; margin-bottom:8px;">Immagini prodotti</div>';
        html += '<div style="display:grid; grid-template-columns:repeat(3,1fr); gap:6px;">';
        d.images.forEach(img => {
          html += '<div style="aspect-ratio:1; overflow:hidden; border-radius:6px; background:rgba(30,41,59,0.8); cursor:pointer;" onclick="window.open(\'' + img + '\',\'_blank\')">';
          html += '<img src="' + img + '" style="width:100%;height:100%;object-fit:cover;" onerror="this.parentElement.style.display=\'none\'">';
          html += '</div>';
        });
        html += '</div></div>';
      } else {
        html += '<div style="margin-top:8px;"><a href="https://www.google.com/search?q=' + query + '&tbm=isch" target="_blank" style="display:inline-block;padding:5px 12px;background:rgba(59,130,245,0.2);border:1px solid rgba(59,130,245,0.4);border-radius:4px;color:#93c5fd;font-size:11px;text-decoration:none;">Cerca immagini</a></div>';
      }
      html += '</div>';
      chat.innerHTML += html;
      chat.scrollTop = chat.scrollHeight;
    })
    .catch(e => {
      const loading = document.getElementById(loadingId);
      if (loading) loading.remove();
      chat.innerHTML += '<div class="message" style="color:#ef4444"><strong>Errore:</strong> ' + e + '</div>';
      chat.scrollTop = chat.scrollHeight;
    });
}

// ---------------------------------------------------------------------------
// BI
// ---------------------------------------------------------------------------
function loadBI() {
  fetch('/api/bi/stats').then(r => r.json()).then(d => {
    const ps = d.per_stato || {};
    let html = '<div style="font-size:11px;font-weight:600;color:#9ca3af;margin-bottom:4px;">Per stato</div>';
    ['bozza','inviata','vinta','persa'].forEach(s => {
      if (ps[s]) html += '<div class="bi-stat"><span>' + s + '</span><span>' + ps[s] + '</span></div>';
    });
    html += '<div class="bi-stat" style="margin-top:4px;"><span>Valore vinto</span><span style="color:#10b981;">€' + (d.valore_vinto||0).toFixed(0) + '</span></div>';
    html += '<div class="bi-stat"><span>Valore aperto</span><span style="color:#3b82f6;">€' + (d.valore_aperto||0).toFixed(0) + '</span></div>';
    if (d.per_commerciale && d.per_commerciale.length > 0) {
      html += '<div style="font-size:11px;font-weight:600;color:#9ca3af;margin-top:8px;margin-bottom:4px;">Per commerciale</div>';
      d.per_commerciale.forEach(pc => {
        html += '<div class="bi-stat"><span>' + pc.nome + '</span><span>' + pc.cantieri + ' | €' + pc.valore.toFixed(0) + '</span></div>';
      });
    }
    document.getElementById('bi-stats').innerHTML = html;
  });
}

function biCancella() {
  const da = document.getElementById('bi-da').value;
  const a = document.getElementById('bi-a').value;
  if (!da || !a) { alert('Seleziona un range di date'); return; }
  const stati = [];
  if (document.getElementById('del-vinta').checked) stati.push('vinta');
  if (document.getElementById('del-persa').checked) stati.push('persa');
  if (document.getElementById('del-bozza').checked) stati.push('bozza');
  if (stati.length === 0) { alert('Seleziona almeno uno stato'); return; }
  if (!confirm('Cancellare i cantieri selezionati? Operazione irreversibile.')) return;
  fetch('/api/bi/cancella', { method:'POST', headers:{'Content-Type':'application/json'}, body: JSON.stringify({da, a, stati}) })
    .then(r => r.json()).then(d => {
      if (d.ok) { alert('Eliminati ' + d.eliminati + ' cantieri'); loadBI(); loadCantieri(); }
    });
}

// ---------------------------------------------------------------------------
// BRANDS / DROPDOWN
// ---------------------------------------------------------------------------
// ---------------------------------------------------------------------------
// MAPPA BRAND → CATEGORIE + FASCIA DI MERCATO
// ---------------------------------------------------------------------------
const BRAND_MAP = {
  // LUXURY
  'Antoniolupi':    { categorie: ['rubinetteria','lavabi','vasche','accessori bagno','specchi'], fascia: 'luxury' },
  'Gessi':          { categorie: ['rubinetteria','doccia','vasche','accessori bagno','wellness'], fascia: 'luxury' },
  'Duscholux':      { categorie: ['box doccia','piatti doccia'], fascia: 'luxury' },
  'Kaldewei':       { categorie: ['vasche','piatti doccia'], fascia: 'luxury' },
  'Glamm Fire':     { categorie: ['camini','fuoco'], fascia: 'luxury' },
  'Tubes':          { categorie: ['radiatori','scaldasalviette'], fascia: 'luxury' },
  'Vismara Vetro':  { categorie: ['box doccia','vetro','partizioni'], fascia: 'luxury' },
  'Decor Walther':  { categorie: ['accessori bagno','specchi'], fascia: 'luxury' },
  'Bisazza':        { categorie: ['mosaico','piastrelle'], fascia: 'luxury' },
  'Cottodeste':     { categorie: ['piastrelle','pavimenti'], fascia: 'luxury' },
  'Sunshower':      { categorie: ['wellness','doccia','sauna'], fascia: 'luxury' },
  'Sunshower Wellness': { categorie: ['wellness','sauna'], fascia: 'luxury' },
  'Trimline Fires': { categorie: ['camini','fuoco'], fascia: 'luxury' },
  'Stuv':           { categorie: ['camini','stufe'], fascia: 'luxury' },
  'Austroflamm':    { categorie: ['camini','stufe'], fascia: 'luxury' },
  'Valdama':        { categorie: ['lavabi','vasche','sanitari'], fascia: 'luxury' },
  'Milldue':        { categorie: ['mobili bagno','arredo bagno'], fascia: 'luxury' },
  'Noorth':         { categorie: ['piastrelle','pavimenti'], fascia: 'luxury' },
  // PREMIUM
  'Duravit':        { categorie: ['sanitari','lavabi','vasche','mobili bagno','rubinetteria'], fascia: 'premium' },
  'Cielo':          { categorie: ['sanitari','lavabi'], fascia: 'premium' },
  'Cerasa':         { categorie: ['mobili bagno','specchi','sanitari'], fascia: 'premium' },
  'Colombo':        { categorie: ['accessori bagno','rubinetteria'], fascia: 'premium' },
  'Grupp Bardelli': { categorie: ['piastrelle','rivestimenti'], fascia: 'premium' },
  'Gruppo Bardelli':{ categorie: ['piastrelle','rivestimenti'], fascia: 'premium' },
  'Gruppo Geromin': { categorie: ['vasche','box doccia','idromassaggio'], fascia: 'premium' },
  'FAP Ceramiche':  { categorie: ['piastrelle','rivestimenti','pavimenti'], fascia: 'premium' },
  'Ariostea':       { categorie: ['piastrelle','grandi lastre','pavimenti'], fascia: 'premium' },
  'Mirage':         { categorie: ['piastrelle','pavimenti'], fascia: 'premium' },
  'FMG':            { categorie: ['piastrelle','pavimenti','grandi lastre'], fascia: 'premium' },
  'Floorim':        { categorie: ['piastrelle','pavimenti'], fascia: 'premium' },
  'Gigacer':        { categorie: ['piastrelle','grandi lastre'], fascia: 6 },
  'Italgraniti':    { categorie: ['piastrelle','pavimenti'], fascia: 'premium' },
  'Bauwerk':        { categorie: ['parquet','legno'], fascia: 'premium' },
  'CP Parquet':     { categorie: ['parquet','legno'], fascia: 'premium' },
  'Iniziativa Legno':{ categorie: ['parquet','legno'], fascia: 'premium' },
  'Madegan':        { categorie: ['parquet','legno'], fascia: 'premium' },
  'Gerflor':        { categorie: ['pavimento vinilico','pavimento tecnico'], fascia: 'premium' },
  'Acquabella':     { categorie: ['piatti doccia','vasche','box doccia'], fascia: 'premium' },
  'Gridiron':       { categorie: ['accessori bagno','portasalviette'], fascia: 'premium' },
  'Wedi':           { categorie: ['impermeabilizzazione','sistemi doccia','edilizia'], fascia: 'premium' },
  'Schluter Systems':{ categorie: ['profili','impermeabilizzazione','piastrelle'], fascia: 'premium' },
  'Tresse':         { categorie: ['piastrelle','mosaico'], fascia: 'premium' },
  'Tonalite':       { categorie: ['piastrelle','rivestimenti'], fascia: 'premium' },
  'Sterneldesign':  { categorie: ['accessori bagno'], fascia: 'premium' },
  // MID
  'Altamarea':      { categorie: ['rubinetteria','doccia'], fascia: 'mid' },
  'Anem':           { categorie: ['rubinetteria','doccia'], fascia: 'mid' },
  'Aparici':        { categorie: ['piastrelle','rivestimenti'], fascia: 'mid' },
  'Apavisa':        { categorie: ['piastrelle','pavimenti'], fascia: 'mid' },
  'Artesia':        { categorie: ['vasche','piatti doccia'], fascia: 'mid' },
  'BGP':            { categorie: ['accessori bagno'], fascia: 'mid' },
  'Blue Design':    { categorie: ['mobili bagno'], fascia: 'mid' },
  'Baufloor':       { categorie: ['pavimenti'], fascia: 'mid' },
  'Caros':          { categorie: ['piastrelle'], fascia: 'mid' },
  'Caesar':         { categorie: ['piastrelle','pavimenti'], fascia: 'mid' },
  'Casalgrande Padana':{ categorie: ['piastrelle','pavimenti'], fascia: 'mid' },
  'Cerasarda':      { categorie: ['piastrelle','rivestimenti'], fascia: 'mid' },
  'CSA':            { categorie: ['rubinetteria'], fascia: 'mid' },
  'Demm':           { categorie: ['accessori bagno'], fascia: 'mid' },
  'DoorAmeda':      { categorie: ['porte','pareti'], fascia: 'mid' },
  'Edimax Astor':   { categorie: ['piastrelle','mosaico'], fascia: 'mid' },
  'Brera':          { categorie: ['sanitari','lavabi'], fascia: 'mid' },
  'GOman':          { categorie: ['accessori bagno'], fascia: 'mid' },
  'Ier Hurne':      { categorie: ['accessori bagno'], fascia: 'mid' },
  'Inklostro Bianco':{ categorie: ['pitture','rivestimenti'], fascia: 'mid' },
  'Iris':           { categorie: ['piastrelle','sanitari'], fascia: 'mid' },
  'Linki':          { categorie: ['accessori bagno'], fascia: 'mid' },
  'Marca Corona':   { categorie: ['piastrelle','rivestimenti'], fascia: 'mid' },
  'Murexin':        { categorie: ['impermeabilizzazione','massetti','posa'], fascia: 'mid' },
  'Omegius':        { categorie: ['accessori bagno'], fascia: 'mid' },
  "Piastrelle d Arredo":{ categorie: ['piastrelle'], fascia: 'mid' },
  'Profiletec':     { categorie: ['profili','bordi'], fascia: 'mid' },
  'SDR':            { categorie: ['sanitari'], fascia: 'mid' },
  'Sichenia':       { categorie: ['piastrelle','pavimenti'], fascia: 'mid' },
  'Simas':          { categorie: ['sanitari','lavabi'], fascia: 'mid' },
  'Remer':          { categorie: ['rubinetteria','accessori bagno'], fascia: 'mid' },
};

const FASCIA_LABEL = {
  luxury:  { label: 'Luxury',  color: '#f59e0b' },
  premium: { label: 'Premium', color: '#8b5cf6' },
  mid:     { label: 'Mid',     color: '#3b82f6' },
  entry:   { label: 'Entry',   color: '#6b7280' },
};

function switchTab(tab) {
  const isBrand = tab === 'brand';
  document.getElementById('tab-content-brand').style.display = isBrand ? '' : 'none';
  document.getElementById('tab-content-cat').style.display = isBrand ? 'none' : '';
  document.getElementById('tab-brand').style.background = isBrand ? '#3b82f6' : 'rgba(59,130,245,0.3)';
  document.getElementById('tab-cat').style.background = isBrand ? 'rgba(59,130,245,0.3)' : '#3b82f6';
  if (!isBrand) document.getElementById('search-cat').focus();
}

function filterPerCategoria() {
  const sv = document.getElementById('search-cat').value.toLowerCase().trim();
  const container = document.getElementById('cat-results');
  if (!sv) { container.innerHTML = '<div style="font-size:10px;color:#6b7280;padding:4px 0;">Digita una categoria...</div>'; return; }

  // Trova brand che matchano la categoria cercata
  const risultati = { luxury: [], premium: [], mid: [], entry: [] };
  Object.entries(BRAND_MAP).forEach(([brand, info]) => {
    if (info.categorie && info.categorie.some(c => c.toLowerCase().includes(sv))) {
      const fascia = info.fascia || 'mid';
      if (risultati[fascia]) risultati[fascia].push(brand);
    }
  });

  const ordine = ['luxury', 'premium', 'mid', 'entry'];
  let html = '';
  let totale = 0;
  ordine.forEach(f => {
    if (risultati[f].length === 0) return;
    const fl = FASCIA_LABEL[f];
    html += '<div style="font-size:9px;font-weight:700;color:' + fl.color + ';text-transform:uppercase;margin:6px 0 3px 0;letter-spacing:0.06em;">' + fl.label + '</div>';
    risultati[f].forEach(brand => {
      totale++;
      const checked = new Set(selected).has(brand) ? 'checked' : '';
      html += '<div class="brand-item cat-item" style="display:flex;align-items:center;gap:6px;">' +
        '<input type="checkbox" value="' + brand + '" ' + checked + ' onchange="updateSelected()">' +
        '<span style="flex:1">' + brand + '</span>' +
        '<span style="font-size:9px;color:' + fl.color + ';font-weight:600;">' + fl.label + '</span>' +
        '</div>';
    });
  });

  if (totale === 0) {
    html = '<div style="font-size:10px;color:#6b7280;padding:4px 0;">Nessun brand trovato per "' + sv + '"</div>';
  }
  container.innerHTML = html;
}

function loadBrands() {
  fetch('/api/get-brands').then(r => r.json()).then(d => {
    BRANDS = d.brands || [];
    loadGroups();
  }).catch(e => console.error("Errore brand:", e));
}

function toggleDropdown() {
  const dd = document.getElementById('dropdown');
  if (!dd) return;
  if (dd.classList.contains('show')) {
    dd.classList.remove('show');
  } else {
    dd.classList.add('show');
    // Costruisce la lista solo se è vuota
    const brandsList = document.getElementById('brands-list');
    if (brandsList && brandsList.querySelectorAll('.brand-item').length === 0) {
      filterBrands();
    }
  }
}

function filterBrands() {
  const search = document.getElementById('search');
  const brandsList = document.getElementById('brands-list');
  if (!search || !brandsList) return;
  const sv = search.value.toLowerCase();

  // Se la lista è già popolata, filtra per visibilità senza ricostruire il DOM
  const existing = brandsList.querySelectorAll('.brand-item');
  if (existing.length > 0) {
    existing.forEach(item => {
      const inp = item.querySelector('input');
      const val = inp ? inp.value.toLowerCase() : '';
      item.style.display = val.includes(sv) ? '' : 'none';
    });
    return;
  }

  // Prima costruzione — popola con stato checked corrente
  const alreadySelected = new Set(selected);
  brandsList.innerHTML = BRANDS.map(b =>
    '<div class="brand-item"><input type="checkbox" value="' + b + '" ' +
    (alreadySelected.has(b) ? 'checked' : '') +
    ' onchange="updateSelected()">' + b + '</div>'
  ).join('');
}

function updateSelected() {
  selected = [];
  document.querySelectorAll('.brand-item input:checked').forEach(cb => { selected.push(cb.value); });
  document.getElementById('selected').innerHTML = selected.map(b => '<span class="badge">' + b + ' x</span>').join('');
  // Propaga brand selezionato in tutti i campi brand dell'app
  if (selected.length > 0) {
    const b = selected[0];
    // Upload documenti
    const uploadInp = document.getElementById('upload-brand-input');
    const uploadVal = document.getElementById('upload-brand-val');
    if (uploadInp) uploadInp.value = b;
    if (uploadVal) uploadVal.value = b;
  }
  // Se drawer aperto, aggiorna quick-select
  if (cantiereAttivo) precompilaBrandDrawer();
}

function toggleWeb() {
  webEnabled = !webEnabled;
  const btn = document.getElementById('web-toggle');
  btn.textContent = webEnabled ? 'ON' : 'OFF';
  btn.className = webEnabled ? 'toggle-btn toggle-on' : 'toggle-btn toggle-off';
}

function toggleAccess() {
  const code = document.getElementById('access-code').value;
  if (code) { accessCode = code; accessLevel = "private"; document.getElementById('access-status').textContent = 'Accesso: PRIVATO (' + code + ')'; }
  else { accessCode = null; accessLevel = "public"; document.getElementById('access-status').textContent = 'Accesso: PUBBLICO'; }
}

function saveGroup() {
  const name = document.getElementById('group-name').value;
  if (!name || selected.length === 0) { alert('Nome gruppo e brand richiesti'); return; }
  groups[name] = selected;
  localStorage.setItem('oracolo_groups', JSON.stringify(groups));
  document.getElementById('group-name').value = '';
  loadGroups();
}

function loadGroups() {
  document.getElementById('saved-groups').innerHTML = Object.keys(groups).map(name =>
    '<div style="padding:4px;background:rgba(59,130,245,0.2);border-radius:4px;margin:3px 0;font-size:11px"><strong>' + name + '</strong> <button onclick="loadGroup(\'' + name + '\')" style="padding:1px 5px;font-size:9px">carica</button> <button onclick="deleteGroup(\'' + name + '\')" style="padding:1px 5px;font-size:9px;background:#ef4444">x</button></div>'
  ).join('');
}

function loadGroup(name) {
  selected = groups[name] || [];
  updateSelected();
  document.getElementById('dropdown').classList.remove('show');
}

function deleteGroup(name) {
  delete groups[name];
  localStorage.setItem('oracolo_groups', JSON.stringify(groups));
  loadGroups();
}

function setAdminPassword() {
  if (selected.length === 0) { alert('Seleziona prima un brand'); return; }
  const brand = selected[0];
  const pwd = document.getElementById('admin-pwd').value.trim();
  if (!pwd) { alert('Inserisci una password'); return; }
  fetch('/api/set-admin-password', { method:'POST', headers:{'Content-Type':'application/json'}, body: JSON.stringify({brand, admin_password: pwd}) })
    .then(r => r.json()).then(d => {
      if (d.ok) alert('OK: ' + d.message);
      else alert('Errore: ' + d.error);
      document.getElementById('admin-pwd').value = '';
    });
}

function addCassetto() {
  const nome = document.getElementById('new-cassetto').value.trim();
  if (!nome) { alert('Nome richiesto'); return; }
  fetch('/api/add-azienda', { method:'POST', headers:{'Content-Type':'application/json'}, body:JSON.stringify({nome}) })
    .then(r => r.json()).then(d => {
      if (d.ok) {
        alert('Cassetto aggiunto!');
        document.getElementById('new-cassetto').value = '';
        fetch('/api/get-brands').then(r => r.json()).then(data => { BRANDS = data.brands || []; });
      } else alert('Errore: ' + d.error);
    });
}

function doUpload(input, tipo) {
  const brand = document.getElementById('upload-brand-val').value;
  if (!brand) { document.getElementById('upload-status').textContent = 'Seleziona prima un brand!'; document.getElementById('upload-status').style.color = '#ef4444'; input.value=''; return; }
  const file = input.files[0];
  if (!file) return;
  document.getElementById('upload-status').textContent = 'Caricamento...';
  document.getElementById('upload-status').style.color = '#9ca3af';
  const reader = new FileReader();
  reader.onload = function(e) {
    const filename = tipo === 'excel' ? file.name + ' [EXCEL]' : file.name;
    fetch('/api/upload-document', { method:'POST', headers:{'Content-Type':'application/json'},
      body: JSON.stringify({filename, content: e.target.result, brand, visibility: accessLevel, access_code: accessCode}) })
      .then(r => r.json()).then(d => {
        if (d.ok) { document.getElementById('upload-status').textContent = 'Caricato: ' + file.name; document.getElementById('upload-status').style.color = '#10b981'; }
        else { document.getElementById('upload-status').textContent = 'Errore: ' + d.error; document.getElementById('upload-status').style.color = '#ef4444'; }
        input.value = '';
      });
  };
  reader.readAsDataURL(file);
}

function apriGestisciDoc() {
  document.getElementById('gestisci-doc-panel').style.display = 'flex';
  document.getElementById('filtro-doc-brand').value = selected.length > 0 ? selected[0] : '';
  filtraDocumenti();
}

function chiudiGestisciDoc() {
  document.getElementById('gestisci-doc-panel').style.display = 'none';
}

function filtraDocumenti(tutti) {
  const brand = tutti ? '' : document.getElementById('filtro-doc-brand').value.trim();
  const url = brand ? '/api/list-documents?brand=' + encodeURIComponent(brand) : '/api/list-documents';
  const container = document.getElementById('doc-list-panel');
  container.innerHTML = '<div style="color:#6b7280;font-size:11px;padding:12px 0;">Caricamento...</div>';

  fetch(url).then(r => r.json()).then(d => {
    const docs = d.documents || [];
    if (docs.length === 0) {
      container.innerHTML = '<div style="color:#6b7280;font-size:11px;padding:12px 0;">Nessun documento trovato' + (brand ? ' per "' + brand + '"' : '') + '</div>';
      return;
    }
    // Raggruppa per brand
    const grouped = {};
    docs.forEach(doc => {
      const b = doc.brand || '—';
      if (!grouped[b]) grouped[b] = [];
      grouped[b].push(doc);
    });

    let html = '';
    Object.entries(grouped).sort().forEach(([b, bdocs]) => {
      html += '<div style="font-size:10px;font-weight:700;color:#60a5fa;text-transform:uppercase;letter-spacing:0.06em;margin:10px 0 6px 0;">' + b + ' <span style="color:#6b7280;font-weight:400;">(' + bdocs.length + ')</span></div>';
      bdocs.forEach(doc => {
        const isExcel = doc.filename.includes('[EXCEL]');
        const tipoHtml = isExcel
          ? '<span class="doc-tipo-excel">📊 Excel</span>'
          : '<span class="doc-tipo-doc">📄 Doc</span>';
        const dataStr = doc.date ? doc.date.substring(0, 16).replace('T', ' ') : '—';
        const visHtml = doc.visibility === 'private'
          ? '<span style="font-size:9px;color:#f59e0b;">🔒 Privato</span>'
          : '<span style="font-size:9px;color:#6b7280;">🌐 Pubblico</span>';
        const nomeFile = doc.filename.replace(' [EXCEL]', '');
        html += '<div class="doc-row" id="docrow-' + doc.id + '">' +
          tipoHtml +
          '<div style="flex:1;min-width:0;">' +
          '<div style="font-weight:600;color:#e0e0e0;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;">' + nomeFile + '</div>' +
          '<div style="color:#6b7280;font-size:10px;">' + dataStr + ' · ' + visHtml + '</div>' +
          '</div>' +
          '<button onclick="eliminaDocumento(' + doc.id + ',\'' + b.replace(/'/g,"\\'") + '\')" class="btn-red btn-sm" style="margin-bottom:0;white-space:nowrap;">✕ Elimina</button>' +
          '</div>';
      });
    });
    container.innerHTML = html;
  });
}

function eliminaDocumento(id, brand) {
  if (!confirm('Eliminare questo documento di ' + brand + '?')) return;
  fetch('/api/delete-document/' + id, { method: 'DELETE' })
    .then(r => r.json())
    .then(d => {
      if (d.ok) {
        const row = document.getElementById('docrow-' + id);
        if (row) { row.style.opacity = '0.3'; row.style.pointerEvents = 'none'; setTimeout(() => row.remove(), 600); }
      } else {
        alert('Errore: ' + (d.error || 'sconosciuto'));
      }
    });
}

function generateOfferta() {
  if (!selected.length) { alert('Seleziona brand'); return; }
  document.getElementById('question').value = 'Genera una proposta commerciale da presentare al cliente finale, illustrando i vantaggi e le caratteristiche dei brand selezionati: ' + selected.join(', ') + '. Il documento deve essere professionale, orientato al cliente e convincente.';
  ask();
}
function generateAnalisi() {
  if (!selected.length) { alert('Seleziona brand'); return; }
  document.getElementById('question').value = 'Crea una scheda analitica da condividere con il cliente sui brand selezionati: ' + selected.join(', ') + '. Evidenzia punti di forza, materiali, design e perché sono la scelta giusta per un progetto di qualità.';
  ask();
}
function generateProposta() {
  if (!selected.length) { alert('Seleziona brand'); return; }
  document.getElementById('question').value = 'Prepara una proposta strategica da presentare al cliente per un progetto che utilizzi i brand: ' + selected.join(', ') + '. Includi posizionamento, benefici concreti e suggerimenti di abbinamento.';
  ask();
}

function copyRisposta(msgId) {
  const el = document.getElementById(msgId);
  if (!el) return;
  const testo = el.innerText.replace('Copia\n', '').replace('Oracolo:\n', '').trim();
  navigator.clipboard.writeText(testo).then(() => {
    const btn = el.querySelector('.copy-btn');
    if (btn) { btn.textContent = 'Copiato!'; setTimeout(() => { btn.textContent = 'Copia'; }, 2000); }
  }).catch(() => {
    const range = document.createRange();
    range.selectNode(el);
    window.getSelection().removeAllRanges();
    window.getSelection().addRange(range);
    document.execCommand('copy');
    window.getSelection().removeAllRanges();
  });
}

function parseMarkdown(text) {
  return text
    .replace(/### (.+)/g, '<h3 style="color:#60a5fa;margin:10px 0 4px 0;font-size:13px">$1</h3>')
    .replace(/## (.+)/g, '<h2 style="color:#3b82f6;margin:12px 0 6px 0;font-size:14px">$1</h2>')
    .replace(/# (.+)/g, '<h1 style="color:#3b82f6;margin:12px 0 6px 0;font-size:15px">$1</h1>')
    .replace(/\*\*(.+?)\*\*/g, '<strong>$1</strong>')
    .replace(/\*(.+?)\*/g, '<em>$1</em>')
    .replace(/^[-\u2013\u2014] (.+)/gm, '<li style="margin-left:16px;margin-bottom:3px">$1</li>')
    .replace(/(<li.*<\/li>)/gs, '<ul style="margin:6px 0">$1</ul>')
    .replace(/\n{2,}/g, '</p><p style="margin:6px 0">')
    .replace(/\n/g, '<br>');
}

function ask() {
  if (!selected.length) { alert('Seleziona brand'); return; }
  const q = document.getElementById('question').value;
  if (!q) return;
  document.getElementById('question').value = '';
  const chat = document.getElementById('chat');
  chat.innerHTML += '<div class="message"><strong>Tu:</strong> ' + q + '</div>';
  const loadingId = 'loading_' + Date.now();
  chat.innerHTML += '<div class="message" id="' + loadingId + '" style="opacity:0.6;font-style:italic">Oracolo sta elaborando...</div>';
  chat.scrollTop = chat.scrollHeight;
  fetch('/api/ask', { method:'POST', headers:{'Content-Type':'application/json'},
    body: JSON.stringify({question: q, brands: selected, web: webEnabled, access_code: accessCode}) })
    .then(r => r.json())
    .then(d => {
      const loading = document.getElementById(loadingId);
      if (loading) loading.remove();
      const formatted = parseMarkdown(d.answer || 'Nessuna risposta');
      const msgId = 'msg_' + Date.now();
      let html = '<div class="message oracolo-msg" id="' + msgId + '">';
      html += '<button class="copy-btn" onclick="copyRisposta(\'' + msgId + '\')">Copia</button>';
      html += '<div style="margin-top:6px;line-height:1.6">' + formatted + '</div>';
      const query = encodeURIComponent(selected.join(' ') + ' ' + q);
      if (d.images && d.images.length > 0) {
        html += '<div style="margin-top:12px; border-top:1px solid rgba(59,130,245,0.2); padding-top:10px;">';
        html += '<div style="font-size:10px; color:#6b7280; font-weight:700; text-transform:uppercase; letter-spacing:0.05em; margin-bottom:8px;">Immagini prodotti</div>';
        html += '<div style="display:grid; grid-template-columns:repeat(3,1fr); gap:6px;">';
        d.images.forEach(img => {
          html += '<div style="aspect-ratio:1; overflow:hidden; border-radius:6px; background:rgba(30,41,59,0.8); cursor:pointer;" onclick="window.open(\'' + img + '\',\'_blank\')">';
          html += '<img src="' + img + '" style="width:100%;height:100%;object-fit:cover;" onerror="this.parentElement.style.display=\'none\'">';
          html += '</div>';
        });
        html += '</div></div>';
      } else {
        html += '<div style="margin-top:8px;"><a href="https://www.google.com/search?q=' + query + '&tbm=isch" target="_blank" style="display:inline-block;padding:5px 12px;background:rgba(59,130,245,0.2);border:1px solid rgba(59,130,245,0.4);border-radius:4px;color:#93c5fd;font-size:11px;text-decoration:none;">Cerca immagini</a></div>';
      }
      // --- BOTTONE AGGIUNGI AL CARRELLO ---
      const safeBrand = (selected[0] || '').replace(/'/g, "\\'");
      html += '<div style="margin-top:10px; border-top:1px solid rgba(16,185,129,0.2); padding-top:8px;">';
      html += '<button id="btn-carrello-' + msgId + '" onclick="apriFormCarrello(\'' + msgId + '\',\'' + safeBrand + '\')" ' +
        'style="background:rgba(16,185,129,0.2); border:1px solid rgba(16,185,129,0.5); color:#10b981; padding:5px 12px; border-radius:5px; font-size:11px; font-weight:600; cursor:pointer; margin-bottom:0;">✓ Aggiungi al carrello</button>';
      html += '</div>';
      // --- FORM CARRELLO (hidden) — desc verrà riempita da AI dopo apertura ---
      html += '<div id="form-carrello-' + msgId + '" style="display:none; margin-top:8px; background:rgba(16,185,129,0.08); border:1px solid rgba(16,185,129,0.3); border-radius:6px; padding:10px;">';
      html += '<div style="font-size:10px; color:#10b981; font-weight:700; text-transform:uppercase; margin-bottom:8px;">Aggiungi al carrello</div>';
      html += '<input type="text" id="fc-codice-' + msgId + '" placeholder="Codice prodotto (opzionale)" style="width:100%; margin-bottom:6px; font-size:11px;">';
      html += '<textarea id="fc-desc-' + msgId + '" rows="2" placeholder="Sintesi in caricamento..." style="width:100%; margin-bottom:6px; font-size:11px; background:rgba(30,41,59,0.8); border:1px solid rgba(59,130,245,0.3); color:white; border-radius:6px; padding:6px; resize:vertical;"></textarea>';
      html += '<div style="display:flex; gap:6px;">';
      html += '<input type="number" id="fc-prezzo-' + msgId + '" placeholder="Prezzo €" style="flex:1; font-size:11px;">';
      html += '<button onclick="confermaDaChat(\'' + msgId + '\',\'' + safeBrand + '\')" class="btn-green" style="flex:1; margin-bottom:0; font-size:11px;">✓ Conferma</button>';
      html += '<button onclick="document.getElementById(\'form-carrello-' + msgId + '\').style.display=\'none\'" class="btn-gray" style="flex:none; margin-bottom:0; font-size:11px;">✕</button>';
      html += '</div></div>';
      html += '</div>';
      chat.innerHTML += html;
      // Salva testo grezzo sul DOM per uso successivo
      const msgEl = document.getElementById(msgId);
      if (msgEl) msgEl.dataset.rawAnswer = d.answer || '';
      chat.scrollTop = chat.scrollHeight;
    })
    .catch(e => {
      const loading = document.getElementById(loadingId);
      if (loading) loading.remove();
      chat.innerHTML += '<div class="message" style="color:#ef4444"><strong>Errore:</strong> ' + e + '</div>';
    });
}

// ---------------------------------------------------------------------------
// AGGIUNGI DA CHAT AL CARRELLO
// ---------------------------------------------------------------------------
function apriFormCarrello(msgId, brand) {
  if (!cantiereAttivo) {
    alert('Apri prima un cantiere dal pannello destra');
    return;
  }
  const form = document.getElementById('form-carrello-' + msgId);
  if (!form) return;
  const isOpen = form.style.display !== 'none';
  form.style.display = isOpen ? 'none' : 'block';
  if (!isOpen) {
    // Se la textarea è ancora vuota, chiedi sintesi AI
    const ta = document.getElementById('fc-desc-' + msgId);
    if (ta && !ta.value.trim()) {
      const msgEl = document.getElementById(msgId);
      const rawAnswer = msgEl ? msgEl.dataset.rawAnswer || '' : '';
      ta.value = 'Sintesi in caricamento...';
      fetch('/api/arricchisci-prodotto', {
        method: 'POST',
        headers: {'Content-Type': 'application/json'},
        body: JSON.stringify({ descrizione: rawAnswer, brand: brand })
      })
      .then(r => r.json())
      .then(d => { if (d.ok) ta.value = d.descrizione_ai; else ta.value = rawAnswer.substring(0, 200); })
      .catch(() => { ta.value = rawAnswer.substring(0, 200); });
    }
  }
}

function confermaDaChat(msgId, brand) {
  if (!cantiereAttivo) { alert('Apri prima un cantiere'); return; }
  const codice = document.getElementById('fc-codice-' + msgId).value.trim();
  const descEl = document.getElementById('fc-desc-' + msgId);
  const prezzoEl = document.getElementById('fc-prezzo-' + msgId);
  const desc = descEl ? descEl.value.trim() : '';
  const prezzo = parseFloat(prezzoEl ? prezzoEl.value : '') || 0;
  if (!desc) { alert('La descrizione è vuota'); return; }
  const descrizione = codice ? '[' + codice + '] ' + desc : desc;
  // Brand = sempre quello selezionato nella sidebar
  const brandEffettivo = (selected && selected.length > 0) ? selected[0] : brand;

  // Disabilita subito il bottone per evitare doppio click
  const form = document.getElementById('form-carrello-' + msgId);
  const confirmBtn = form ? form.querySelector('.btn-green') : null;
  if (confirmBtn) { confirmBtn.disabled = true; confirmBtn.textContent = '⏳'; }

  fetch('/api/cantieri/' + cantiereAttivo + '/righe', {
    method: 'POST',
    headers: {'Content-Type': 'application/json'},
    body: JSON.stringify({ brand: brandEffettivo, categoria: '', descrizione: descrizione, importo: prezzo })
  })
  .then(r => r.json())
  .then(d => {
    if (d.ok) {
      if (form) {
        form.innerHTML = '<div style="color:#10b981; font-size:11px; font-weight:600; padding:4px 0;">✓ Aggiunto al carrello!</div>';
        setTimeout(() => { form.style.display = 'none'; }, 2000);
      }
      // Disabilita anche il bottone esterno
      const outerBtn = document.getElementById('btn-carrello-' + msgId);
      if (outerBtn) { outerBtn.textContent = '✓ Aggiunto'; outerBtn.disabled = true; outerBtn.style.opacity = '0.5'; }
      loadRighe();
    } else {
      if (confirmBtn) { confirmBtn.disabled = false; confirmBtn.textContent = '✓ Conferma'; }
    }
  });
}

// ---------------------------------------------------------------------------
// EXCEL INTELLIGENTE
// ---------------------------------------------------------------------------
let excelRighe = [];  // righe caricate dall'Excel

function toggleExcelPanel() {
  const panel = document.getElementById('excel-panel');
  const arrow = document.getElementById('excel-panel-arrow');
  const open = panel.style.display !== 'none';
  panel.style.display = open ? 'none' : 'block';
  arrow.textContent = open ? '▼' : '▲';
}

function scaricaImmaginiGessi() {
  if (!confirm('Scarica URL immagini Gessi (prodotti + accessori)? Questo potrebbe richiedere 1-2 minuti...')) return;
  
  const btn = event.target;
  btn.disabled = true;
  btn.textContent = '⏳ In corso...';
  
  fetch('/api/scarica-immagini/Gessi', {method: 'POST'})
    .then(r => r.json())
    .then(d => {
      btn.disabled = false;
      if (d.ok) {
        const msg = `✓ Prodotti: ${d.prodotti_aggiornati}/${d.prodotti_totali} | Accessori: ${d.accessori_aggiornati}/${d.accessori_totali}`;
        btn.textContent = msg;
        btn.style.background = '#10b981';
        setTimeout(() => {
          btn.textContent = '🖼️ Scarica URL Immagini Gessi';
          btn.style.background = '#06b6d4';
        }, 4000);
      } else {
        btn.textContent = '❌ Errore: ' + d.error;
        btn.style.background = '#ef4444';
      }
    })
    .catch(e => {
      btn.disabled = false;
      btn.textContent = '❌ Errore';
      console.error(e);
    });
}

function caricaAbbinamentiEProdotti() {
  const btn = event.target;
  btn.disabled = true;
  btn.textContent = '⏳ Caricamento...';
  
  // 1. Carica listino
  fetch('/api/listino/Gessi')
    .then(r => r.json())
    .then(d1 => {
      const prodotti = d1.ok ? d1.prodotti.length : 0;
      
      // 2. Carica abbinamenti
      return fetch('/api/carica-abbinamenti-excel/Gessi', {method:'POST'})
        .then(r => r.json())
        .then(d2 => {
          const abbinamenti = d2.ok ? d2.count : 0;
          
          if (d1.ok && d2.ok) {
            btn.textContent = `✅ Caricati ${prodotti} prodotti + ${abbinamenti} abbinamenti`;
            btn.style.background = '#10b981';
          } else {
            btn.textContent = `⚠️ ${prodotti} prodotti, ${abbinamenti} abbinamenti`;
            btn.style.background = '#f59e0b';
          }
          
          setTimeout(() => {
            btn.textContent = '📋 Carica Listino + Abbinamenti';
            btn.style.background = '#f59e0b';
            btn.disabled = false;
          }, 4000);
        });
    })
    .catch(e => {
      btn.textContent = '❌ Errore: ' + e.message;
      btn.style.background = '#ef4444';
      btn.disabled = false;
    });
}

function caricaExcelListino(input) {
  const file = input.files[0];
  if (!file) return;
  document.getElementById('excel-status').textContent = 'Lettura Excel...';
  document.getElementById('excel-status').style.color = '#9ca3af';
  const reader = new FileReader();
  reader.onload = function(e) {
    fetch('/api/parse-excel', {
      method: 'POST',
      headers: {'Content-Type': 'application/json'},
      body: JSON.stringify({content: e.target.result})
    })
    .then(r => r.json())
    .then(d => {
      if (d.ok) {
        excelRighe = d.righe;
        document.getElementById('excel-status').textContent = '✓ ' + d.totale + ' prodotti trovati';
        document.getElementById('excel-status').style.color = '#10b981';
        renderExcelRighe();
      } else {
        document.getElementById('excel-status').textContent = 'Errore: ' + d.error;
        document.getElementById('excel-status').style.color = '#ef4444';
      }
    })
    .catch(err => {
      document.getElementById('excel-status').textContent = 'Errore: ' + err;
      document.getElementById('excel-status').style.color = '#ef4444';
    });
  };
  reader.readAsDataURL(file);
  input.value = '';
}

function renderExcelRighe() {
  const container = document.getElementById('excel-righe-list');
  if (!excelRighe.length) { container.innerHTML = ''; return; }

  // Campo ricerca
  let html = '<input type="text" id="excel-search" placeholder="Filtra prodotti..." ' +
    'oninput="renderExcelRigheFiltered()" style="width:100%; margin-bottom:8px; font-size:11px;">';
  html += '<div id="excel-righe-inner"></div>';
  container.innerHTML = html;
  renderExcelRigheFiltered();
}

function renderExcelRigheFiltered() {
  const searchEl = document.getElementById('excel-search');
  const sv = searchEl ? searchEl.value.toLowerCase() : '';
  const filtered = sv
    ? excelRighe.filter(r => (r.codice + ' ' + r.descrizione).toLowerCase().includes(sv))
    : excelRighe;

  const inner = document.getElementById('excel-righe-inner');
  if (!inner) return;

  inner.innerHTML = filtered.slice(0, 50).map((r, i) => {
    const idx = excelRighe.indexOf(r);
    const prezzoHtml = r.prezzo !== null && r.prezzo !== undefined
      ? '<span class="' + (r.prezzo_src === 'excel' ? 'prezzo-excel' : 'prezzo-web') + '">€' +
        parseFloat(r.prezzo).toFixed(2) + (r.prezzo_src !== 'excel' ? ' ⚠web' : '') + '</span>'
      : '<span style="color:#6b7280; font-size:10px;">prezzo mancante</span>';

    const aiDesc = r.descrizione_ai
      ? '<div class="excel-ai-desc">' + r.descrizione_ai + '</div>'
      : '';

    return '<div class="excel-row" id="excel-row-' + idx + '">' +
      '<div class="excel-row-header">' +
      '<span class="excel-codice">' + (r.codice || '—') + '</span>' +
      '<span class="excel-desc">' + (r.descrizione || '—') + '</span>' +
      prezzoHtml +
      '</div>' +
      aiDesc +
      '<div class="excel-actions">' +
      '<button onclick="arricchisciRiga(' + idx + ')" class="btn-sm" style="background:rgba(59,130,245,0.3);color:#93c5fd;margin-bottom:0;">✨ Arricchisci</button>' +
      '<input type="number" id="prezzo-edit-' + idx + '" placeholder="Modifica €" value="' + (r.prezzo !== null ? r.prezzo : '') + '" ' +
        'style="width:90px; font-size:10px; padding:3px 6px; flex:none;" ' +
        'onchange="aggiornaPrezzo(' + idx + ')">' +
      '<button onclick="aggiungiAlCarrello(' + idx + ')" class="btn-sm btn-green" style="margin-bottom:0;">✓ Carrello</button>' +
      '</div>' +
      '</div>';
  }).join('');

  if (filtered.length > 50) {
    inner.innerHTML += '<div style="font-size:10px; color:#9ca3af; text-align:center; padding:6px;">Mostrati 50 di ' + filtered.length + ' — usa il filtro per trovare</div>';
  }
}

function aggiornaPrezzo(idx) {
  const input = document.getElementById('prezzo-edit-' + idx);
  if (!input) return;
  const val = parseFloat(input.value);
  if (!isNaN(val)) {
    excelRighe[idx].prezzo = val;
    excelRighe[idx].prezzo_src = 'excel';  // manuale = trattato come excel (verde)
    renderExcelRigheFiltered();
  }
}

function arricchisciRiga(idx) {
  const r = excelRighe[idx];
  const btn = document.querySelector('#excel-row-' + idx + ' button');
  if (btn) { btn.textContent = '⏳'; btn.disabled = true; }

  fetch('/api/arricchisci-prodotto', {
    method: 'POST',
    headers: {'Content-Type': 'application/json'},
    body: JSON.stringify({
      codice: r.codice,
      descrizione: r.descrizione,
      prezzo: r.prezzo,
      brand: selected.length === 1 ? selected[0] : (selected[0] || '')
    })
  })
  .then(res => res.json())
  .then(d => {
    if (d.ok) {
      excelRighe[idx].descrizione_ai = d.descrizione_ai;
      renderExcelRigheFiltered();
    }
  })
  .catch(() => { if (btn) { btn.textContent = '✨ Arricchisci'; btn.disabled = false; } });
}

function aggiungiAlCarrello(idx) {
  if (!cantiereAttivo) { alert('Apri prima un cantiere'); return; }
  const r = excelRighe[idx];
  const descrizione = r.descrizione_ai || r.descrizione || '';
  const importo = r.prezzo || 0;
  const brand = selected.length === 1 ? selected[0] : (selected[0] || '');

  fetch('/api/cantieri/' + cantiereAttivo + '/righe-da-ai', {
    method: 'POST',
    headers: {'Content-Type': 'application/json'},
    body: JSON.stringify({
      brand: brand,
      categoria: '',
      codice: r.codice,
      descrizione: descrizione,
      importo: importo
    })
  })
  .then(res => res.json())
  .then(d => {
    if (d.ok) {
      // Feedback visivo sul bottone
      const btns = document.querySelectorAll('#excel-row-' + idx + ' button');
      const addBtn = btns[btns.length - 1];
      if (addBtn) { addBtn.textContent = '✓ Aggiunto!'; addBtn.style.background = '#10b981'; setTimeout(() => { addBtn.textContent = '✓ Carrello'; addBtn.style.background = ''; }, 2000); }
      loadRighe();
    }
  });
}

function addVoceManuale() {
  if (!cantiereAttivo) { alert('Apri prima un cantiere'); return; }
  const desc = document.getElementById('voce-desc').value.trim();
  const importo = parseFloat(document.getElementById('voce-importo').value) || 0;
  if (!desc) { alert('Inserisci una descrizione'); return; }

  fetch('/api/cantieri/' + cantiereAttivo + '/righe', {
    method: 'POST',
    headers: {'Content-Type': 'application/json'},
    body: JSON.stringify({brand: '', categoria: 'Voce manuale', descrizione: desc, importo: importo})
  })
  .then(r => r.json())
  .then(d => {
    if (d.ok) {
      document.getElementById('voce-desc').value = '';
      document.getElementById('voce-importo').value = '';
      loadRighe();
    }
  });
}

// ---------------------------------------------------------------------------
// RICERCA RAPIDA LISTINO DALLA CHAT
// ---------------------------------------------------------------------------
let cercaRapidaTimer = null;

function cercaRapidaListino(val) {
  const qsr = document.getElementById('quick-search-results');
  clearTimeout(cercaRapidaTimer);
  if (!val || val.length < 3 || !selected.length) {
    qsr.style.display = 'none';
    return;
  }
  cercaRapidaTimer = setTimeout(() => {
    fetch('/api/cerca-prodotto', {
      method: 'POST',
      headers: {'Content-Type': 'application/json'},
      body: JSON.stringify({ query: val, brand: selected[0] })
    })
    .then(r => r.json())
    .then(d => {
      if (!d.prodotti || d.prodotti.length === 0) {
        qsr.style.display = 'none';
        return;
      }
      const listinoTipoAttuale = listinoTipo || 'cliente';
      let html = '<div style="padding:6px 10px; font-size:9px; color:#6b7280; border-bottom:1px solid rgba(59,130,245,0.15);">📄 Trovato nel listino Excel — clicca per aggiungere al carrello</div>';
      d.prodotti.forEach((p, idx) => {
        const prezzo = listinoTipoAttuale === 'rivenditore' && p.prezzo_rivenditore ? p.prezzo_rivenditore : p.prezzo;
        const prezzoRiv = p.prezzo_rivenditore;
        const prezzoLabel = prezzo ? '<span style="color:#10b981;font-weight:700;">€' + parseFloat(prezzo).toFixed(0) + '</span>' + (prezzoRiv && listinoTipoAttuale === 'cliente' ? '<span style="color:#f59e0b;font-size:9px;margin-left:4px;">riv.€' + parseFloat(prezzoRiv).toFixed(0) + '</span>' : '') : '<span style="color:#6b7280;">—</span>';
        const disp = (p.disponibilita||'').toLowerCase().includes('ordine') ? '⏳' : '✓';
        const dispColor = (p.disponibilita||'').toLowerCase().includes('ordine') ? '#f59e0b' : '#10b981';
        html += '<div style="padding:8px 10px; border-bottom:1px solid rgba(59,130,245,0.1); display:flex; align-items:center; gap:8px; cursor:pointer;" ' +
          'onmouseover="this.style.background=\'rgba(59,130,245,0.1)\'" onmouseout="this.style.background=\'\'">' +
          '<div style="flex:1;">' +
          '<div style="font-size:10px; font-weight:600; color:#e0e0e0;">' + (p.nome||p.codice) + '</div>' +
          '<div style="font-size:9px; color:#9ca3af;">' + (p.codice||'') + (p.collezione ? ' · ' + p.collezione : '') + ' <span style="color:' + dispColor + ';">' + disp + '</span></div>' +
          '</div>' +
          '<div style="text-align:right;">' + prezzoLabel + '</div>' +
          '<button onclick="aggiungiDaRicercaRapida(' + idx + ',this)" class="btn-green btn-sm" style="margin-bottom:0;white-space:nowrap;">+ Carrello</button>' +
          '<button onclick="usaDescrizioneRapida(' + idx + ')" class="btn-sm" style="background:rgba(59,130,245,0.3);color:#93c5fd;margin-bottom:0;">AI ✨</button>' +
          '</div>';
      });
      // Salva risultati per uso nei bottoni
      window._qsResults = d.prodotti;
      qsr.innerHTML = html;
      qsr.style.display = 'block';
    });
  }, 350);
}

function aggiungiDaRicercaRapida(idx, btn) {
  if (!cantiereAttivo) { alert('Apri prima un cantiere'); return; }
  const p = window._qsResults[idx];
  if (!p) return;
  const listinoTipoAttuale = listinoTipo || 'cliente';
  const importo = listinoTipoAttuale === 'rivenditore' && p.prezzo_rivenditore ? p.prezzo_rivenditore : (p.prezzo || 0);
  const descrizione = (p.codice ? '[' + p.codice + '] ' : '') + (p.nome || p.descrizione || '');
  if (btn) { btn.textContent = '⏳'; btn.disabled = true; }
  fetch('/api/cantieri/' + cantiereAttivo + '/righe', {
    method: 'POST',
    headers: {'Content-Type': 'application/json'},
    body: JSON.stringify({ brand: selected[0] || '', categoria: p.categoria||'', descrizione, importo })
  })
  .then(r => r.json())
  .then(d => {
    if (d.ok) {
      if (btn) { btn.textContent = '✓'; btn.style.background = '#10b981'; }
      loadRighe();
      setTimeout(() => { document.getElementById('quick-search-results').style.display = 'none'; }, 1000);
    }
  });
}

function usaDescrizioneRapida(idx) {
  const p = window._qsResults[idx];
  if (!p) return;
  document.getElementById('question').value = 'Descrivi commercialmente: ' + (p.nome||p.codice) + ' [' + (p.codice||'') + '] — ' + (p.descrizione||'');
  document.getElementById('quick-search-results').style.display = 'none';
  ask();
}

// Chiudi risultati rapidi se clicco fuori
document.addEventListener('click', function(e) {
  const qsr = document.getElementById('quick-search-results');
  const q = document.getElementById('question');
  if (qsr && q && !qsr.contains(e.target) && e.target !== q) {
    qsr.style.display = 'none';
  }
});

// =============================================================================
// INTERFACCIA 2 — PIANI/STANZE/VOCI
// =============================================================================

let pianiBrowserData = {};
let stanzaAttivaPerCarrello = null;

function loadInterfacciaPiani(cantiere_id) {
  const pannello = document.getElementById('pannello-piani');
  if (!pannello) return;

  pannello.innerHTML = '<div style="padding:12px; background:rgba(59,130,245,0.1); border-radius:6px; color:#93c5fd; font-size:11px;">⏳ Caricamento struttura piani...</div>';

  fetch('/api/cantieri/' + cantiere_id + '/struttura')
    .then(r => r.json())
    .then(d => {
      if (!d.ok) {
        pannello.innerHTML = '<div style="color:#ef4444; padding:12px;">❌ Errore: ' + d.error + '</div>';
        return;
      }

      pianiBrowserData = d.piani || [];
      renderInterfacciaPiani();
    })
    .catch(e => {
      pannello.innerHTML = '<div style="color:#ef4444; padding:12px;">❌ ' + e.message + '</div>';
    });
}

function renderInterfacciaPiani() {
  const pannello = document.getElementById('pannello-piani');
  if (!pannello || !pianiBrowserData) return;

  if (pianiBrowserData.length === 0) {
    pannello.innerHTML = '<div style="padding:20px; text-align:center; color:#9ca3af;"><div style="font-size:14px; margin-bottom:12px;">📐 Nessun piano ancora</div><button onclick="aggiungiPianoUI()" class="btn-green" style="padding:8px 16px;">➕ Crea primo piano</button></div>';
    return;
  }

  let html = '<div style="padding:0;">';

  pianiBrowserData.forEach((piano, pIdx) => {
    const isOpen = localStorage.getItem('piano_open_' + piano.id) === '1';
    
    html += '<div style="background:rgba(59,130,245,0.12); border:1px solid rgba(59,130,245,0.3); border-radius:8px; margin-bottom:12px; overflow:hidden;">' +
      '<div onclick="togglePiano(' + piano.id + ')" style="padding:12px; background:rgba(59,130,245,0.2); cursor:pointer; display:flex; justify-content:space-between; align-items:center; user-select:none;">' +
      '<div style="display:flex; align-items:center; gap:8px; flex:1;"><span style="font-size:14px; color:#60a5fa; font-weight:bold;">' + (isOpen ? '▼' : '▶') + ' ' + piano.nome + '</span><span style="font-size:10px; color:#9ca3af;">(' + piano.stanze.length + ' stanze)</span></div>' +
      '<div style="font-size:12px; color:#10b981; font-weight:bold; background:rgba(16,185,129,0.2); padding:3px 8px; border-radius:4px;">€' + (piano.totale_piano || 0).toFixed(2) + '</div>' +
      '</div>' +
      '<div id="piano-body-' + piano.id + '" style="display:' + (isOpen ? 'block' : 'none') + '; padding:8px;">';

    (piano.stanze || []).forEach((stanza, sIdx) => {
      const stanzaOpen = localStorage.getItem('stanza_open_' + stanza.id) === '1';
      const voci = stanza.voci || [];
      
      html += '<div style="background:rgba(30,41,59,0.8); border-left:3px solid #8b5cf6; border-radius:6px; margin-bottom:6px; overflow:hidden;">' +
        '<div onclick="toggleStanza(' + stanza.id + ')" style="padding:8px 10px; cursor:pointer; display:flex; justify-content:space-between; align-items:center; user-select:none; background:rgba(139,92,246,0.15);">' +
        '<div style="display:flex; align-items:center; gap:6px; flex:1;"><span style="font-size:11px; color:#d1d5db; font-weight:bold;">' + (stanzaOpen ? '▼' : '▶') + ' 🏠 ' + stanza.nome + '</span></div>' +
        '<div style="display:flex; gap:8px; align-items:center;"><span style="font-size:11px; color:#10b981; font-weight:bold;">€' + (stanza.totale_stanza || 0).toFixed(2) + '</span>' +
        '<button onclick="event.stopPropagation(); aggiungiVoceStanza(' + stanza.id + ', \'' + stanza.nome.replace(/'/g, "\\'") + '\')" class="btn-green btn-sm" style="padding:2px 6px; font-size:9px; margin-bottom:0;">+ Voce</button></div>' +
        '</div>' +
        '<div id="stanza-body-' + stanza.id + '" style="display:' + (stanzaOpen ? 'block' : 'none') + '; padding:6px;">';

      if (voci.length === 0) {
        html += '<div style="padding:6px; font-size:10px; color:#6b7280; font-style:italic;">Nessuna voce</div>';
      } else {
        voci.forEach(voce => {
          html += '<div style="padding:5px 6px; margin:3px 0; background:rgba(59,130,245,0.1); border-radius:4px; font-size:10px; color:#e0e0e0; display:flex; justify-content:space-between; align-items:center;">' +
            '<div style="flex:1;"><div style="font-weight:600;">[' + (voce.codice||'—') + '] ' + (voce.brand || '—') + '</div>' +
            '<div style="font-size:9px; color:#9ca3af;">' + voce.descrizione + '</div>' +
            '<div style="font-size:9px; color:#6b7280; margin-top:2px;">Qty: ' + voce.quantita + ' | €' + voce.prezzo_unitario + ' = €' + voce.subtotale.toFixed(2) + '</div></div>' +
            '<button onclick="cancellaVoce(' + voce.id + ', ' + stanza.id + ')" class="btn-red btn-sm" style="padding:2px 4px; font-size:8px; margin-bottom:0; white-space:nowrap;">✕</button>' +
            '</div>';
        });
      }

      html += '</div></div>';
    });

    html += '<button onclick="aggiungiStanzaUI(' + piano.id + ')" class="btn-purple btn-sm" style="width:100%; margin-top:6px; margin-bottom:0;">➕ Nuova stanza</button>' +
      '</div></div>';
  });

  html += '<button onclick="aggiungiPianoUI()" class="btn-green" style="width:100%; margin-top:8px;">➕ Nuovo piano</button></div>';

  pannello.innerHTML = html;
}

function togglePiano(pianoId) {
  const body = document.getElementById('piano-body-' + pianoId);
  const isOpen = body.style.display !== 'none';
  body.style.display = isOpen ? 'none' : 'block';
  localStorage.setItem('piano_open_' + pianoId, isOpen ? '0' : '1');
}

function toggleStanza(stanzaId) {
  const body = document.getElementById('stanza-body-' + stanzaId);
  const isOpen = body.style.display !== 'none';
  body.style.display = isOpen ? 'none' : 'block';
  localStorage.setItem('stanza_open_' + stanzaId, isOpen ? '0' : '1');
}

function aggiungiPianoUI() {
  if (!cantiereAttivo) return;
  const nome = prompt('Nome piano (es. "Piano Terra", "Primo Livello"):');
  if (!nome || !nome.trim()) return;

  fetch('/api/cantieri/' + cantiereAttivo + '/piani', {
    method: 'POST',
    headers: {'Content-Type': 'application/json'},
    body: JSON.stringify({ numero: pianiBrowserData.length + 1, nome: nome.trim() })
  })
  .then(r => r.json())
  .then(d => {
    if (d.ok) {
      loadInterfacciaPiani(cantiereAttivo);
    } else {
      alert('❌ ' + (d.error || 'Errore'));
    }
  });
}

function aggiungiStanzaUI(pianoId) {
  if (!cantiereAttivo) return;
  const nome = prompt('Nome stanza (es. "Bagno principale", "Doccia"):');
  if (!nome || !nome.trim()) return;

  fetch('/api/piani/' + pianoId + '/stanze', {
    method: 'POST',
    headers: {'Content-Type': 'application/json'},
    body: JSON.stringify({ nome: nome.trim() })
  })
  .then(r => r.json())
  .then(d => {
    if (d.ok) {
      loadInterfacciaPiani(cantiereAttivo);
    } else {
      alert('❌ ' + (d.error || 'Errore'));
    }
  });
}

function aggiungiVoceStanza(stanzaId, stanzaNome) {
  if (!cantiereAttivo) return;
  
  // Controlla che ci sia UN brand selezionato nella sidebar
  if (!selected || selected.length === 0) {
    alert('Seleziona un brand nella sidebar SX prima di aggiungere prodotti');
    return;
  }
  
  // Brand FISSO dal sidebar (il primo / l'unico selezionato)
  const brandScelto = selected[0];
  
  // State: ricorda quale stanza stiamo riempiendo
  stanzaAttivaPerCarrello = { id: stanzaId, nome: stanzaNome };
  stanzaSelezionataPiani = { id: stanzaId, nome: stanzaNome };
  
  // APRI IL LISTINO CENTRALE (quello che funziona!)
  // Con un flag speciale per sapere che stiamo aggiungendo a una stanza
  window._modalitaStanza = true;
  window._stanzaAggiunta = { id: stanzaId, nome: stanzaNome };
  
  // Apri il listino del brand
  apriListino(brandScelto);
}

function salvaVoceStanza() {
  if (!stanzaAttivaPerCarrello) return;

  const brand = document.getElementById('voce-brand-val').value || document.getElementById('voce-brand-input').value;
  const codice = document.getElementById('voce-codice').value.trim();
  const desc = document.getElementById('voce-desc').value.trim();
  const qty = parseFloat(document.getElementById('voce-qty').value) || 1;
  const prezzo = parseFloat(document.getElementById('voce-prezzo').value) || 0;

  if (!desc) {
    alert('Inserisci almeno una descrizione');
    return;
  }

  fetch('/api/stanze/' + stanzaAttivaPerCarrello.id + '/voci', {
    method: 'POST',
    headers: {'Content-Type': 'application/json'},
    body: JSON.stringify({
      codice: codice,
      brand: brand,
      descrizione: desc,
      quantita: qty,
      prezzo_unitario: prezzo,
      sconto_percentuale: 0,
      colore: 'verde'
    })
  })
  .then(r => r.json())
  .then(d => {
    if (d.ok) {
      document.getElementById('form-voce-stanza').remove();
      stanzaAttivaPerCarrello = null;
      loadInterfacciaPiani(cantiereAttivo);
    } else {
      alert('❌ ' + (d.error || 'Errore'));
    }
  });
}

// ============================================================================
// GRID PRODOTTI DA LISTINO PER STANZA (MODALITA' PIANI)
// ============================================================================

function apriGridProdottiStanza(brand, stanzaId, stanzaNome) {
  // Modale fullscreen con GRID CARD di prodotti
  const modalHtml = `
    <div id="modal-grid-stanza" style="position:fixed; top:0; left:0; width:100%; height:100%; background:rgba(0,0,0,0.8); z-index:6000; display:flex; flex-direction:column; padding:0;">
      <div style="background:#0f172e; border-bottom:2px solid #3b82f6; padding:16px 20px; display:flex; justify-content:space-between; align-items:center; flex-shrink:0;">
        <div style="font-size:16px; font-weight:bold; color:#60a5fa;">📋 Seleziona prodotti per: <strong>${stanzaNome}</strong> (${brand})</div>
        <div style="display:flex; gap:12px; align-items:center;">
          <input type="text" id="grid-search-stanza" placeholder="Cerca codice/nome..." style="padding:8px 12px; background:rgba(30,41,59,0.8); border:1px solid rgba(59,130,245,0.3); color:white; border-radius:6px; width:250px; font-size:11px;" oninput="filtraGridStanza()">
          <button onclick="chiudiGridStanza()" style="background:#ef4444; color:white; border:none; padding:8px 16px; border-radius:6px; cursor:pointer; font-weight:600;">✕ Chiudi</button>
        </div>
      </div>
      <div id="grid-container-stanza" style="flex:1; overflow-y:auto; padding:20px; background:#0a0f23;">
        <div style="color:#6b7280; text-align:center; padding:40px;">⏳ Caricamento listino...</div>
      </div>
    </div>
  `;
  
  document.body.insertAdjacentHTML('beforeend', modalHtml);
  
  // Carica listino del brand
  fetch('/api/listino/' + encodeURIComponent(brand))
    .then(r => r.json())
    .then(d => {
      if (!d.ok || !d.prodotti) {
        document.getElementById('grid-container-stanza').innerHTML = '<div style="color:#ef4444; text-align:center; padding:40px;">❌ Nessun listino trovato per ' + brand + '</div>';
        return;
      }
      
      window._gridProdottiStanza = d.prodotti;
      window._gridBrandStanza = brand;
      window._gridStanzaId = stanzaId;
      
      filtraGridStanza();
      
      // 🔑 CARICA ABBINAMENTI PER TUTTE LE CARD
      console.log('📦 Caricamento abbinamenti per ' + d.prodotti.length + ' prodotti...');
      caricaAbbinationiPerCards(d.prodotti);
    })
    .catch(e => {
      document.getElementById('grid-container-stanza').innerHTML = '<div style="color:#ef4444; text-align:center; padding:40px;">❌ ' + e.message + '</div>';
    });
}

function filtraGridStanza() {
  const search = (document.getElementById('grid-search-stanza') || {}).value || '';
  const sv = search.toLowerCase();
  
  if (!window._gridProdottiStanza) return;
  
  const filtered = sv
    ? window._gridProdottiStanza.filter(p => 
        (p.codice || '').toLowerCase().includes(sv) ||
        (p.nome || '').toLowerCase().includes(sv) ||
        (p.descrizione || '').toLowerCase().includes(sv)
      )
    : window._gridProdottiStanza;
  
  renderGridStanza(filtered);
}

function renderGridStanza(prodotti) {
  if (!prodotti || prodotti.length === 0) {
    document.getElementById('grid-container-stanza').innerHTML = '<div style="color:#6b7280; text-align:center; padding:40px;">Nessun prodotto trovato</div>';
    return;
  }
  
  const html = `
    <div style="display:grid; grid-template-columns:repeat(auto-fill, minmax(280px, 1fr)); gap:16px;">
      ${prodotti.map((p, idx) => `
        <div style="background:#1e293b; border:1px solid #334155; border-radius:8px; overflow:hidden; cursor:pointer; transition:all 0.2s; display:flex; flex-direction:column;" 
             onmouseover="this.style.borderColor='#3b82f6'; this.style.boxShadow='0 0 15px rgba(59,130,245,0.3)'" 
             onmouseout="this.style.borderColor='#334155'; this.style.boxShadow='none'"
             data-product-idx="${idx}">
          
          <!-- IMMAGINE THUMBNAIL -->
          <div style="width:100%; height:160px; background:rgba(30,41,59,0.8); display:flex; align-items:center; justify-content:center; overflow:hidden; position:relative; border-bottom:1px solid #334155;">
            ${p.immagine_url ? 
              `<img src="${p.immagine_url}" style="width:100%; height:100%; object-fit:cover; transition:transform 0.3s;" 
                    onmouseover="this.style.transform='scale(1.05)'" 
                    onmouseout="this.style.transform='scale(1)'" />` 
              : 
              `<div style="color:#6b7280; font-size:40px;">📦</div>`
            }
            <div style="position:absolute; top:8px; right:8px; background:#10b981; color:white; padding:4px 8px; border-radius:4px; font-size:9px; font-weight:bold;">✓ DISPONIBILE</div>
          </div>
          
          <!-- CONTENUTO CARD -->
          <div style="padding:12px; flex:1; display:flex; flex-direction:column;">
            <div style="font-size:10px; color:#9ca3af; margin-bottom:4px; font-family:monospace; font-weight:bold;">${p.codice || '—'}</div>
            <div style="font-size:11px; font-weight:600; color:#e0e0e0; margin-bottom:6px; line-height:1.3; min-height:32px;">${p.nome || '—'}</div>
            
            <!-- DESCRIZIONE CON SLIDER INTERATTIVO -->
            <div style="margin-bottom:8px;">
              <div id="desc-text-${idx}" style="font-size:9px; color:#9ca3af; line-height:1.4; margin-bottom:6px; min-height:24px; max-height:60px; overflow-y:auto; padding-right:4px;">
                ${p.descrizione ? p.descrizione.substring(0, 100) : 'Nessuna descrizione'}
              </div>
              
              <!-- SLIDER CARATTERI -->
              <div style="display:flex; align-items:center; gap:6px; padding:4px 0;">
                <span style="font-size:7px; color:#6b7280; white-space:nowrap;">50</span>
                <input type="range" id="desc-slider-${idx}" 
                       min="50" max="${p.descrizione ? p.descrizione.length : 100}" value="100" step="10"
                       style="flex:1; height:4px; cursor:pointer; accent-color:#8b5cf6;"
                       oninput="aggiornaDescrizioneSlider(${idx}, '${p.descrizione ? p.descrizione.replace(/'/g, "\\'") : ''}')">
                <span id="desc-count-${idx}" style="font-size:7px; color:#c084fc; font-weight:bold; white-space:nowrap; width:25px; text-align:right;">100</span>
                <span style="font-size:7px; color:#6b7280; white-space:nowrap;">All</span>
              </div>
            </div>
            
            <!-- BADGE COLORE/TIPO -->
            <div style="font-size:8px; color:#6b7280; margin-bottom:8px; display:flex; gap:4px; flex-wrap:wrap;">
              ${p.categoria ? `<span style="background:rgba(59,130,245,0.2); padding:2px 6px; border-radius:3px;">${p.categoria}</span>` : ''}
              ${p.tipo ? `<span style="background:rgba(168,85,247,0.2); padding:2px 6px; border-radius:3px;">${p.tipo}</span>` : ''}
            </div>
            
            <!-- ABBINAMENTI PREVIEW (se ce ne sono) -->
            <div id="abbinamenti-preview-${idx}" style="font-size:8px; color:#c084fc; margin-bottom:8px; display:none;">
              <div style="font-weight:bold; margin-bottom:2px;">🔗 Abbinamenti:</div>
              <div id="abbinamenti-list-${idx}" style="display:flex; gap:2px; flex-wrap:wrap; max-height:30px; overflow:hidden;"></div>
            </div>
            
            <!-- PREZZO -->
            <div style="font-size:14px; color:#10b981; font-weight:bold; margin-bottom:12px;">€${p.prezzo ? parseFloat(p.prezzo).toFixed(0) : '—'}</div>
            
            <!-- BOTTONI AZIONI -->
            <div style="display:flex; gap:6px; margin-top:auto; flex-direction:column;">
              <div style="display:flex; gap:6px;">
                <button id="btn-abbina-${idx}" onclick="event.stopPropagation(); apriModaleAbbinamenti(${idx})" 
                        style="flex:1; padding:6px; background:#ef4444; color:white; border:none; border-radius:4px; font-size:10px; cursor:pointer; font-weight:600; transition:background 0.2s; display:none;" 
                        onmouseover="this.style.background='#dc2626'" 
                        onmouseout="this.style.background='#ef4444'">
                  🔗 Abbina
                </button>
                <button onclick="event.stopPropagation(); aggiungiProdottoStanza(${idx})" 
                        style="flex:1; padding:6px; background:#3b82f6; color:white; border:none; border-radius:4px; font-size:10px; cursor:pointer; font-weight:600; transition:background 0.2s;" 
                        onmouseover="this.style.background='#2563eb'" 
                        onmouseout="this.style.background='#3b82f6'">
                  ✓ Aggiungi
                </button>
                <button onclick="event.stopPropagation(); aggiungiAlCarrello(${idx})" 
                        style="flex:1; padding:6px; background:#10b981; color:white; border:none; border-radius:4px; font-size:10px; cursor:pointer; font-weight:600; transition:background 0.2s;" 
                        onmouseover="this.style.background='#059669'" 
                        onmouseout="this.style.background='#10b981'">
                  🛒 Carrello
                </button>
              </div>
              <button onclick="event.stopPropagation(); apriModaleImmagine(${idx})" 
                      style="width:100%; padding:6px; background:#8b5cf6; color:white; border:none; border-radius:4px; font-size:10px; cursor:pointer; font-weight:600; transition:background 0.2s;" 
                      onmouseover="this.style.background='#7c3aed'" 
                      onmouseout="this.style.background='#8b5cf6'">
                🖼️ Cerca Immagine
              </button>
            </div>
          </div>
        </div>
      `).join('')}
    </div>
  `;
  
  document.getElementById('grid-container-stanza').innerHTML = html;
  
  // DOPO il render, carica gli abbinamenti e mostra/nascondi bottone
  caricaAbbinationiPerCards(prodotti);
}

function caricaAbbinationiPerCards(prodotti) {
  const brand = window._gridBrandStanza;
  if (!brand) return;
  
  prodotti.forEach((p, idx) => {
    fetch('/api/abbinamenti/' + encodeURIComponent(brand) + '/' + encodeURIComponent(p.codice))
      .then(r => r.json())
      .then(d => {
        const abbinamenti = (d.ok && d.abbinamenti) ? d.abbinamenti : [];
        
        // Se ci sono abbinamenti:
        if (abbinamenti && abbinamenti.length > 0) {
          // 1. Mostra bottone Abbina
          const btnAbbina = document.getElementById('btn-abbina-' + idx);
          if (btnAbbina) {
            btnAbbina.style.display = 'block';
          }
          
          // 2. Mostra preview abbinamenti nella card
          const previewDiv = document.getElementById('abbinamenti-preview-' + idx);
          const listDiv = document.getElementById('abbinamenti-list-' + idx);
          
          if (previewDiv && listDiv) {
            previewDiv.style.display = 'block';
            
            // Mostra max 3 abbinamenti
            const abbinatiMostrarti = abbinamenti.slice(0, 3);
            const html = abbinatiMostrarti.map(a => 
              `<span style="background:rgba(192,132,252,0.3); padding:1px 4px; border-radius:2px; white-space:nowrap;">${a.nome || a.codice}</span>`
            ).join('');
            
            listDiv.innerHTML = html + (abbinamenti.length > 3 ? `<span style="color:#6b7280; font-size:7px;">+${abbinamenti.length - 3}</span>` : '');
          }
        }
      })
      .catch(e => {
        // Silenzioso se fallisce
        console.log('⚠️ Abbinamenti non trovati per:', p.codice);
      });
  });
}

// ============================================================================
// SLIDER DESCRIZIONE — AGGIORNA TESTO AL MOVIMENTO
// ============================================================================

function aggiornaDescrizioneSlider(idx, testoCompleto) {
  const slider = document.getElementById('desc-slider-' + idx);
  const textDiv = document.getElementById('desc-text-' + idx);
  const countSpan = document.getElementById('desc-count-' + idx);
  
  console.log('🎚️ Slider movimento idx=' + idx + ', elementi trovati:', {slider: !!slider, textDiv: !!textDiv, countSpan: !!countSpan});
  
  if (!slider || !textDiv) {
    console.warn('⚠️ Slider o textDiv non trovati per idx', idx);
    return;
  }
  
  const numCaratteri = parseInt(slider.value);
  const testoTroncato = testoCompleto.substring(0, numCaratteri);
  
  textDiv.textContent = testoTroncato + (numCaratteri < testoCompleto.length ? '' : '');
  countSpan.textContent = numCaratteri;
  
  console.log('✅ Descrizione aggiornata: ' + numCaratteri + ' caratteri');
}

function aggiungiProdottoStanza(idx) {
  if (!window._gridProdottiStanza || !window._gridStanzaId) return;
  
  const prodotto = window._gridProdottiStanza[idx];
  const brand = window._gridBrandStanza;
  
  // CONTROLLA SE CI SONO ABBINAMENTI PER QUESTO PRODOTTO
  fetch('/api/abbinamenti/' + encodeURIComponent(brand) + '/' + encodeURIComponent(prodotto.codice))
    .then(r => r.json())
    .then(d => {
      const abbinamenti = (d.ok && d.abbinamenti) ? d.abbinamenti : [];
      
      if (abbinamenti.length > 0) {
        // ✅ HA ABBINAMENTI → OBBLIGA A PASSARE DALLA MODALE
        console.log('🔗 Prodotto ha ' + abbinamenti.length + ' abbinamenti. Apri modale.');
        apriModaleAbbinamenti(idx);
      } else {
        // ❌ NESSUN ABBINAMENTO → AGGIUNGI DIRETTAMENTE
        console.log('✓ Prodotto NON ha abbinamenti. Aggiungi direttamente.');
        aggiungiProdottoSenzaAbbinamenti(idx);
      }
    })
    .catch(e => {
      // In caso di errore, aggiungi direttamente
      console.error('Errore controllo abbinamenti, aggiungi comunque:', e);
      aggiungiProdottoSenzaAbbinamenti(idx);
    });
}

function aggiungiProdottoSenzaAbbinamenti(idx) {
  if (!window._gridProdottiStanza || !window._gridStanzaId) return;
  
  const prodotto = window._gridProdottiStanza[idx];
  const stanzaId = window._gridStanzaId;
  const brand = window._gridBrandStanza;
  
  const descrizione = (prodotto.codice ? '[' + prodotto.codice + '] ' : '') + (prodotto.nome || '');
  const prezzo = prodotto.prezzo || 0;
  
  // Salva nella stanza
  fetch('/api/stanze/' + stanzaId + '/voci', {
    method: 'POST',
    headers: {'Content-Type': 'application/json'},
    body: JSON.stringify({
      codice: prodotto.codice || '',
      brand: brand,
      descrizione: descrizione,
      quantita: 1,
      prezzo_unitario: prezzo,
      sconto_percentuale: 0,
      colore: 'verde',
      immagine_url: prodotto.immagine_url || ''
    })
  })
  .then(r => r.json())
  .then(d => {
    if (d.ok) {
      console.log('✓ Aggiunto (senza abbinamenti):', prodotto.nome);
      loadInterfacciaPiani(window.cantiere_attivo_id);
      alert('✓ Aggiunto: ' + (prodotto.nome || 'Prodotto'));
    } else {
      alert('❌ ' + (d.error || 'Errore'));
    }
  });
}

// ============================================================================
// MODALE RICERCA IMMAGINE
// ============================================================================

function apriModaleImmagine(idx) {
  if (!window._gridProdottiStanza) return;
  
  const prodotto = window._gridProdottiStanza[idx];
  const brand = window._gridBrandStanza;
  
  const modalHtml = `
    <div id="modal-immagine" style="position:fixed; top:0; left:0; width:100%; height:100%; background:rgba(0,0,0,0.85); z-index:7500; display:flex; flex-direction:column; padding:0;">
      
      <!-- HEADER -->
      <div style="background:#0f172e; border-bottom:2px solid #8b5cf6; padding:16px 20px; display:flex; justify-content:space-between; align-items:center; flex-shrink:0;">
        <div style="font-size:14px; font-weight:bold; color:#c084fc;">🖼️ Gestisci Immagine: <strong>${prodotto.nome}</strong></div>
        <button onclick="chiudiModaleImmagine()" style="background:#ef4444; color:white; border:none; padding:8px 16px; border-radius:6px; cursor:pointer; font-weight:600;">✕ Chiudi</button>
      </div>
      
      <!-- CONTENUTO -->
      <div style="display:flex; gap:20px; flex:1; overflow:hidden; padding:20px;">
        
        <!-- SINISTRA: ANTEPRIMA IMMAGINE -->
        <div style="flex:0 0 300px; display:flex; flex-direction:column; background:rgba(30,41,59,0.6); border:1px solid #334155; border-radius:8px; padding:16px; overflow-y:auto;">
          <div style="font-size:12px; font-weight:bold; color:#c084fc; margin-bottom:12px;">📸 Anteprima</div>
          
          <div id="anteprima-immagine" style="width:100%; aspect-ratio:1; background:rgba(59,130,245,0.1); border-radius:6px; display:flex; align-items:center; justify-content:center; margin-bottom:12px; overflow:hidden; border:2px dashed #334155;">
            ${prodotto.immagine_url ? 
              `<img src="${prodotto.immagine_url}" style="width:100%; height:100%; object-fit:cover;" />` 
              : 
              `<div style="color:#6b7280; text-align:center; font-size:40px;">📦<div style="font-size:11px; margin-top:8px; color:#9ca3af;">Ricerca in corso...</div></div>`
            }
          </div>
          
          <div style="font-size:10px; color:#9ca3af; margin-top:auto;">
            <div style="font-weight:bold; margin-bottom:4px; color:#d1d5db;">Info Prodotto:</div>
            Codice: <strong>${prodotto.codice}</strong><br>
            Brand: <strong>${brand}</strong><br>
            Nome: <strong>${prodotto.nome}</strong>
          </div>
        </div>
        
        <!-- DESTRA: RISULTATI RICERCA -->
        <div style="flex:1; display:flex; flex-direction:column;">
          <div style="font-size:12px; font-weight:bold; color:#c084fc; margin-bottom:12px;">🔍 Ricerca Automatica</div>
          
          <div style="background:rgba(30,41,59,0.6); border:1px solid #334155; border-radius:8px; padding:16px; flex:1; display:flex; flex-direction:column; overflow:hidden;">
            <div id="ricerca-status" style="font-size:10px; color:#93c5fd; margin-bottom:12px; text-align:center;">⏳ Ricerca immagini per: <strong>${prodotto.codice}</strong></div>
            
            <div id="risultati-grid" style="display:grid; grid-template-columns:repeat(2, 1fr); gap:12px; flex:1; overflow-y:auto; margin-bottom:12px;"></div>
            
            <!-- Metodo alternativo: Incolla URL manuale -->
            <div style="border-top:1px solid #334155; padding-top:12px; margin-top:12px;">
              <div style="font-size:9px; font-weight:bold; color:#d1d5db; margin-bottom:6px;">O Incolla URL manuale:</div>
              <div style="display:flex; gap:6px;">
                <input type="text" id="url-immagine-input" placeholder="https://example.com/image.jpg" style="flex:1; padding:8px; background:rgba(30,41,59,0.8); border:1px solid #334155; color:white; border-radius:4px; font-size:10px;" value="${prodotto.immagine_url || ''}">
                <button onclick="salvaURLImmagineManuale('${idx}')" style="padding:8px 12px; background:#10b981; color:white; border:none; border-radius:4px; cursor:pointer; font-weight:600; font-size:10px;">✓ Salva</button>
              </div>
            </div>
          </div>
        </div>
      </div>
      
      <!-- FOOTER -->
      <div style="background:#0f172e; border-top:2px solid #8b5cf6; padding:16px 20px; display:flex; gap:12px; justify-content:flex-end; flex-shrink:0;">
        <button onclick="chiudiModaleImmagine()" style="padding:10px 20px; background:#6b7280; color:white; border:none; border-radius:6px; cursor:pointer; font-weight:600;">✕ Chiudi</button>
      </div>
    </div>
  `;
  
  document.body.insertAdjacentHTML('beforeend', modalHtml);
  
  // RICERCA AUTOMATICA subito
  cercaImmaginiAutomatica(idx, prodotto, brand);
}

function cercaImmaginiAutomatica(idx, prodotto, brand) {
  const codice = prodotto.codice;
  const nome = prodotto.nome;
  
  console.log('🔍 Ricerca automatica per:', codice, nome);
  
  // Chiama backend per cercare immagini
  fetch('/api/cerca-immagine-prodotto', {
    method: 'POST',
    headers: {'Content-Type': 'application/json'},
    body: JSON.stringify({
      codice: codice,
      nome: nome,
      brand: brand
    })
  })
  .then(r => r.json())
  .then(d => {
    console.log('📦 Risultati ricerca:', d);
    
    const statusEl = document.getElementById('ricerca-status');
    const gridEl = document.getElementById('risultati-grid');
    
    if (!d.ok || !d.risultati || d.risultati.length === 0) {
      statusEl.innerHTML = '❌ Nessuna immagine trovata. Incolla URL manuale.';
      return;
    }
    
    statusEl.innerHTML = `✅ Trovate ${d.risultati.length} immagini`;
    
    // Renderizza risultati
    const html = d.risultati.map((ris, ridx) => `
      <div style="background:#1e293b; border:1px solid #334155; border-radius:6px; overflow:hidden; cursor:pointer; transition:all 0.2s;" 
           onmouseover="this.style.borderColor='#3b82f6'; this.style.boxShadow='0 0 10px rgba(59,130,245,0.3)'" 
           onmouseout="this.style.borderColor='#334155'; this.style.boxShadow='none'"
           onclick="selezionaImmagineRicerca(${idx}, '${ris.url}', '${ris.b64.substring(0, 50)}...', ${ridx})">
        <div style="width:100%; aspect-ratio:1; background:#0a0f23; overflow:hidden; display:flex; align-items:center; justify-content:center;">
          <img src="data:image/jpeg;base64,${ris.b64}" style="width:100%; height:100%; object-fit:cover;" />
        </div>
        <div style="padding:8px; background:rgba(59,130,245,0.1); border-top:1px solid #334155;">
          <div style="font-size:8px; color:#9ca3af; margin-bottom:2px;">${ris.fonte}</div>
          <button style="width:100%; padding:4px; background:#10b981; color:white; border:none; border-radius:3px; font-size:9px; cursor:pointer; font-weight:600;">✓ Seleziona</button>
        </div>
      </div>
    `).join('');
    
    gridEl.innerHTML = html;
  })
  .catch(e => {
    console.error('❌ ERRORE ricerca:', e);
    const statusEl = document.getElementById('ricerca-status');
    statusEl.innerHTML = '❌ Errore: ' + e.message;
  });
}

function selezionaImmagineRicerca(idx, urlOriginale, b64Preview, ridx) {
  console.log('✅ Immagine selezionata:', ridx);
  
  // Aggiorna anteprima
  const anteprimaDiv = document.getElementById('anteprima-immagine');
  anteprimaDiv.innerHTML = '<div style="color:#93c5fd; text-align:center;">⏳ Salvataggio in corso...</div>';
  
  // Salva nel backend
  fetch('/api/prodotti/' + encodeURIComponent(window._gridBrandStanza) + '/' + encodeURIComponent(window._gridProdottiStanza[idx].codice) + '/immagine', {
    method: 'PUT',
    headers: {'Content-Type': 'application/json'},
    body: JSON.stringify({
      immagine_url: urlOriginale,
      immagine_b64: b64Preview  // Per DB (completo viene già salvato dal backend)
    })
  })
  .then(r => r.json())
  .then(d => {
    if (d.ok) {
      // Aggiorna anteprima con l'immagine vera
      anteprimaDiv.innerHTML = `<img src="${urlOriginale}" style="width:100%; height:100%; object-fit:cover;" />`;
      
      // Aggiorna prodotto in memoria
      window._gridProdottiStanza[idx].immagine_url = urlOriginale;
      
      alert('✓ Immagine salvata con successo!');
    } else {
      anteprimaDiv.innerHTML = '<div style="color:#ef4444;">❌ Errore nel salvataggio</div>';
      alert('❌ ' + (d.error || 'Errore'));
    }
  })
  .catch(e => {
    anteprimaDiv.innerHTML = '<div style="color:#ef4444;">❌ ' + e.message + '</div>';
  });
}

function salvaURLImmagineManuale(idx) {
  const urlInput = document.getElementById('url-immagine-input');
  const urlImmagine = urlInput.value.trim();
  
  if (!urlImmagine) {
    alert('❌ Inserisci un URL valido');
    return;
  }
  
  if (!window._gridProdottiStanza) return;
  
  const prodotto = window._gridProdottiStanza[idx];
  
  // Salva URL manuale
  fetch('/api/prodotti/' + encodeURIComponent(window._gridBrandStanza) + '/' + encodeURIComponent(prodotto.codice) + '/immagine', {
    method: 'PUT',
    headers: {'Content-Type': 'application/json'},
    body: JSON.stringify({
      immagine_url: urlImmagine
    })
  })
  .then(r => r.json())
  .then(d => {
    if (d.ok) {
      const anteprima = document.getElementById('anteprima-immagine');
      anteprima.innerHTML = `<img src="${urlImmagine}" style="width:100%; height:100%; object-fit:cover;" />`;
      window._gridProdottiStanza[idx].immagine_url = urlImmagine;
      alert('✓ Immagine salvata con successo!');
    } else {
      alert('❌ ' + (d.error || 'Errore'));
    }
  })
  .catch(e => {
    alert('❌ Errore: ' + e.message);
  });
}

function salvaURLImmagine(idx) {
  salvaURLImmagineManuale(idx);  // Alias per compatibilità
}

function chiudiModaleImmagine() {
  const modal = document.getElementById('modal-immagine');
  if (modal) {
    modal.remove();
  }
}

// ============================================================================
// MODALE ABBINAMENTI RICCO CON IMMAGINI
// ============================================================================

function apriModaleAbbinamenti(idx) {
  if (!window._gridProdottiStanza) return;
  
  const prodotto = window._gridProdottiStanza[idx];
  const brand = window._gridBrandStanza;
  
  // Salva state per il modale
  window._prodottoSelezionatoPerAbbinamenti = prodotto;
  window._abbinamenti_selezionati = [];
  
  // Carica abbinamenti da API
  fetch('/api/abbinamenti/' + encodeURIComponent(brand) + '/' + encodeURIComponent(prodotto.codice))
    .then(r => r.json())
    .then(d => {
      const abbinamenti = (d.ok && d.abbinamenti) ? d.abbinamenti : [];
      renderModaleAbbinamenti(prodotto, abbinamenti, brand);
    })
    .catch(e => {
      renderModaleAbbinamenti(prodotto, [], brand);
    });
}

function renderModaleAbbinamenti(prodotto, abbinamenti, brand) {
  const modalHtml = `
    <div id="modal-abbinamenti" style="position:fixed; top:0; left:0; width:100%; height:100%; background:rgba(0,0,0,0.85); z-index:7000; display:flex; flex-direction:column; padding:0; overflow:hidden;">
      
      <!-- HEADER -->
      <div style="background:#0f172e; border-bottom:2px solid #3b82f6; padding:16px 20px; display:flex; justify-content:space-between; align-items:center; flex-shrink:0;">
        <div style="font-size:14px; font-weight:bold; color:#60a5fa;">🔗 Abbinamenti per: <strong>${prodotto.nome}</strong></div>
        <button onclick="chiudiModaleAbbinamenti()" style="background:#ef4444; color:white; border:none; padding:8px 16px; border-radius:6px; cursor:pointer; font-weight:600;">✕ Chiudi</button>
      </div>
      
      <!-- CONTENUTO -->
      <div style="display:flex; gap:20px; flex:1; overflow:hidden; padding:20px;">
        
        <!-- SINISTRA: PRODOTTO PRINCIPALE -->
        <div style="flex:0 0 300px; display:flex; flex-direction:column; background:rgba(30,41,59,0.6); border:1px solid #334155; border-radius:8px; padding:16px; overflow-y:auto;">
          <div style="font-size:12px; font-weight:bold; color:#60a5fa; margin-bottom:12px;">Prodotto Principale</div>
          
          <!-- Immagine prodotto -->
          <div style="width:100%; aspect-ratio:1; background:rgba(59,130,245,0.1); border-radius:6px; display:flex; align-items:center; justify-content:center; margin-bottom:12px; overflow:hidden;">
            ${prodotto.immagine_url ? 
              `<img src="${prodotto.immagine_url}" style="width:100%; height:100%; object-fit:cover;" />` 
              : 
              `<div style="color:#6b7280; font-size:60px;">📦</div>`
            }
          </div>
          
          <!-- Info prodotto -->
          <div style="font-size:11px; color:#d1d5db; margin-bottom:4px;">
            <div style="font-weight:bold; margin-bottom:4px;">Codice:</div>
            <div style="color:#9ca3af; font-family:monospace;">${prodotto.codice}</div>
          </div>
          
          <div style="font-size:11px; color:#d1d5db; margin-bottom:4px;">
            <div style="font-weight:bold; margin-bottom:4px;">Nome:</div>
            <div style="color:#9ca3af;">${prodotto.nome}</div>
          </div>
          
          <div style="font-size:11px; color:#d1d5db; margin-bottom:4px;">
            <div style="font-weight:bold; margin-bottom:4px;">Descrizione:</div>
            <div style="color:#9ca3af; font-size:10px;">${prodotto.descrizione || '—'}</div>
          </div>
          
          <div style="font-size:14px; color:#10b981; font-weight:bold; margin-top:12px; padding-top:12px; border-top:1px solid #334155;">
            €${prodotto.prezzo ? parseFloat(prodotto.prezzo).toFixed(0) : '—'}
          </div>
          
          <!-- ARRICCHIMENTO DESCRIZIONE -->
          <div style="margin-top:16px; padding-top:16px; border-top:1px solid #334155;">
            <div style="font-size:11px; font-weight:bold; color:#60a5fa; margin-bottom:8px;">🎨 Arricchisci Descrizione</div>
            <textarea id="desc-arricchita" placeholder="Aggiungi note, colori, dettagli..." style="width:100%; height:80px; padding:8px; background:rgba(30,41,59,0.8); border:1px solid #334155; color:white; border-radius:4px; font-size:10px; resize:none; font-family:monospace;"></textarea>
            <button onclick="generaDescrizioneIA()" style="width:100%; margin-top:8px; padding:8px; background:#8b5cf6; color:white; border:none; border-radius:4px; cursor:pointer; font-size:10px; font-weight:600;">✨ Genera con IA</button>
          </div>
        </div>
        
        <!-- DESTRA: ABBINAMENTI GRID -->
        <div style="flex:1; display:flex; flex-direction:column; overflow:hidden;">
          <div style="font-size:12px; font-weight:bold; color:#60a5fa; margin-bottom:12px;">Abbinamenti Disponibili (${abbinamenti.length})</div>
          
          <div id="abbinamenti-grid" style="flex:1; overflow-y:auto; display:grid; grid-template-columns:repeat(auto-fill, minmax(160px, 1fr)); gap:12px; padding-right:12px;">
            ${abbinamenti.length === 0 ? 
              `<div style="grid-column:1/-1; color:#6b7280; text-align:center; padding:40px;">Nessun abbinamento disponibile</div>` 
              : 
              abbinamenti.map((a, aidx) => `
                <div id="abbinamento-${aidx}" style="background:#1e293b; border:2px solid #334155; border-radius:6px; padding:10px; cursor:pointer; transition:all 0.2s;" 
                     data-codice="${a.codice || ''}" 
                     data-nome="${a.nome || ''}" 
                     data-prezzo="${a.prezzo || 0}"
                     onclick="toggleAbbinamento(${aidx})" 
                     onmouseover="this.style.borderColor='#3b82f6'" 
                     onmouseout="this.style.borderColor='#334155'">
                  
                  <!-- Immagine abbinamento -->
                  <div style="width:100%; aspect-ratio:1; background:rgba(59,130,245,0.1); border-radius:4px; display:flex; align-items:center; justify-content:center; margin-bottom:8px; overflow:hidden;">
                    ${a.immagine_url ? 
                      `<img src="${a.immagine_url}" style="width:100%; height:100%; object-fit:cover;" />` 
                      : 
                      `<div style="color:#6b7280; font-size:30px;">📎</div>`
                    }
                  </div>
                  
                  <!-- Checkbox -->
                  <div style="display:flex; align-items:center; margin-bottom:6px;">
                    <input type="checkbox" id="check-${aidx}" style="width:16px; height:16px; cursor:pointer;" 
                           onchange="toggleAbbinamento(${aidx})">
                    <label for="check-${aidx}" style="flex:1; margin-left:6px; font-size:10px; font-weight:bold; color:#d1d5db; cursor:pointer;">Seleziona</label>
                  </div>
                  
                  <!-- Info -->
                  <div style="font-size:9px; color:#9ca3af; margin-bottom:2px;">${a.codice || '—'}</div>
                  <div style="font-size:9px; color:#d1d5db; font-weight:bold; margin-bottom:4px; line-height:1.2;">${a.nome || '—'}</div>
                  <div style="font-size:10px; color:#10b981; font-weight:bold;">€${a.prezzo ? parseFloat(a.prezzo).toFixed(0) : '—'}</div>
                </div>
              `).join('')
            }
          </div>
        </div>
      </div>
      
      <!-- FOOTER -->
      <div style="background:#0f172e; border-top:2px solid #3b82f6; padding:16px 20px; display:flex; gap:12px; justify-content:flex-end; flex-shrink:0;">
        <button onclick="chiudiModaleAbbinamenti()" style="padding:10px 20px; background:#6b7280; color:white; border:none; border-radius:6px; cursor:pointer; font-weight:600;">✕ Annulla</button>
        <button onclick="event.preventDefault(); event.stopPropagation(); salvaConAbbinamenti();" style="padding:10px 20px; background:#10b981; color:white; border:none; border-radius:6px; cursor:pointer; font-weight:600;">✓ Aggiungi alla Stanza</button>
      </div>
    </div>
  `;
  
  document.body.insertAdjacentHTML('beforeend', modalHtml);
}

function toggleAbbinamento(idx) {
  console.log('🔵 toggleAbbinamento() chiamata con idx:', idx);
  
  const checkbox = document.getElementById('check-' + idx);
  const card = document.getElementById('abbinamento-' + idx);
  
  console.log('✓ Checkbox trovato:', checkbox);
  console.log('✓ Card trovata:', card);
  console.log('✓ Card dataset:', card ? card.dataset : 'NULL');
  
  if (!window._abbinamenti_selezionati) {
    window._abbinamenti_selezionati = [];
  }
  
  if (checkbox.checked) {
    // Salva i DATI veri, non solo l'indice
    const codice = card.dataset.codice || '';
    const nome = card.dataset.nome || '';
    const prezzo = card.dataset.prezzo || 0;
    
    console.log('✅ AGGIUNTO abbinamento:', {codice, nome, prezzo});
    
    // Evita duplicati
    if (!window._abbinamenti_selezionati.find(a => a.codice === codice)) {
      window._abbinamenti_selezionati.push({
        codice: codice,
        nome: nome,
        prezzo: prezzo
      });
    }
    
    card.style.borderColor = '#10b981';
    card.style.background = 'rgba(16,185,129,0.1)';
  } else {
    // Rimuovi per codice
    const codice = card.dataset.codice || '';
    window._abbinamenti_selezionati = window._abbinamenti_selezionati.filter(a => a.codice !== codice);
    
    console.log('❌ RIMOSSO abbinamento:', codice);
    
    card.style.borderColor = '#334155';
    card.style.background = '#1e293b';
  }
  
  console.log('📊 State attuale window._abbinamenti_selezionati:', window._abbinamenti_selezionati);
}

function generaDescrizioneIA() {
  const prodotto = window._prodottoSelezionatoPerAbbinamenti;
  if (!prodotto) return;
  
  const textarea = document.getElementById('desc-arricchita');
  textarea.value = '⏳ Generazione in corso...';
  textarea.disabled = true;
  
  // Chiama OpenAI API per arricchire descrizione
  fetch('/api/arricchisci-descrizione', {
    method: 'POST',
    headers: {'Content-Type': 'application/json'},
    body: JSON.stringify({
      codice: prodotto.codice,
      nome: prodotto.nome,
      descrizione_attuale: prodotto.descrizione || ''
    })
  })
  .then(r => r.json())
  .then(d => {
    if (d.ok && d.descrizione_arricchita) {
      textarea.value = d.descrizione_arricchita;
    } else {
      textarea.value = prodotto.descrizione || '';
    }
    textarea.disabled = false;
  })
  .catch(e => {
    textarea.value = prodotto.descrizione || '';
    textarea.disabled = false;
  });
}

function salvaConAbbinamenti() {
  console.log('🟢 salvaConAbbinamenti() INIZIATA');
  
  const prodotto = window._prodottoSelezionatoPerAbbinamenti;
  const stanzaId = window._gridStanzaId;
  const brand = window._gridBrandStanza;
  
  console.log('📦 Prodotto:', prodotto);
  console.log('🏠 StanzaId:', stanzaId);
  console.log('🏢 Brand:', brand);
  
  if (!prodotto || !stanzaId) {
    console.error('❌ ERRORE: Prodotto o StanzaId mancanti');
    return;
  }
  
  const descArricchita = document.getElementById('desc-arricchita').value.trim() || 
                         ((prodotto.codice ? '[' + prodotto.codice + '] ' : '') + (prodotto.nome || ''));
  
  // ✅ USA GLI ABBINAMENTI SELEZIONATI (da toggleAbbinamento)
  const abbinamenti_list = window._abbinamenti_selezionati || [];
  
  console.log('🔗 Abbinamenti da salvare:', abbinamenti_list);
  console.log('📊 Numero abbinamenti:', abbinamenti_list.length);
  
  const prezzo = prodotto.prezzo || 0;
  
  const payload = {
    codice: prodotto.codice || '',
    brand: brand,
    descrizione: descArricchita,
    quantita: 1,
    prezzo_unitario: prezzo,
    sconto_percentuale: 0,
    colore: 'verde',
    immagine_url: prodotto.immagine_url || '',
    abbinamenti_selezionati: abbinamenti_list
  };
  
  console.log('📤 Payload inviato:', payload);
  
  // Salva nella stanza
  fetch('/api/stanze/' + stanzaId + '/voci', {
    method: 'POST',
    headers: {'Content-Type': 'application/json'},
    body: JSON.stringify(payload)
  })
  .then(r => r.json())
  .then(d => {
    console.log('📥 Risposta backend:', d);
    
    if (d.ok) {
      chiudiModaleAbbinamenti();
      loadInterfacciaPiani(window.cantiere_attivo_id);
      alert('✓ Aggiunto con ' + abbinamenti_list.length + ' abbinamenti');
    } else {
      alert('❌ ' + (d.error || 'Errore'));
    }
  })
  .catch(e => {
    console.error('❌ Errore fetch:', e);
    alert('❌ Errore: ' + e.message);
  });
}

function chiudiModaleAbbinamenti() {
  console.log('🔴 chiudiModaleAbbinamenti() INIZIATA');
  
  const modal = document.getElementById('modal-abbinamenti');
  if (modal) {
    modal.style.display = 'none';  // Nascondi prima
    modal.remove();  // Poi rimuovi
    console.log('✅ Modale rimosso');
  }
  
  window._prodottoSelezionatoPerAbbinamenti = null;
  window._abbinamenti_selezionati = [];
  
  console.log('✅ State resettato');
}

function aggiungiAlCarrello(idx) {
  if (!window._gridProdottiStanza) return;
  
  const prodotto = window._gridProdottiStanza[idx];
  const stanzaId = window._gridStanzaId;
  const brand = window._gridBrandStanza;
  
  const descrizione = (prodotto.codice ? '[' + prodotto.codice + '] ' : '') + (prodotto.nome || '');
  const prezzo = prodotto.prezzo || 0;
  
  // POST al carrello (stesso endpoint stanze/voci per adesso)
  fetch('/api/stanze/' + stanzaId + '/voci', {
    method: 'POST',
    headers: {'Content-Type': 'application/json'},
    body: JSON.stringify({
      codice: prodotto.codice || '',
      brand: brand,
      descrizione: descrizione,
      quantita: 1,
      prezzo_unitario: prezzo,
      sconto_percentuale: 0,
      colore: 'verde',
      immagine_url: prodotto.immagine_url || '',
      flag_carrello: true
    })
  })
  .then(r => r.json())
  .then(d => {
    if (d.ok) {
      alert('🛒 Aggiunto al carrello: ' + (prodotto.nome || 'Prodotto'));
    } else {
      alert('❌ ' + (d.error || 'Errore'));
    }
  });
}

function chiudiGridStanza() {
  const modal = document.getElementById('modal-grid-stanza');
  if (modal) {
    modal.remove();
    // Ricarica la struttura piani
    loadInterfacciaPiani(cantiereAttivo);
  }
}

function cancellaVoce(voceId, stanzaId) {
  if (!confirm('Eliminare questa voce?')) return;

  fetch('/api/stanza_voci/' + voceId, { method: 'DELETE' })
    .then(r => r.json())
    .then(d => {
      if (d.ok) {
        loadInterfacciaPiani(cantiereAttivo);
      } else {
        alert('❌ ' + (d.error || 'Errore'));
      }
    });
}

// =============================================================================
// LISTINO DINAMICO INTERFACCIA 2 (PIANI)
// =============================================================================

let listinoStorePiani = [];
let stanzaSelezionataPiani = null;

function aggiungiVoceStanzaFromListino(prodotto) {
  if (!stanzaSelezionataPiani) {
    alert('Seleziona prima una stanza cliccando "+ Voce"');
    return;
  }

  const brand = selected.length > 0 ? selected[0] : 'Gessi';
  const descrizione = '[' + (prodotto.codice || '—') + '] ' + (prodotto.nome || prodotto.descrizione || '');
  const prezzo = prodotto.prezzo || 0;

  fetch('/api/stanze/' + stanzaSelezionataPiani.id + '/voci', {
    method: 'POST',
    headers: {'Content-Type': 'application/json'},
    body: JSON.stringify({
      codice: prodotto.codice || '',
      brand: brand,
      descrizione: descrizione,
      quantita: 1,
      prezzo_unitario: prezzo,
      sconto_percentuale: 0,
      colore: 'verde'
    })
  })
  .then(r => r.json())
  .then(d => {
    if (d.ok) {
      loadInterfacciaPiani(cantiereAttivo);
      document.getElementById('listino-piani-search').value = '';
    } else {
      alert('❌ ' + (d.error || 'Errore'));
    }
  });
}

function filtraListinoPiani() {
  const sv = document.getElementById('listino-piani-search').value.toLowerCase();
  const grid = document.getElementById('listino-piani-grid');
  
  if (!stanzaSelezionataPiani) {
    grid.innerHTML = '<div style="color:#6b7280; font-size:10px;">Seleziona una stanza per caricare il listino</div>';
    return;
  }

  if (listinoStorePiani.length === 0) {
    grid.innerHTML = '<div style="color:#6b7280; font-size:10px;">⏳ Caricamento listino...</div>';
    
    // Carica listino del brand selezionato
    const brand = selected.length > 0 ? selected[0] : 'Gessi';
    fetch('/api/listino/' + encodeURIComponent(brand))
      .then(r => r.json())
      .then(d => {
        listinoStorePiani = d.prodotti || [];
        filtraListinoPiani(); // Ricorsivo per renderizzare
      })
      .catch(e => {
        grid.innerHTML = '<div style="color:#ef4444; font-size:10px;">❌ ' + e.message + '</div>';
      });
    return;
  }

  let filtered = listinoStorePiani;
  if (sv) {
    filtered = listinoStorePiani.filter(p =>
      (p.nome || '').toLowerCase().includes(sv) ||
      (p.codice || '').toLowerCase().includes(sv) ||
      (p.descrizione || '').toLowerCase().includes(sv)
    );
  }

  if (filtered.length === 0) {
    grid.innerHTML = '<div style="color:#6b7280; font-size:10px;">Nessun prodotto trovato</div>';
    return;
  }

  grid.innerHTML = filtered.slice(0, 20).map(p => {
    const prezzo = p.prezzo || 0;
    const colore = prezzo > 500 ? '#10b981' : '#3b82f6';
    return '<div style="background:rgba(30,41,59,0.9); border:1px solid rgba(59,130,245,0.2); border-radius:4px; padding:6px; cursor:pointer;" onclick="aggiungiVoceStanzaFromListino(this.parentElement.dataset.prod)" data-prod=\'' + JSON.stringify(p).replace(/'/g, "\\'") + '\'>' +
      '<div style="font-size:10px; font-weight:600; color:#e0e0e0;">' + (p.nome || p.codice) + '</div>' +
      '<div style="font-size:9px; color:#9ca3af;">[' + (p.codice || '—') + ']</div>' +
      '<div style="font-size:11px; color:' + colore + '; font-weight:bold; margin-top:3px;">€' + parseFloat(prezzo).toFixed(0) + '</div>' +
      '</div>';
  }).join('');
}

// =============================================================================
// CARICAMENTO PLANIMETRIA CON OPENAI VISION
// =============================================================================

async function caricaPlanimetria(input) {
  const file = input.files[0];
  if (!file) return;

  const statusEl = document.getElementById('planimetria-status');
  statusEl.textContent = '⏳ Caricamento...';
  statusEl.style.color = '#93c5fd';

  // Converti immagine a base64
  const reader = new FileReader();
  reader.onload = async (e) => {
    const base64 = e.target.result;

    try {
      statusEl.textContent = '🤖 Analizzando disegno...';

      // Chiama endpoint che userà DeepSeek Vision
      const resp = await fetch('/api/analizza-planimetria', {
        method: 'POST',
        headers: {'Content-Type': 'application/json'},
        body: JSON.stringify({
          cantiere_id: cantiereAttivo,
          immagine_base64: base64
        })
      });

      const data = await resp.json();

      if (!data.ok) {
        statusEl.textContent = '❌ ' + (data.error || 'Errore');
        statusEl.style.color = '#ef4444';
        input.value = '';
        return;
      }

      // CREAZIONE AUTOMATICA
      statusEl.textContent = `✅ ${data.piani_creati} piani, ${data.stanze_create} stanze create!`;
      statusEl.style.color = '#10b981';
      input.value = '';

      // Ricarica interfaccia
      setTimeout(() => {
        loadInterfacciaPiani(cantiereAttivo);
      }, 1500);

    } catch (err) {
      statusEl.textContent = '❌ ' + err.message;
      statusEl.style.color = '#ef4444';
      input.value = '';
    }
  };

  reader.readAsDataURL(file);
}

// ---------------------------------------------------------------------------
// LISTINO DASHBOARD — flusso 1-click
// ---------------------------------------------------------------------------
let listinoData = [];
let listinoBrand = '';
let filtroLinea = 'tutti';
let filtroCategoria = 'tutti';
let listinoTipo = 'cliente';

function setListinoTipo(tipo) {
  listinoTipo = tipo;
  document.getElementById('btn-listino-cliente').style.background = tipo === 'cliente' ? '#3b82f6' : 'rgba(59,130,245,0.2)';
  document.getElementById('btn-listino-riv').style.background = tipo === 'rivenditore' ? '#f59e0b' : 'rgba(245,158,11,0.2)';
  document.getElementById('btn-listino-riv').style.color = tipo === 'rivenditore' ? 'white' : '#f59e0b';
  filtraListino();
}

// Apre listino automaticamente quando brand selezionato ha un listino caricato
function apriListino() {
  if (!selected || selected.length === 0) { alert('Seleziona prima un brand'); return; }
  listinoBrand = selected[0];
  filtroLinea = 'tutti';
  filtroCategoria = 'tutti';
  document.getElementById('listino-brand-tag').textContent = listinoBrand;
  document.getElementById('listino-panel').classList.add('open');
  document.getElementById('listino-search').value = '';
  caricaListino();
}

function chiudiListino() {
  document.getElementById('listino-panel').classList.remove('open');
}

function caricaListino() {
  document.getElementById('prodotti-grid').innerHTML = '<div style="color:#6b7280;font-size:11px;padding:20px 0;">Caricamento prodotti...</div>';
  fetch('/api/listino/' + encodeURIComponent(listinoBrand))
    .then(r => r.json())
    .then(d => {
      listinoData = d.prodotti || [];
      costruisciFiltriLinea();
      costruisciFiltriCat();
      costruisciDomande();
      filtraListino();
      const n = listinoData.length;
      const fonte = d.fonte
        ? '<span style="color:#10b981;">📄 ' + n + ' prodotti da listino</span>'
        : '<span style="color:#ef4444;">⚠ Nessun listino caricato per ' + listinoBrand + ' — carica un Excel prima</span>';
      document.getElementById('listino-count').innerHTML = fonte;
      
      // CARICA ABBINAMENTI AUTOMATICAMENTE QUANDO CARICHI LISTINO
      fetch('/api/carica-abbinamenti-excel/' + encodeURIComponent(listinoBrand), {method:'POST'})
        .then(r => r.json())
        .catch(() => {}); // Silenzioso se fallisce
    });
}

function costruisciFiltriLinea() {
  const linee = [...new Set(listinoData.map(p => p.collezione).filter(Boolean))].sort();
  let html = '<button class="filtro-btn active" onclick="setFiltroLinea(\'tutti\',this)">Tutte le linee</button>';
  linee.forEach(l => {
    const n = listinoData.filter(p => p.collezione === l).length;
    html += '<button class="filtro-btn" onclick="setFiltroLinea(\'' + l.replace(/'/g,"\\'") + '\',this)">' + l + ' <span style="opacity:0.6;">(' + n + ')</span></button>';
  });
  document.getElementById('filtri-linea').innerHTML = html;
}

function costruisciFiltriCat() {
  const cats = [...new Set(listinoData.map(p => p.categoria).filter(Boolean))].sort();
  let html = '<button class="filtro-btn active" onclick="setFiltroCategoria(\'tutti\',this)">Tutte le cat.</button>';
  cats.forEach(c => {
    html += '<button class="filtro-btn" onclick="setFiltroCategoria(\'' + c.replace(/'/g,"\\'") + '\',this)">' + c + '</button>';
  });
  document.getElementById('filtri-cat').innerHTML = html;
}

function costruisciDomande() {
  const domande = [
    'Quali prodotti abbinare per un bagno moderno?',
    'Cosa è disponibile subito?',
    'Prodotti entry level sotto €400',
    'Soluzioni premium per progetto di lusso',
    'Cosa abbinare a ' + listinoBrand + ' per doccia?',
    'Differenze tra le collezioni',
    'Novità e best seller'
  ];
  document.getElementById('domande-bar').innerHTML = domande.map(d =>
    '<button class="domanda-chip" onclick="faiDomandaListino(\'' + d.replace(/'/g,"\\'") + '\')">' + d + '</button>'
  ).join('');
}

function setFiltroLinea(val, btn) {
  filtroLinea = val;
  document.querySelectorAll('#filtri-linea .filtro-btn').forEach(b => b.classList.remove('active'));
  btn.classList.add('active');
  filtraListino();
}

function setFiltroCategoria(val, btn) {
  filtroCategoria = val;
  document.querySelectorAll('#filtri-cat .filtro-btn').forEach(b => b.classList.remove('active'));
  btn.classList.add('active');
  filtraListino();
}

function filtraListino() {
  const sv = (document.getElementById('listino-search') ? document.getElementById('listino-search').value : '').toLowerCase();
  let prodotti = listinoData;
  if (filtroLinea !== 'tutti') prodotti = prodotti.filter(p => p.collezione === filtroLinea);
  if (filtroCategoria !== 'tutti') prodotti = prodotti.filter(p => p.categoria === filtroCategoria);
  if (sv) prodotti = prodotti.filter(p =>
    (p.nome||'').toLowerCase().includes(sv) ||
    (p.codice||'').toLowerCase().includes(sv) ||
    (p.descrizione||'').toLowerCase().includes(sv)
  );

  if (prodotti.length === 0) {
    document.getElementById('prodotti-grid').innerHTML = '<div style="color:#6b7280;font-size:11px;padding:20px 0;">Nessun prodotto trovato</div>';
    return;
  }

  document.getElementById('prodotti-grid').innerHTML = prodotti.map(p => {
    const idx = listinoData.indexOf(p);
    const dispClass = (p.disponibilita||'').toLowerCase().includes('ordine') ? 'su-ordine' : 'disponibile';
    const dispBadge = (p.disponibilita||'').toLowerCase().includes('ordine') ? 'disp-ord' : 'disp-ok';
    const dispLabel = (p.disponibilita||'').toLowerCase().includes('ordine') ? '⏳ Su ordine' : '✓ Disponibile';

    // Prezzo principale + secondario
    const pCliente = p.prezzo;
    const pRiv = p.prezzo_rivenditore;
    const pUsato = listinoTipo === 'rivenditore' && pRiv ? pRiv : pCliente;
    const pSecondario = listinoTipo === 'rivenditore' ? pCliente : pRiv;
    const clsPrezzo = p.fonte === 'excel' ? 'prodotto-prezzo-excel' : 'prodotto-prezzo-web';
    const iconWeb = p.fonte !== 'excel' ? ' ⚠' : '';

    let prezzoHtml = pUsato
      ? '<span class="' + clsPrezzo + '">€' + parseFloat(pUsato).toFixed(0) + iconWeb + '</span>'
      : '<span style="color:#6b7280;font-size:10px;">—</span>';
    if (pSecondario) {
      const lbl = listinoTipo === 'cliente' ? 'riv.' : 'cl.';
      const clr = listinoTipo === 'cliente' ? '#f59e0b' : '#9ca3af';
      prezzoHtml += '<span style="font-size:9px;color:' + clr + ';margin-left:6px;">' + lbl + ' €' + parseFloat(pSecondario).toFixed(0) + '</span>';
    }

    return '<div class="prodotto-card ' + dispClass + '" id="pcard-' + idx + '">' +
      '<div style="display:flex;justify-content:space-between;align-items:flex-start;margin-bottom:3px;">' +
      '<span class="prodotto-codice">' + (p.codice||'—') + '</span>' +
      '<span class="prodotto-disp ' + dispBadge + '">' + dispLabel + '</span>' +
      '</div>' +
      '<div class="prodotto-nome">' + (p.nome||'—') + '</div>' +
      '<div class="prodotto-cat">' + [p.collezione, p.categoria].filter(Boolean).join(' · ') + '</div>' +
      (p.descrizione ? '<div style="font-size:10px;color:#9ca3af;margin:3px 0;line-height:1.3;">' + p.descrizione + '</div>' : '') +
      (p.finiture ? '<div style="font-size:9px;color:#6b7280;">🎨 ' + p.finiture + '</div>' : '') +
      '<div style="display:flex;justify-content:space-between;align-items:center;margin-top:6px;">' +
      '<div>' + prezzoHtml + '</div>' +
      '</div>' +
      '<div class="prodotto-actions" style="margin-top:8px;">' +
      '<button onclick="event.stopPropagation();verificaAbbinamenti(' + idx + ',\'' + (p.codice||'').replace(/'/g,"\\'") + '\')" class="btn-sm" id="abbina-btn-' + idx + '" style="flex:1; background:rgba(107,114,128,0.3);color:#d1d5db;">📋 Abbina</button>' +
      '<button onclick="event.stopPropagation();chiediAIprodotto(' + idx + ',\'descrizione\')" class="btn-sm" style="background:rgba(139,92,246,0.2);color:#a78bfa;flex:1;">✍ Arricchisci</button>' +
      '<button onclick="event.stopPropagation();aggiungiDaListino(' + idx + ')" class="btn-sm btn-green" style="flex:1;" id="addbtn-' + idx + '">+ Carrello</button>' +
      '</div></div>';
  }).join('');
  
  // Controlla abbinamenti per TUTTI i prodotti visibili
  prodotti.forEach((p, i) => {
    const idx = listinoData.indexOf(p);
    setTimeout(() => verificaAbbinamenti(idx, p.codice), 100 * i);
  });
}

function chiediAIprodotto(idx, tipo) {
  const p = listinoData[idx];
  if (!p) return;
  let domanda = '';
  if (tipo === 'abbinamenti') domanda = 'Cosa abbinare al prodotto ' + (p.nome||p.codice) + ' di ' + listinoBrand + '? Suggerisci prodotti complementari dello stesso brand o di altri brand che trattiamo.';
  if (tipo === 'descrizione') domanda = 'Crea una descrizione commerciale breve (max 120 caratteri) per: ' + (p.nome||p.codice) + ' — ' + (p.descrizione||'') + '. Prezzo: €' + (p.prezzo||'da definire') + '. Brand: ' + listinoBrand;
  chiudiListino();
  document.getElementById('question').value = domanda;
  ask();
}

function faiDomandaListino(domanda) {
  chiudiListino();
  document.getElementById('question').value = domanda;
  ask();
}

function aggiungiAccessorio(accId, accNome) {
  if (!cantiereAttivo) {
    alert('Nessun cantiere aperto');
    return;
  }
  
  fetch('/api/cantieri/' + cantiereAttivo + '/righe', {
    method: 'POST',
    headers: {'Content-Type': 'application/json'},
    body: JSON.stringify({ 
      brand: 'Gessi', 
      categoria: 'Accessori', 
      descrizione: accNome,
      importo: 0 // Accessori senza prezzo iniziale
    })
  })
  .then(r => r.json())
  .then(d => {
    if (d.ok) {
      loadRighe(); // Ricarica carrello
      // Feedback visuale
      const msg = document.createElement('div');
      msg.innerHTML = '✓ Accessorio aggiunto!';
      msg.style.cssText = 'position:fixed;top:20px;right:20px;background:#10b981;color:white;padding:12px 20px;border-radius:6px;z-index:9999;font-size:12px;';
      document.body.appendChild(msg);
      setTimeout(() => msg.remove(), 2000);
    }
  });
}

function aggiungiDaListino(idx) {
  // 🆕 SE SIAMO IN MODALITA' STANZA, AGGIUNG DIRETTAMENTE ALLA STANZA
  if (window._modalitaStanza && window._stanzaAggiunta) {
    aggiungiProdottoAllaStanza(idx);
    return;
  }
  
  // ALTRIMENTI: comportamento NORMALE (aggiungi al carrello cantieri)
  if (!cantiereAttivo) {
    if (confirm('Nessun cantiere aperto. Vuoi aprire il pannello cantieri?')) {
      chiudiListino();
      document.getElementById('mod-cantieri') && toggleModule('cantieri-body');
    }
    return;
  }
  const p = listinoData[idx];
  if (!p) return;
  const descrizione = (p.codice ? '[' + p.codice + '] ' : '') + (p.nome || p.descrizione || '');
  const importo = listinoTipo === 'rivenditore' && p.prezzo_rivenditore ? p.prezzo_rivenditore : (p.prezzo || 0);

  const btn = document.getElementById('addbtn-' + idx);
  if (btn) { btn.textContent = '⏳'; btn.disabled = true; }

  fetch('/api/cantieri/' + cantiereAttivo + '/righe', {
    method: 'POST',
    headers: {'Content-Type': 'application/json'},
    body: JSON.stringify({ brand: listinoBrand, categoria: p.categoria||'', descrizione, importo })
  })
  .then(r => r.json())
  .then(d => {
    if (d.ok) {
      if (btn) { btn.textContent = '✓ Aggiunto'; btn.style.background = '#10b981'; }
      const card = document.getElementById('pcard-' + idx);
      if (card) card.style.opacity = '0.6';
      loadRighe();
      // Carica accessori consigliati
      const prodottoId = p.codice || '';
      if (prodottoId) caricaAccessoriProdotto(prodottoId, listinoBrand);
    } else {
      if (btn) { btn.textContent = '+ Carrello'; btn.disabled = false; }
    }
    
    // Controlla abbinamenti e colora il bottone
    verificaAbbinamenti(idx, p.codice);
  });
}

function aggiungiProdottoAllaStanza(idx) {
  const p = listinoData[idx];
  if (!p || !window._stanzaAggiunta) return;
  
  const stanza = window._stanzaAggiunta;
  const descrizione = (p.codice ? '[' + p.codice + '] ' : '') + (p.nome || p.descrizione || '');
  const prezzo = listinoTipo === 'rivenditore' && p.prezzo_rivenditore ? p.prezzo_rivenditore : (p.prezzo || 0);
  
  const btn = document.getElementById('addbtn-' + idx);
  if (btn) { btn.textContent = '⏳'; btn.disabled = true; }
  
  // INVIA DIRETTAMENTE AL BACKEND STANZA
  fetch('/api/stanze/' + stanza.id + '/voci', {
    method: 'POST',
    headers: {'Content-Type': 'application/json'},
    body: JSON.stringify({
      codice: p.codice || '',
      brand: listinoBrand,
      descrizione: descrizione,
      quantita: 1,
      prezzo_unitario: prezzo,
      sconto_percentuale: 0,
      colore: 'verde',
      immagine_url: p.immagine_url || ''
    })
  })
  .then(r => r.json())
  .then(d => {
    console.log('📦 Risposta aggiunta stanza:', d);
    if (d.ok) {
      if (btn) { btn.textContent = '✓ Aggiunto'; btn.style.background = '#10b981'; }
      const card = document.getElementById('pcard-' + idx);
      if (card) card.style.opacity = '0.6';
      
      // CHIUDI il listino e RICARICA la stanza
      chiudiListino();
      if (window.loadInterfacciaPiani) {
        loadInterfacciaPiani(window.cantiere_attivo_id);
      }
      alert('✓ Aggiunto a ' + stanza.nome);
    } else {
      if (btn) { btn.textContent = '+ Aggiungi'; btn.disabled = false; }
      alert('❌ ' + (d.error || 'Errore'));
    }
  })
  .catch(e => {
    console.error('Errore:', e);
    if (btn) { btn.textContent = '+ Aggiungi'; btn.disabled = false; }
    alert('❌ Errore: ' + e.message);
  });
}

function verificaAbbinamenti(idx, codice) {
  if (!codice) return;
  
  // Fetch diretto — controlla se abbinamenti sono nel DB
  fetch('/api/abbina/' + encodeURIComponent(codice))
    .then(r => r.json())
    .then(d => {
      const btn = document.getElementById('abbina-btn-' + idx);
      if (!btn) return;
      
      // Se ha abbinamenti ufficiali O alternative
      const hasAbbinamenti = (d.ufficiali && d.ufficiali.length > 0) || 
                             (d.alternative && d.alternative.length > 0);
      
      if (hasAbbinamenti) {
        // 🔴 ROSSO - Ha abbinamenti
        btn.style.background = '#ef4444';
        btn.style.color = 'white';
        btn.style.cursor = 'pointer';
        btn.disabled = false;
        btn.onclick = () => apriModalAbbinamenti(codice, d);
      } else {
        // 🔘 Grigio - No abbinamenti
        btn.style.background = 'rgba(107,114,128,0.3)';
        btn.style.color = '#d1d5db';
        btn.style.cursor = 'not-allowed';
        btn.disabled = true;
      }
    })
    .catch(e => {
      console.error('Errore verifica:', e);
      const btn = document.getElementById('abbina-btn-' + idx);
      if (btn) btn.style.background = 'rgba(107,114,128,0.3)';
    });
}

function apriModalAbbinamenti(codice, data) {
  if (!cantiereAttivo) {
    alert('Apri prima un cantiere');
    return;
  }

  // Aggiungi DIRETTAMENTE tutti gli abbinamenti ufficiali
  const abbinamenti = data.ufficiali || [];
  
  if (abbinamenti.length === 0) {
    alert('Nessun abbinamento disponibile');
    return;
  }

  console.log('🔵 Aggiungo', abbinamenti.length, 'abbinamenti');
  
  let aggiunto = 0;
  abbinamenti.forEach(acc => {
    fetch('/api/cantieri/' + cantiereAttivo + '/righe', {
      method: 'POST',
      headers: {'Content-Type': 'application/json'},
      body: JSON.stringify({
        brand: 'Gessi',
        categoria: 'Accessori',
        descrizione: `[${acc.accessorio_id}] ${acc.nome}`,
        importo: 0
      })
    })
    .then(r => r.json())
    .then(d => {
      if (d.ok) {
        aggiunto++;
        console.log(`✅ ${acc.nome}`);
        if (aggiunto === abbinamenti.length) {
          loadRighe();
          const msg = document.createElement('div');
          msg.textContent = `✓ ${aggiunto} abbinamenti aggiunti`;
          msg.style.cssText = 'position:fixed;top:20px;right:20px;background:#10b981;color:white;padding:12px 20px;border-radius:6px;z-index:9999;font-size:12px;font-weight:600;';
          document.body.appendChild(msg);
          setTimeout(() => msg.remove(), 2000);
        }
      } else {
        console.error('Errore:', acc.accessorio_id, d.error);
      }
    })
    .catch(e => console.error('Exception:', e));
  });
}

// Controlla se già loggato
fetch('/api/me').then(r => r.json()).then(d => {
  if (d.logged) {
    document.getElementById('login-overlay').style.display = 'none';
    document.getElementById('main-app').style.display = 'flex';
    initApp();
  }
});
</script>

<!-- ============================================================================
     MODAL ABBINAMENTI — HTML/CSS/JavaScript
     ============================================================================ -->

<style>
.modal-abbinamenti {
    display: none;
    position: fixed;
    top: 0;
    left: 0;
    width: 100%;
    height: 100%;
    background: rgba(0, 0, 0, 0.7);
    z-index: 10000;
    justify-content: center;
    align-items: center;
    padding: 20px;
    box-sizing: border-box;
}

.modal-abbinamenti.show {
    display: flex;
}

.modal-abbinamenti-content {
    background: #0f172a;
    border: 2px solid #3b82f6;
    border-radius: 12px;
    width: 100%;
    max-width: 1000px;
    max-height: 85vh;
    overflow-y: auto;
    padding: 30px;
    color: #e0e0e0;
}

.modal-header {
    display: flex;
    justify-content: space-between;
    align-items: center;
    margin-bottom: 25px;
    border-bottom: 2px solid #1e40af;
    padding-bottom: 15px;
}

.modal-header h2 {
    margin: 0;
    color: #60a5fa;
    font-size: 24px;
    font-weight: bold;
}

.modal-close {
    background: #ef4444;
    color: white;
    border: none;
    border-radius: 50%;
    width: 40px;
    height: 40px;
    cursor: pointer;
    font-size: 20px;
    display: flex;
    align-items: center;
    justify-content: center;
}

.modal-close:hover {
    background: #dc2626;
}

.prodotto-principale {
    background: rgba(59, 130, 245, 0.1);
    border-left: 4px solid #3b82f6;
    padding: 15px;
    border-radius: 6px;
    margin-bottom: 25px;
}

.accessorio-codice {
    background: #3b82f6;
    color: white;
    font-size: 16px;
    font-weight: bold;
    padding: 10px 12px;
    border-radius: 4px;
    display: inline-block;
    margin-bottom: 10px;
    letter-spacing: 1px;
    word-break: break-all;
}

.accessorio-nome {
    font-size: 15px;
    font-weight: 600;
    color: #e0e0e0;
    margin: 8px 0;
    line-height: 1.4;
}

.accessorio-brand {
    font-size: 12px;
    color: #94a3b8;
    margin: 5px 0;
}

.accessorio-categoria {
    font-size: 11px;
    background: rgba(59, 130, 245, 0.2);
    color: #93c5fd;
    padding: 4px 8px;
    border-radius: 3px;
    display: inline-block;
    margin: 8px 0;
}

.abbinamenti-section {
    margin-bottom: 30px;
}

.abbinamenti-section h3 {
    margin: 0 0 15px 0;
    font-size: 18px;
    font-weight: bold;
    display: flex;
    align-items: center;
    gap: 10px;
}

.section-ufficiali h3 {
    color: #10b981;
}

.section-ufficiali h3::before {
    content: "✅";
    font-size: 20px;
}

.section-alternative h3 {
    color: #f59e0b;
}

.section-alternative h3::before {
    content: "🔹";
    font-size: 20px;
}

.section-esclusi h3 {
    color: #ef4444;
}

.section-esclusi h3::before {
    content: "❌";
    font-size: 20px;
}

.abbinamenti-grid {
    display: grid;
    grid-template-columns: repeat(auto-fill, minmax(280px, 1fr));
    gap: 15px;
}

.accessorio-card {
    background: #1e293b;
    border: 1px solid #334155;
    border-radius: 8px;
    padding: 16px;
    transition: all 0.3s ease;
}

.accessorio-card:hover {
    border-color: #3b82f6;
    box-shadow: 0 0 15px rgba(59, 130, 245, 0.3);
    transform: translateY(-2px);
}

.accessorio-card.escluso {
    opacity: 0.6;
    border-color: #ef4444;
    background: rgba(239, 68, 68, 0.1);
}

.accessorio-card.escluso .accessorio-codice {
    background: #ef4444;
}

.accessorio-note {
    font-size: 12px;
    color: #cbd5e1;
    margin: 8px 0;
    font-style: italic;
    line-height: 1.3;
}

.accessorio-vincolo {
    font-size: 11px;
    background: rgba(245, 158, 11, 0.2);
    color: #fcd34d;
    padding: 6px 8px;
    border-radius: 3px;
    margin: 8px 0;
    border-left: 2px solid #f59e0b;
}

.accessorio-escluso-motivo {
    font-size: 11px;
    background: rgba(239, 68, 68, 0.2);
    color: #fca5a5;
    padding: 6px 8px;
    border-radius: 3px;
    margin: 8px 0;
    border-left: 2px solid #ef4444;
}

.accessorio-azioni {
    display: flex;
    gap: 8px;
    margin-top: 12px;
}

.btn-aggiungi {
    background: #10b981;
    color: white;
    border: none;
    padding: 8px 12px;
    border-radius: 4px;
    font-size: 12px;
    cursor: pointer;
    flex: 1;
    font-weight: 600;
    transition: all 0.2s;
}

.btn-aggiungi:hover {
    background: #059669;
}

.btn-aggiungi:disabled {
    background: #6b7280;
    cursor: not-allowed;
}

.accessorio-prezzo {
    font-size: 13px;
    color: #10b981;
    font-weight: 600;
    margin-top: 5px;
}
</style>

<div class="modal-abbinamenti" id="modalAbbinamenti">
    <div class="modal-abbinamenti-content">
        <div class="modal-header">
            <h2>Abbinamenti Prodotto</h2>
            <button class="modal-close" onclick="chiudiModalAbbinamenti()">✕</button>
        </div>
        
        <div class="prodotto-principale" id="prodottoPrincipale"></div>
        <div class="abbinamenti-section section-ufficiali" id="sezioneUfficiali"></div>
        <div class="abbinamenti-section section-alternative" id="sezioneAlternative"></div>
        <div class="abbinamenti-section section-esclusi" id="sezioneEsclusi"></div>
        
        <div style="display:flex; gap:10px; margin-top:20px; padding-top:15px; border-top:1px solid #e5e7eb;">
            <button onclick="chiudiModalAbbinamenti()" style="flex:1; padding:12px; background:#d1d5db; color:#374151; border:none; border-radius:6px; cursor:pointer; font-weight:600;">Annulla</button>
            <button onclick="aggiungiAbbinamenti()" style="flex:1; padding:12px; background:#10b981; color:white; border:none; border-radius:6px; cursor:pointer; font-weight:600;">✓ Aggiungi Voce</button>
        </div>
    </div>
</div>

<script>
function mostraAbbinamenti(prodottoId) {
    fetch(`/api/abbina/${prodottoId}`)
        .then(r => r.json())
        .then(data => {
            if (!data.prodotto || !data.prodotto.codice) {
                alert("Prodotto non trovato");
                return;
            }
            // Aggiorna il drawer a destra INVECE di aprire un modal
            aggiornaPannelloAccessori(data);
        })
        .catch(e => {
            console.error("Errore:", e);
            alert("Errore nel caricamento degli abbinamenti");
        });
}

function aggiornaPannelloAccessori(data) {
    const prod = data.prodotto;
    
    // Titolo prodotto
    let html = `<div style="font-size:11px;color:#ef4444;font-weight:bold;margin-bottom:12px;padding:8px;background:rgba(239,68,68,0.1);border-radius:4px;">
        <div>${prod.codice}</div>
        <div>${prod.nome}</div>
    </div>`;
    
    // UFFICIALI
    if (data.ufficiali && data.ufficiali.length > 0) {
        html += `<div style="font-size:10px;color:#ef4444;font-weight:bold;margin-bottom:6px;">✓ UFFICIALI</div>`;
        data.ufficiali.forEach(acc => {
            html += `
            <div style="padding:8px;margin-bottom:6px;background:rgba(239,68,68,0.1);border-left:2px solid #ef4444;border-radius:3px;cursor:pointer;" onclick="aggiungiAccessorio('${acc.accessorio_id}','${acc.nome}')">
                <div style="font-size:10px;color:#fff;font-weight:bold;margin-bottom:2px;">${acc.nome}</div>
                <div style="font-size:9px;color:#9ca3af;">ID: ${acc.accessorio_id}</div>
                <div style="font-size:10px;color:#10b981;margin-top:4px;">→ Clicca per aggiungere</div>
            </div>
            `;
        });
    }
    
    // ALTERNATIVE
    if (data.alternative && data.alternative.length > 0) {
        html += `<div style="font-size:10px;color:#f59e0b;font-weight:bold;margin-top:12px;margin-bottom:6px;">★ ALTERNATIVE</div>`;
        data.alternative.forEach(acc => {
            html += `
            <div style="padding:8px;margin-bottom:6px;background:rgba(245,158,11,0.1);border-left:2px solid #f59e0b;border-radius:3px;cursor:pointer;" onclick="aggiungiAccessorio('${acc.accessorio_id}','${acc.nome}')">
                <div style="font-size:10px;color:#fff;font-weight:bold;margin-bottom:2px;">${acc.nome}</div>
                <div style="font-size:9px;color:#9ca3af;">ID: ${acc.accessorio_id}</div>
                <div style="font-size:10px;color:#f59e0b;margin-top:4px;">→ Clicca per aggiungere</div>
            </div>
            `;
        });
    }
    
    document.getElementById('sezioneUfficiali').innerHTML = html;
    document.getElementById('sezioneAlternative').innerHTML = '';
}

function popolaModalAbbinamenti(data) {
    const prod = data.prodotto;
    const htmlProd = `
        <div class="accessorio-codice">${prod.codice}</div>
        <div class="accessorio-nome">${prod.nome}</div>
        <div class="accessorio-brand">Collezione: ${prod.collezione}</div>
        <div class="accessorio-categoria">Categoria: ${prod.categoria}</div>
        <div class="accessorio-prezzo">💰 Cliente: €${prod.prezzo_cliente} | Riv: €${prod.prezzo_riv}</div>
    `;
    document.getElementById('prodottoPrincipale').innerHTML = htmlProd;
    
    if (data.ufficiali.length > 0) {
        const html = `<h3>Abbinamenti Ufficiali</h3><div class="abbinamenti-grid">${data.ufficiali.map(acc => creaCardAccessorio(acc)).join('')}</div>`;
        document.getElementById('sezioneUfficiali').innerHTML = html;
    } else {
        document.getElementById('sezioneUfficiali').innerHTML = '';
    }
    
    if (data.alternative.length > 0) {
        const html = `<h3>Abbinamenti Alternativi</h3><div class="abbinamenti-grid">${data.alternative.map(acc => creaCardAccessorio(acc)).join('')}</div>`;
        document.getElementById('sezioneAlternative').innerHTML = html;
    } else {
        document.getElementById('sezioneAlternative').innerHTML = '';
    }
    
    if (data.esclusi.length > 0) {
        const html = `<h3>Non Compatibili</h3><div class="abbinamenti-grid">${data.esclusi.map(acc => creaCardAccessorio(acc, true)).join('')}</div>`;
        document.getElementById('sezioneEsclusi').innerHTML = html;
    } else {
        document.getElementById('sezioneEsclusi').innerHTML = '';
    }
}

function creaCardAccessorio(acc, escluso = false) {
    const classeCard = escluso ? 'accessorio-card escluso' : 'accessorio-card';
    const htmlVincolo = acc.vincolo_messaggio ? `<div class="accessorio-vincolo">⚠️ ${acc.vincolo_messaggio}</div>` : '';
    const htmlMotivo = escluso ? `<div class="accessorio-escluso-motivo">❌ Non compatibile</div>` : '';
    
    return `
        <div class="${classeCard}">
            <div class="accessorio-codice">${acc.id}</div>
            <div class="accessorio-nome">${acc.nome}</div>
            <div class="accessorio-brand">${acc.brand}</div>
            <div class="accessorio-categoria">${acc.categoria}</div>
            ${acc.nota_prodotto ? `<div class="accessorio-note">${acc.nota_prodotto}</div>` : ''}
            ${htmlVincolo}
            ${htmlMotivo}
            <div class="accessorio-azioni">
                <label style="display:flex; align-items:center; gap:8px; cursor:${escluso ? 'not-allowed' : 'pointer'};">
                    <input type="checkbox" data-abbinamento-id="${acc.id}" ${escluso ? 'disabled' : ''} style="width:18px; height:18px;">
                    <span>${escluso ? '❌ Non disponibile' : '✓ Seleziona'}</span>
                </label>
            </div>
        </div>
    `;
}

function aggiungiAccessorio(accessorioId) {
    console.log("Aggiunto accessorio:", accessorioId);
    alert("Accessorio aggiunto al cantiere");
}

function aggiungiAbbinamenti() {
    // RACCOGLIAMO: 1 PADRE + N ABBINAMENTI SELEZIONATI
    const modal = document.getElementById('modalAbbinamenti');
    const prodPrincipaleDiv = document.getElementById('prodottoPrincipale');
    
    // Estrai il PADRE dal div (è il primo che abbiamo aperto)
    const codicePadre = prodPrincipaleDiv.querySelector('.accessorio-codice')?.textContent || '';
    const nomePadre = prodPrincipaleDiv.querySelector('.accessorio-nome')?.textContent || '';
    const collezionePadre = prodPrincipaleDiv.querySelector('.accessorio-brand')?.textContent?.replace('Collezione: ', '') || '';
    
    // Raccogliamo SOLO gli abbinamenti SELEZIONATI (checkbox checked)
    const checkboxSelezionati = Array.from(
        document.querySelectorAll('input[data-abbinamento-id]:checked')
    );
    
    if (checkboxSelezionati.length === 0) {
        alert('Seleziona almeno un abbinamento');
        return;
    }
    
    const abbinamenti = checkboxSelezionati.map(checkbox => {
        const card = checkbox.closest('.accessorio-card');
        return {
            id: checkbox.dataset.abbinamentoId,
            codice: card.querySelector('.accessorio-codice')?.textContent || '',
            nome: card.querySelector('.accessorio-nome')?.textContent || '',
            brand: card.querySelector('.accessorio-brand')?.textContent || '',
            categoria: card.querySelector('.accessorio-categoria')?.textContent || ''
        };
    });
    
    // PAYLOAD: 1 PADRE + N ABBINAMENTI (STESSA LOGICA DEL LISTINO)
    const payload = {
        padre: {
            codice: codicePadre,
            nome: nomePadre,
            collezione: collezionePadre
        },
        abbinamenti: abbinamenti
    };
    
    console.log('📦 Payload aggiungi abbinamenti:', payload);
    
    // INVIA AL BACKEND
    fetch('/api/add-abbinamenti-voce', {
        method: 'POST',
        headers: { 'Content-Type': 'application/json' },
        body: JSON.stringify(payload)
    })
    .then(r => r.json())
    .then(data => {
        if (data.success) {
            alert(`✅ Voce aggiunta: ${nomePadre} + ${abbinamenti.length} abbinamenti`);
            chiudiModalAbbinamenti();
            // Ricarica la tabella
            if (window.caricaTabellaStanza) window.caricaTabellaStanza();
        } else {
            alert('❌ Errore: ' + (data.error || 'Sconosciuto'));
        }
    })
    .catch(err => console.error('Errore:', err));
}

document.addEventListener('click', function(event) {
    const modal = document.getElementById('modalAbbinamenti');
    if (event.target === modal) {
        chiudiModalAbbinamenti();
    }
});
</script>

</body>
</html>''')


# ============ ENDPOINT V4: LIVELLI/AMBIENTI ============

@app.route('/api/create-cantiere', methods=['POST'])
def api_create_cantiere():
    try:
        data = request.json
        nome = data.get('nome')
        config = data.get('configurazione')
        
        if not nome or not config:
            return jsonify({'error': 'Dati incompleti'}), 400
        
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        
        now = datetime.now().isoformat()
        c.execute("INSERT INTO cantieri (cliente_id, commerciale_id, nome, configurazione_piani, data_creazione, data_aggiornamento) VALUES (?,?,?,?,?,?)",
                  (1, 1, nome, json.dumps(config), now, now))
        
        cid = c.lastrowid
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'cantiere_id': cid, 'message': f'Cantiere {nome} creato'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/get-cantiere/<int:cid>', methods=['GET'])
def api_get_cantiere(cid):
    try:
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        c = conn.cursor()
        
        c.execute('SELECT * FROM cantieri WHERE id = ?', (cid,))
        cant = dict(c.fetchone() or {})
        
        if cant and cant.get('configurazione_piani'):
            cant['configurazione'] = json.loads(cant['configurazione_piani'])
        
        c.execute('SELECT * FROM cantiere_righe WHERE cantiere_id = ? ORDER BY piano, ambiente', (cid,))
        righe = [dict(r) for r in c.fetchall()]
        
        conn.close()
        return jsonify({'cantiere': cant, 'righe': righe})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/crea-cantiere')
def crea_cantiere():
    html = """<!DOCTYPE html>
<html><head><meta charset="UTF-8"><title>Crea Cantiere</title>
<style>
body {font-family:Arial; background:#f5f5f5; padding:20px;}
.container {max-width:900px; margin:0 auto; background:white; padding:30px; border-radius:12px;}
h1 {color:#667eea;}
.piano-card {background:#f9f9f9; border:2px solid #e0e0e0; padding:20px; margin:15px 0; border-radius:8px;}
.piano-title {font-size:1.2em; font-weight:bold; color:#667eea;}
input {width:100%; padding:10px; margin:8px 0; border:1px solid #ddd; border-radius:4px;}
button {background:#667eea; color:white; padding:10px 20px; border:none; border-radius:4px; cursor:pointer; margin:5px;}
button:hover {background:#764ba2;}
.stats {background:#667eea; color:white; padding:15px; border-radius:4px; margin:20px 0;}
</style>
</head><body>
<div class="container">
<h1>🏗️ Crea Cantiere</h1>
<form id="form">
<label>Nome Cantiere:</label>
<input type="text" id="nome" placeholder="Es: Progetto Bagno Milano" required>
<div class="stats">
<div>Piani: <span id="stat-piani">1</span></div>
<div>Ambienti: <span id="stat-ambienti">0</span></div>
</div>
<div id="piani"></div>
<button type="button" onclick="aggiungiPiano()">➕ Aggiungi Piano</button>
<button type="submit" style="background:#28a745;">✨ Crea Cantiere</button>
</form>
<div id="msg"></div>
</div>
<script>
let piani = {"Piano 1": ["Bagno principale"]};
function aggiungiPiano() {let n=Object.keys(piani).length; piani["Piano "+(n+1)]=[""]; render();}
function rimuoviPiano(p) {delete piani[p]; render();}
function aggiungiAmbiente(p) {piani[p].push(""); render();}
function rimuoviAmbiente(p,i) {piani[p].splice(i,1); render();}
function updateAmbiente(p,i,v) {piani[p][i]=v;}
function updateStats() {document.getElementById('stat-piani').textContent=Object.keys(piani).length; let tot=0; Object.values(piani).forEach(a=>{tot+=a.filter(x=>x.trim()).length}); document.getElementById('stat-ambienti').textContent=tot;}
function render() {let html=""; Object.entries(piani).forEach(([p,a])=>{html+=`<div class="piano-card"><div class="piano-title">${p} <button type="button" onclick="rimuoviPiano('${p}')">❌</button></div>`; a.forEach((amb,i)=>{html+=`<input type="text" placeholder="Ambiente" value="${amb}" onchange="updateAmbiente('${p}',${i},this.value)" onkeyup="updateAmbiente('${p}',${i},this.value)"><button type="button" onclick="rimuoviAmbiente('${p}',${i})">✕</button>`}); html+=`<button type="button" onclick="aggiungiAmbiente('${p}')">➕ Ambiente</button></div>`;}); document.getElementById('piani').innerHTML=html; updateStats();}
document.getElementById('form').addEventListener('submit', async (e) => {
    e.preventDefault();
    let config={}, nome=document.getElementById('nome').value;
    Object.entries(piani).forEach(([p,a])=>{let v=a.filter(x=>x.trim()); if(v.length>0) config[p]=v;});
    if(Object.keys(config).length===0) {alert('Aggiungi almeno un ambiente'); return;}
    let res=await fetch('/api/create-cantiere', {method:'POST', headers:{'Content-Type':'application/json'}, body:JSON.stringify({nome,configurazione:config})});
    let data=await res.json();
    document.getElementById('msg').innerHTML=res.ok ? `<p style="color:green">${data.message}</p>` : `<p style="color:red">${data.error}</p>`;
});
render();
</script>
</body></html>"""
    return render_template_string(html)


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=10000)
