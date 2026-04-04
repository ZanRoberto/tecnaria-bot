import os
import json
import re
import sqlite3
from typing import List, Dict, Any, Optional
from datetime import datetime

from fastapi import FastAPI, HTTPException, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel

# ============================================================
# CONFIG BASE
# ============================================================

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
STATIC_DIR = os.path.join(BASE_DIR, "static")
DATA_DIR = os.path.join(STATIC_DIR, "data")
UPLOADS_DIR = os.path.join(DATA_DIR, "uploads")
DB_PATH = os.path.join(DATA_DIR, "oracolo_covolo.db")

os.makedirs(UPLOADS_DIR, exist_ok=True)
os.makedirs(DATA_DIR, exist_ok=True)

# ============================================================
# DATABASE INIT
# ============================================================

def init_db():
    """Initialize SQLite for document storage and queries."""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    
    c.execute('''CREATE TABLE IF NOT EXISTS documents
                 (id INTEGER PRIMARY KEY,
                  filename TEXT UNIQUE,
                  upload_date TIMESTAMP,
                  file_size INTEGER,
                  content TEXT,
                  preview TEXT)''')
    
    c.execute('''CREATE TABLE IF NOT EXISTS queries
                 (id INTEGER PRIMARY KEY,
                  question TEXT,
                  answer TEXT,
                  documents_used TEXT,
                  timestamp TIMESTAMP)''')
    
    conn.commit()
    conn.close()

init_db()

# ============================================================
# FASTAPI APP
# ============================================================

app = FastAPI(title="Oracolo Covolo - Universal Bot")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

if not os.path.isdir(STATIC_DIR):
    os.makedirs(STATIC_DIR, exist_ok=True)

app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")

# ============================================================
# MODELLI PYDANTIC
# ============================================================

class QuestionRequest(BaseModel):
    question: str

class AnswerResponse(BaseModel):
    answer: str
    sources: List[str]
    timestamp: str
    meta: Dict[str, Any]

class DocumentUploadResponse(BaseModel):
    filename: str
    status: str
    message: str

# ============================================================
# DOCUMENT PROCESSING
# ============================================================

def extract_text_from_pdf(file_path: str) -> str:
    """Extract text from PDF."""
    try:
        import PyPDF2
        text = ""
        with open(file_path, 'rb') as f:
            reader = PyPDF2.PdfReader(f)
            for page_num, page in enumerate(reader.pages, 1):
                text += f"[PAGE {page_num}]\n"
                text += page.extract_text() + "\n"
        return text
    except Exception as e:
        return f"Error reading PDF: {str(e)}"

def extract_text_from_excel(file_path: str) -> str:
    """Extract text from Excel."""
    try:
        import openpyxl
        text = ""
        wb = openpyxl.load_workbook(file_path)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            text += f"[SHEET: {sheet_name}]\n"
            for row in ws.iter_rows(values_only=True):
                text += " | ".join(str(cell) if cell else "" for cell in row) + "\n"
        return text
    except Exception as e:
        return f"Error reading Excel: {str(e)}"

def extract_text_from_docx(file_path: str) -> str:
    """Extract text from Word document."""
    try:
        from docx import Document
        doc = Document(file_path)
        text = ""
        for para in doc.paragraphs:
            text += para.text + "\n"
        return text
    except Exception as e:
        return f"Error reading Word: {str(e)}"

def extract_text_from_file(file_path: str) -> str:
    """Extract text from any supported file."""
    ext = os.path.splitext(file_path)[1].lower()
    
    if ext == '.pdf':
        return extract_text_from_pdf(file_path)
    elif ext in ['.xlsx', '.xls']:
        return extract_text_from_excel(file_path)
    elif ext in ['.docx', '.doc']:
        return extract_text_from_docx(file_path)
    elif ext == '.txt':
        with open(file_path, 'r', encoding='utf-8') as f:
            return f.read()
    else:
        return "File type not supported"

# ============================================================
# DOCUMENT STORAGE
# ============================================================

def save_document_to_db(filename: str, content: str, file_size: int) -> bool:
    """Save document to database."""
    try:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        
        preview = content[:500] + "..." if len(content) > 500 else content
        
        c.execute('''INSERT OR REPLACE INTO documents 
                     (filename, upload_date, file_size, content, preview)
                     VALUES (?, ?, ?, ?, ?)''',
                  (filename, datetime.now().isoformat(), file_size, content, preview))
        
        conn.commit()
        conn.close()
        return True
    except Exception as e:
        print(f"Error saving to DB: {e}")
        return False

def get_all_documents() -> List[Dict[str, Any]]:
    """Get all documents from database."""
    try:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute('SELECT filename, upload_date, file_size, preview FROM documents')
        docs = c.fetchall()
        conn.close()
        
        return [
            {
                "filename": doc[0],
                "upload_date": doc[1],
                "file_size": doc[2],
                "preview": doc[3][:100]
            }
            for doc in docs
        ]
    except Exception as e:
        print(f"Error fetching documents: {e}")
        return []

def search_documents(query: str) -> List[Dict[str, Any]]:
    """Search documents by keyword."""
    try:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        
        search_term = f"%{query.lower()}%"
        c.execute('''SELECT filename, content FROM documents 
                     WHERE LOWER(content) LIKE ? OR LOWER(filename) LIKE ?''',
                  (search_term, search_term))
        
        results = c.fetchall()
        conn.close()
        
        return [
            {
                "filename": result[0],
                "content": result[1],
                "match_preview": result[1][max(0, result[1].lower().find(query.lower())-50):
                                          min(len(result[1]), result[1].lower().find(query.lower())+150)]
            }
            for result in results
        ]
    except Exception as e:
        print(f"Error searching documents: {e}")
        return []

# ============================================================
# TEXT NORMALIZATION
# ============================================================

def normalize(text: str) -> str:
    """Normalize text for matching."""
    text = text.lower()
    text = re.sub(r"[^\w\sàèéìòóùç]", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text

# ============================================================
# RESPONSE BUILDING
# ============================================================

def build_oracolo_prompt(question: str, documents: List[Dict], web_results: List[str]) -> str:
    """Build prompt for Claude with document context."""
    
    doc_context = ""
    if documents:
        doc_context = "DOCUMENTI CARICATI:\n"
        for doc in documents[:3]:  # Use top 3 matches
            doc_context += f"\n📄 {doc['filename']}:\n{doc['content'][:1000]}\n"
    
    web_context = ""
    if web_results:
        web_context = "\nINFORMAZIONI DA WEB (AGGIORNAMENTI):\n"
        for result in web_results[:2]:
            web_context += f"\n🌐 {result}\n"
    
    prompt = f"""Tu sei l'Oracolo Covolo - un assistente esperto che conosce TUTTI i dati aziendali di Covolo.

DOCUMENTI DISPONIBILI:
{doc_context}

{web_context}

DOMANDA DELL'UTENTE: {question}

REGOLE:
1. Rispondi SOLO basandoti sui documenti caricati + web search se appropriato
2. Cita sempre il documento da cui hai preso l'informazione
3. Se l'informazione non è nei documenti, dillo chiaramente
4. Sii preciso, chiaro, tecnico
5. Usa codici prodotto, prezzi, lead time quando disponibili
6. Se serve informazione web per aggiornare, includila

Fornisci una risposta completa e accurata:"""
    
    return prompt

# ============================================================
# ENDPOINTS
# ============================================================

@app.get("/")
async def root() -> FileResponse:
    """Serve l'interfaccia HTML."""
    index_path = os.path.join(STATIC_DIR, "index.html")
    if not os.path.exists(index_path):
        raise HTTPException(status_code=500, detail="index.html non trovato")
    return FileResponse(index_path)

@app.get("/api/status")
async def status():
    """Status dell'Oracolo."""
    docs = get_all_documents()
    return {
        "status": "Oracolo Covolo Online",
        "documents_loaded": len(docs),
        "timestamp": datetime.now().isoformat()
    }

@app.get("/api/documents")
async def list_documents():
    """List all uploaded documents."""
    return {"documents": get_all_documents()}

@app.post("/api/upload", response_model=DocumentUploadResponse)
async def upload_document(file: UploadFile = File(...)):
    """Upload and process document."""
    try:
        # Save file
        file_path = os.path.join(UPLOADS_DIR, file.filename)
        with open(file_path, 'wb') as f:
            content = await file.read()
            f.write(content)
        
        # Extract text
        text_content = extract_text_from_file(file_path)
        
        # Save to DB
        file_size = len(content)
        saved = save_document_to_db(file.filename, text_content, file_size)
        
        if saved:
            return DocumentUploadResponse(
                filename=file.filename,
                status="success",
                message=f"Documento caricato: {file.filename}"
            )
        else:
            return DocumentUploadResponse(
                filename=file.filename,
                status="error",
                message="Errore nel salvare il documento"
            )
    
    except Exception as e:
        return DocumentUploadResponse(
            filename=file.filename,
            status="error",
            message=f"Errore: {str(e)}"
        )

@app.post("/api/ask", response_model=AnswerResponse)
async def api_ask(req: QuestionRequest):
    """Ask a question to the Oracolo."""
    question_raw = (req.question or "").strip()
    if not question_raw:
        raise HTTPException(status_code=400, detail="Domanda vuota")
    
    try:
        # 1. Search documents
        matching_docs = search_documents(question_raw)
        
        # 2. Build prompt
        prompt = build_oracolo_prompt(question_raw, matching_docs, [])
        
        # 3. Call Claude (using Anthropic API via artifact)
        # For now, simulate response
        answer = f"""Basandomi sui documenti caricati:

{question_raw}

Risposta costruita dal tuo database Covolo e web research."""
        
        sources = [doc['filename'] for doc in matching_docs]
        
        return AnswerResponse(
            answer=answer,
            sources=sources,
            timestamp=datetime.now().isoformat(),
            meta={
                "documents_searched": len(matching_docs),
                "response_type": "from_documents"
            }
        )
    
    except Exception as e:
        print(f"[ERROR] /api/ask: {e}")
        return AnswerResponse(
            answer=f"Errore nell'elaborazione: {str(e)}",
            sources=[],
            timestamp=datetime.now().isoformat(),
            meta={"error": str(e)}
        )

@app.delete("/api/documents/{filename}")
async def delete_document(filename: str):
    """Delete a document."""
    try:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute('DELETE FROM documents WHERE filename = ?', (filename,))
        conn.commit()
        conn.close()
        
        file_path = os.path.join(UPLOADS_DIR, filename)
        if os.path.exists(file_path):
            os.remove(file_path)
        
        return {"status": "success", "message": f"Documento {filename} eliminato"}
    except Exception as e:
        return {"status": "error", "message": str(e)}

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
