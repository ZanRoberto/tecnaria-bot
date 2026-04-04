"""
ORACOLO COVOLO CON NARRATOR
============================
Backend DeepSeek + MacroruleEngine + NarratorSystem
Sistema intelligente che RACCONTA le soluzioni.
Non risposte fredde - narrativa calda e persuasiva.
"""

import os
import json
import re
import sqlite3
from typing import List, Dict, Any, Optional
from datetime import datetime

from fastapi import FastAPI, HTTPException, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
import httpx

# Import custom systems
from macrorule_engine import MacroruleEngine
from narrator_system import NarratorSystem

# ============================================================
# CONFIG
# ============================================================

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
STATIC_DIR = os.path.join(BASE_DIR, "static")
DATA_DIR = os.path.join(STATIC_DIR, "data")
UPLOADS_DIR = os.path.join(DATA_DIR, "uploads")
DB_PATH = os.path.join(DATA_DIR, "oracolo_covolo.db")
MACRORULE_FILE = os.path.join(BASE_DIR, "macroregole_covolo_universe.json")

# DeepSeek API
DEEPSEEK_API_KEY = os.getenv("DEEPSEEK_API_KEY", "").strip()
DEEPSEEK_API_URL = "https://api.deepseek.com/chat/completions"
DEEPSEEK_MODEL = "deepseek-chat"

os.makedirs(UPLOADS_DIR, exist_ok=True)
os.makedirs(DATA_DIR, exist_ok=True)

# ============================================================
# INITIALIZE INTELLIGENT SYSTEMS
# ============================================================

# MacroruleEngine: Sistema di macroregole
macrorule_engine = MacroruleEngine(MACRORULE_FILE)

# NarratorSystem: Narrative calde
narrator = NarratorSystem()

# ============================================================
# DATABASE
# ============================================================

def init_db():
    """Initialize SQLite."""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    
    c.execute('''CREATE TABLE IF NOT EXISTS documents
                 (id INTEGER PRIMARY KEY,
                  filename TEXT UNIQUE,
                  upload_date TIMESTAMP,
                  file_size INTEGER,
                  content TEXT,
                  preview TEXT)''')
    
    c.execute('''CREATE TABLE IF NOT EXISTS queries
                 (id INTEGER PRIMARY KEY,
                  question TEXT,
                  answer TEXT,
                  documents_used TEXT,
                  timestamp TIMESTAMP,
                  macrorules_applied TEXT)''')
    
    conn.commit()
    conn.close()

init_db()

# ============================================================
# FASTAPI APP
# ============================================================

app = FastAPI(title="Oracolo Covolo - DeepSeek + MacroruleEngine + Narrator")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

if not os.path.isdir(STATIC_DIR):
    os.makedirs(STATIC_DIR, exist_ok=True)

app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")

# ============================================================
# MODELLI
# ============================================================

class QuestionRequest(BaseModel):
    question: str

class AnswerResponse(BaseModel):
    answer: str
    sources: List[str]
    timestamp: str
    meta: Dict[str, Any]

class DocumentUploadResponse(BaseModel):
    filename: str
    status: str
    message: str

# ============================================================
# DOCUMENT PROCESSING
# ============================================================

def extract_text_from_pdf(file_path: str) -> str:
    """Extract text from PDF."""
    try:
        import PyPDF2
        text = ""
        with open(file_path, 'rb') as f:
            reader = PyPDF2.PdfReader(f)
            for page_num, page in enumerate(reader.pages, 1):
                text += f"[PAGE {page_num}]\n"
                text += page.extract_text() + "\n"
        return text
    except Exception as e:
        return f"Error: {str(e)}"

def extract_text_from_excel(file_path: str) -> str:
    """Extract text from Excel."""
    try:
        import openpyxl
        text = ""
        wb = openpyxl.load_workbook(file_path)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            text += f"[SHEET: {sheet_name}]\n"
            for row in ws.iter_rows(values_only=True):
                text += " | ".join(str(cell) if cell else "" for cell in row) + "\n"
        return text
    except Exception as e:
        return f"Error: {str(e)}"

def extract_text_from_docx(file_path: str) -> str:
    """Extract text from Word."""
    try:
        from docx import Document
        doc = Document(file_path)
        text = ""
        for para in doc.paragraphs:
            text += para.text + "\n"
        return text
    except Exception as e:
        return f"Error: {str(e)}"

def extract_text_from_file(file_path: str) -> str:
    """Extract text from any supported file."""
    ext = os.path.splitext(file_path)[1].lower()
    
    if ext == '.pdf':
        return extract_text_from_pdf(file_path)
    elif ext in ['.xlsx', '.xls']:
        return extract_text_from_excel(file_path)
    elif ext in ['.docx', '.doc']:
        return extract_text_from_docx(file_path)
    elif ext == '.txt':
        with open(file_path, 'r', encoding='utf-8') as f:
            return f.read()
    else:
        return "File type not supported"

# ============================================================
# DOCUMENT STORAGE
# ============================================================

def save_document_to_db(filename: str, content: str, file_size: int) -> bool:
    """Save document to database."""
    try:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        preview = content[:500] + "..." if len(content) > 500 else content
        c.execute('''INSERT OR REPLACE INTO documents 
                     (filename, upload_date, file_size, content, preview)
                     VALUES (?, ?, ?, ?, ?)''',
                  (filename, datetime.now().isoformat(), file_size, content, preview))
        conn.commit()
        conn.close()
        return True
    except Exception as e:
        print(f"Error saving: {e}")
        return False

def get_all_documents() -> List[Dict]:
    """Get all documents."""
    try:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute('SELECT filename, upload_date, file_size, preview FROM documents')
        docs = c.fetchall()
        conn.close()
        return [{"filename": doc[0], "upload_date": doc[1], "file_size": doc[2], "preview": doc[3][:100]}
                for doc in docs]
    except Exception:
        return []

def search_documents(query: str) -> List[Dict]:
    """Search documents."""
    try:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        search_term = f"%{query.lower()}%"
        c.execute('SELECT filename, content FROM documents WHERE LOWER(content) LIKE ? OR LOWER(filename) LIKE ?',
                  (search_term, search_term))
        results = c.fetchall()
        conn.close()
        return [{"filename": result[0], "content": result[1]} for result in results]
    except Exception:
        return []

# ============================================================
# INTELLIGENT RESPONSE WITH NARRATOR
# ============================================================

async def generate_intelligent_answer(question: str, documents: List[Dict]) -> str:
    """
    Genera risposta intelligente:
    1. Controlla macroregole (incompatibilità, requirements)
    2. Se match: ritorna risposta narrativa
    3. Se no match: chiama DeepSeek con contesto
    """
    
    # 1. Parse question per capire cosa chiede
    parsed = parse_question(question)
    
    # 2. Controlla incompatibilità macroregole
    incomp = macrorule_engine.check_incompatibility(parsed)
    if incomp:
        alternatives = incomp.get("alternatives", [])
        context = {
            "product": parsed.get("product", "questo prodotto"),
            "requirement": incomp.get("requirements", ""),
            "problem": incomp.get("message", ""),
            "alternative_1": alternatives[0] if len(alternatives) > 0 else {},
            "alternative_2": alternatives[1] if len(alternatives) > 1 else {},
        }
        return narrator.narrate_answer("incompatibilita_critica", context)
    
    # 3. Controlla requirements
    reqs = macrorule_engine.check_requirements(parsed)
    if reqs:
        context = {
            "product": parsed.get("product", ""),
            "requires": [r.get("requirement", []) for r in reqs],
            "why": reqs[0].get("why", "") if reqs else "",
            "cost": reqs[0].get("cost", 0) if reqs else 0,
        }
        return narrator.narrate_answer("requirement", context)
    
    # 4. Se nessuna macroregola match: chiama DeepSeek
    doc_context = "\n".join([f"📄 {d['filename']}:\n{d['content'][:500]}" for d in documents[:2]])
    
    prompt = f"""Tu sei un consulente ESPERTO di arredo bagno per Covolo SRL.

DOCUMENTI CARICATI:
{doc_context}

DOMANDA DEL CLIENTE: {question}

ISTRUZIONI CRITICHE:
1. Rispondi COME UN CONSULENTE, non come una macchina
2. Racconta la SOLUZIONE, non solo dati
3. Spiega il PERCHÉ, non solo il COSA
4. Proponi ALTERNATIVE quando possibile
5. Sii CALDO ma PROFESSIONALE
6. Citaprodotti/codici dal documento se disponibili
7. Dai TIMELINE REALISTICI

Scrivi come parleresti a un cliente in un caffè - caldo, umano, esperto."""

    try:
        async with httpx.AsyncClient() as client:
            response = await client.post(
                DEEPSEEK_API_URL,
                headers={
                    "Authorization": f"Bearer {DEEPSEEK_API_KEY}",
                    "Content-Type": "application/json"
                },
                json={
                    "model": DEEPSEEK_MODEL,
                    "messages": [{
                        "role": "system",
                        "content": "Sei un consulente professionista di arredo bagno. Rispondi sempre come narrativa calda e esperta, non fredda e tecnica."
                    }, {
                        "role": "user",
                        "content": prompt
                    }],
                    "temperature": 0.4,
                    "max_tokens": 2000,
                },
                timeout=30.0
            )
            
            if response.status_code == 200:
                return response.json()["choices"][0]["message"]["content"]
            else:
                return f"Errore DeepSeek: {response.text}"
    except Exception as e:
        return f"Errore: {str(e)}"

def parse_question(question: str) -> Dict:
    """Parse question per estrarre chiave: product, type, ecc."""
    return {
        "product": question.split("mi serve")[1].split("?")[0].strip() if "mi serve" in question else "",
        "raw": question
    }

# ============================================================
# ENDPOINTS
# ============================================================

@app.get("/")
async def root() -> FileResponse:
    """Serve HTML interface."""
    index_path = os.path.join(STATIC_DIR, "index.html")
    if not os.path.exists(index_path):
        raise HTTPException(status_code=500, detail="index.html not found")
    return FileResponse(index_path)

@app.get("/api/status")
async def status():
    """Status endpoint."""
    docs = get_all_documents()
    macrorules_stats = macrorule_engine.debug_info()
    
    return {
        "status": "Oracolo Covolo Online - DeepSeek + MacroruleEngine + Narrator",
        "documents_loaded": len(docs),
        "deepseek_configured": bool(DEEPSEEK_API_KEY),
        "macroregole_engine": macrorules_stats,
        "timestamp": datetime.now().isoformat()
    }

@app.get("/api/documents")
async def list_documents():
    """List documents."""
    return {"documents": get_all_documents()}

@app.post("/api/upload", response_model=DocumentUploadResponse)
async def upload_document(file: UploadFile = File(...)):
    """Upload document."""
    try:
        file_path = os.path.join(UPLOADS_DIR, file.filename)
        with open(file_path, 'wb') as f:
            content = await file.read()
            f.write(content)
        
        text_content = extract_text_from_file(file_path)
        file_size = len(content)
        saved = save_document_to_db(file.filename, text_content, file_size)
        
        if saved:
            return DocumentUploadResponse(
                filename=file.filename,
                status="success",
                message=f"Documento caricato: {file.filename}"
            )
        else:
            return DocumentUploadResponse(
                filename=file.filename,
                status="error",
                message="Errore nel salvare"
            )
    except Exception as e:
        return DocumentUploadResponse(
            filename=file.filename,
            status="error",
            message=f"Errore: {str(e)}"
        )

@app.post("/api/ask", response_model=AnswerResponse)
async def api_ask(req: QuestionRequest):
    """Ask a question."""
    question_raw = (req.question or "").strip()
    if not question_raw:
        raise HTTPException(status_code=400, detail="Domanda vuota")
    
    try:
        # 1. Search documents
        matching_docs = search_documents(question_raw)
        
        if not matching_docs:
            return AnswerResponse(
                answer="Non ho trovato informazioni nei tuoi documenti. Carica il listino Gessi o i tuoi materiali!",
                sources=[],
                timestamp=datetime.now().isoformat(),
                meta={"documents_searched": 0}
            )
        
        # 2. Generate intelligent answer
        answer = await generate_intelligent_answer(question_raw, matching_docs)
        
        sources = [doc['filename'] for doc in matching_docs]
        
        return AnswerResponse(
            answer=answer,
            sources=sources,
            timestamp=datetime.now().isoformat(),
            meta={
                "documents_searched": len(matching_docs),
                "response_type": "intelligent_narrative",
                "ai_model": DEEPSEEK_MODEL
            }
        )
    except Exception as e:
        print(f"Error: {e}")
        return AnswerResponse(
            answer=f"Errore: {str(e)}",
            sources=[],
            timestamp=datetime.now().isoformat(),
            meta={"error": str(e)}
        )

@app.delete("/api/documents/{filename}")
async def delete_document(filename: str):
    """Delete document."""
    try:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute('DELETE FROM documents WHERE filename = ?', (filename,))
        conn.commit()
        conn.close()
        
        file_path = os.path.join(UPLOADS_DIR, filename)
        if os.path.exists(file_path):
            os.remove(file_path)
        
        return {"status": "success", "message": f"Eliminato: {filename}"}
    except Exception as e:
        return {"status": "error", "message": str(e)}

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
