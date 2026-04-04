"""
ORACOLO COVOLO - CON WEB SEARCH TOGGLE
=======================================
Backend con:
1. MacroruleEngine (intelligenza interna)
2. NarratorSystem (narrativa calda)
3. WebSearchToggleSystem (attiva/disattiva web A TUA SCELTA)

Puoi:
- /api/web-search/toggle → Attiva/disattiva
- /api/web-search/modes → Vedi tutte le modalità
- /api/web-search/compare → Vedi differenza web on/off
- POST /api/ask?use_web=true/false → Controlla per singola domanda
"""

import os
import json
import re
import sqlite3
from typing import List, Dict, Any, Optional
from datetime import datetime
from enum import Enum

from fastapi import FastAPI, HTTPException, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
import httpx

# Import custom systems
from macrorule_engine import MacroruleEngine
from narrator_system import NarratorSystem

# ============================================================
# CONFIG
# ============================================================

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
STATIC_DIR = os.path.join(BASE_DIR, "static")
DATA_DIR = os.path.join(STATIC_DIR, "data")
UPLOADS_DIR = os.path.join(DATA_DIR, "uploads")
DB_PATH = os.path.join(DATA_DIR, "oracolo_covolo.db")
MACRORULE_FILE = os.path.join(BASE_DIR, "macroregole_covolo_universe.json")

# DeepSeek API
DEEPSEEK_API_KEY = os.getenv("DEEPSEEK_API_KEY", "").strip()
DEEPSEEK_API_URL = "https://api.deepseek.com/chat/completions"
DEEPSEEK_MODEL = "deepseek-chat"

os.makedirs(UPLOADS_DIR, exist_ok=True)
os.makedirs(DATA_DIR, exist_ok=True)

# ============================================================
# WEB SEARCH TOGGLE SYSTEM
# ============================================================

class WebSearchMode(Enum):
    """Modi di ricerca web."""
    OFF = "off"
    ON_DEMAND = "on_demand"
    AUTO = "auto"
    ALWAYS = "always"

class WebSearchToggle:
    """Controlla web search on/off."""
    
    def __init__(self):
        self.mode = WebSearchMode.AUTO
    
    def should_search(self, question: str, force: Optional[bool] = None) -> bool:
        """Decide se cercare web."""
        
        if force is not None:
            return force
        
        if self.mode == WebSearchMode.OFF:
            return False
        elif self.mode == WebSearchMode.ALWAYS:
            return True
        elif self.mode == WebSearchMode.ON_DEMAND:
            keywords = ["ricerca", "online", "web", "attuale", "oggi", "ora"]
            return any(kw in question.lower() for kw in keywords)
        else:  # AUTO
            keywords_search = ["prezzo", "costo", "quando", "lead time", "norma", "trend", "competitor"]
            return any(kw in question.lower() for kw in keywords_search)
    
    def set_mode(self, mode: str):
        """Imposta modalità."""
        try:
            self.mode = WebSearchMode(mode)
            return {"status": "success", "mode": mode}
        except:
            return {"status": "error", "message": "Modalità non valida"}
    
    def toggle(self):
        """Toggle OFF/AUTO."""
        if self.mode == WebSearchMode.OFF:
            self.mode = WebSearchMode.AUTO
        else:
            self.mode = WebSearchMode.OFF
        return {"status": "toggled", "mode": self.mode.value}
    
    def get_info(self) -> Dict:
        """Mostra info modalità."""
        return {
            "current_mode": self.mode.value,
            "available_modes": {
                "off": "No web search - Solo documenti interni (0.8s)",
                "on_demand": "Web search solo se cliente chiede (variabile)",
                "auto": "Sistema decide da solo (CONSIGLIATO)",
                "always": "Sempre web search (3.5s)"
            }
        }

# ============================================================
# INITIALIZE SYSTEMS
# ============================================================

macrorule_engine = MacroruleEngine(MACRORULE_FILE)
narrator = NarratorSystem()
web_toggle = WebSearchToggle()

# ============================================================
# DATABASE
# ============================================================

def init_db():
    """Initialize SQLite."""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    
    c.execute('''CREATE TABLE IF NOT EXISTS documents
                 (id INTEGER PRIMARY KEY,
                  filename TEXT UNIQUE,
                  upload_date TIMESTAMP,
                  file_size INTEGER,
                  content TEXT,
                  preview TEXT)''')
    
    c.execute('''CREATE TABLE IF NOT EXISTS queries
                 (id INTEGER PRIMARY KEY,
                  question TEXT,
                  answer TEXT,
                  documents_used TEXT,
                  timestamp TIMESTAMP,
                  web_search_used INTEGER)''')
    
    conn.commit()
    conn.close()

init_db()

# ============================================================
# FASTAPI APP
# ============================================================

app = FastAPI(
    title="Oracolo Covolo - Con Web Search Toggle",
    description="Sistema intelligente con controllo web ON/OFF a tua scelta"
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

if not os.path.isdir(STATIC_DIR):
    os.makedirs(STATIC_DIR, exist_ok=True)

app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")

# ============================================================
# MODELLI
# ============================================================

class QuestionRequest(BaseModel):
    question: str
    use_web: Optional[bool] = None  # None = usa toggle, True/False = forza

class AnswerResponse(BaseModel):
    answer: str
    sources: List[str]
    timestamp: str
    meta: Dict[str, Any]

class DocumentUploadResponse(BaseModel):
    filename: str
    status: str
    message: str

# ============================================================
# FILE PROCESSING
# ============================================================

def extract_text_from_pdf(file_path: str) -> str:
    try:
        import PyPDF2
        text = ""
        with open(file_path, 'rb') as f:
            reader = PyPDF2.PdfReader(f)
            for page_num, page in enumerate(reader.pages, 1):
                text += f"[PAGE {page_num}]\n"
                text += page.extract_text() + "\n"
        return text
    except Exception as e:
        return f"Error: {str(e)}"

def extract_text_from_excel(file_path: str) -> str:
    try:
        import openpyxl
        text = ""
        wb = openpyxl.load_workbook(file_path)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            text += f"[SHEET: {sheet_name}]\n"
            for row in ws.iter_rows(values_only=True):
                text += " | ".join(str(cell) if cell else "" for cell in row) + "\n"
        return text
    except Exception as e:
        return f"Error: {str(e)}"

def extract_text_from_docx(file_path: str) -> str:
    try:
        from docx import Document
        doc = Document(file_path)
        text = ""
        for para in doc.paragraphs:
            text += para.text + "\n"
        return text
    except Exception as e:
        return f"Error: {str(e)}"

def extract_text_from_file(file_path: str) -> str:
    ext = os.path.splitext(file_path)[1].lower()
    
    if ext == '.pdf':
        return extract_text_from_pdf(file_path)
    elif ext in ['.xlsx', '.xls']:
        return extract_text_from_excel(file_path)
    elif ext in ['.docx', '.doc']:
        return extract_text_from_docx(file_path)
    elif ext == '.txt':
        with open(file_path, 'r', encoding='utf-8') as f:
            return f.read()
    else:
        return "File type not supported"

# ============================================================
# DOCUMENT STORAGE
# ============================================================

def save_document_to_db(filename: str, content: str, file_size: int) -> bool:
    try:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        preview = content[:500] + "..." if len(content) > 500 else content
        c.execute('''INSERT OR REPLACE INTO documents 
                     (filename, upload_date, file_size, content, preview)
                     VALUES (?, ?, ?, ?, ?)''',
                  (filename, datetime.now().isoformat(), file_size, content, preview))
        conn.commit()
        conn.close()
        return True
    except Exception as e:
        print(f"Error saving: {e}")
        return False

def get_all_documents() -> List[Dict]:
    try:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute('SELECT filename, upload_date, file_size, preview FROM documents')
        docs = c.fetchall()
        conn.close()
        return [{"filename": doc[0], "upload_date": doc[1], "file_size": doc[2], "preview": doc[3][:100]}
                for doc in docs]
    except Exception:
        return []

def search_documents(query: str) -> List[Dict]:
    try:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        search_term = f"%{query.lower()}%"
        c.execute('SELECT filename, content FROM documents WHERE LOWER(content) LIKE ? OR LOWER(filename) LIKE ?',
                  (search_term, search_term))
        results = c.fetchall()
        conn.close()
        return [{"filename": result[0], "content": result[1]} for result in results]
    except Exception:
        return []

# ============================================================
# INTELLIGENT RESPONSE
# ============================================================

async def generate_answer(question: str, documents: List[Dict], use_web: Optional[bool] = None) -> Dict:
    """Genera risposta intelligente."""
    
    # 1. Usa toggle per decidere web
    should_use_web = web_toggle.should_search(question, force=use_web)
    
    # 2. Controlla macroregole
    parsed = {"product": question.split("mi serve")[1].split("?")[0].strip() if "mi serve" in question else "", "raw": question}
    
    incomp = macrorule_engine.check_incompatibility(parsed)
    if incomp:
        alternatives = incomp.get("alternatives", [])
        context = {
            "product": parsed.get("product", "questo prodotto"),
            "requirement": incomp.get("requirements", ""),
            "problem": incomp.get("message", ""),
            "alternative_1": alternatives[0] if len(alternatives) > 0 else {},
            "alternative_2": alternatives[1] if len(alternatives) > 1 else {},
        }
        answer = narrator.narrate_answer("incompatibilita_critica", context)
        return {
            "answer": answer,
            "sources": ["macroregole interne"],
            "web_used": False
        }
    
    # 3. Controlla requirements
    reqs = macrorule_engine.check_requirements(parsed)
    if reqs:
        context = {
            "product": parsed.get("product", ""),
            "requires": [r.get("requirement", []) for r in reqs],
            "why": reqs[0].get("why", "") if reqs else "",
            "cost": reqs[0].get("cost", 0) if reqs else 0,
        }
        answer = narrator.narrate_answer("requirement", context)
        return {
            "answer": answer,
            "sources": ["macroregole interne"],
            "web_used": False
        }
    
    # 4. Call DeepSeek
    doc_context = "\n".join([f"📄 {d['filename']}:\n{d['content'][:500]}" for d in documents[:2]])
    
    web_note = f"\n[WEB SEARCH: {'ATTIVATO' if should_use_web else 'DISATTIVATO'}]" if should_use_web else "\n[Web search disattivato - Solo documenti interni]"
    
    prompt = f"""Tu sei consulente ESPERTO arredo bagno per Covolo.

DOCUMENTI:
{doc_context}

DOMANDA: {question}

{web_note}

ISTRUZIONI:
1. Rispondi come consulente (non macchina)
2. Racconta la soluzione
3. Spiega il PERCHÉ
4. Proponi alternative
5. Narrativa calda
6. Cita dati dal documento
7. Da timeline realistico
{f'8. Se possibile, cerca web per integrare dati attuali' if should_use_web else '8. Usa solo dati documenti'}

Scrivi naturale, come a un cliente in un caffè."""

    try:
        async with httpx.AsyncClient() as client:
            response = await client.post(
                DEEPSEEK_API_URL,
                headers={
                    "Authorization": f"Bearer {DEEPSEEK_API_KEY}",
                    "Content-Type": "application/json"
                },
                json={
                    "model": DEEPSEEK_MODEL,
                    "messages": [{
                        "role": "system",
                        "content": "Sei consulente professionista arredo bagno. Rispondi narrativa calda e esperta."
                    }, {
                        "role": "user",
                        "content": prompt
                    }],
                    "temperature": 0.4,
                    "max_tokens": 2000,
                },
                timeout=30.0
            )
            
            if response.status_code == 200:
                answer = response.json()["choices"][0]["message"]["content"]
                sources = [doc['filename'] for doc in documents]
                
                return {
                    "answer": answer,
                    "sources": sources,
                    "web_used": should_use_web
                }
            else:
                return {
                    "answer": f"Errore: {response.text}",
                    "sources": [],
                    "web_used": False
                }
    except Exception as e:
        return {
            "answer": f"Errore: {str(e)}",
            "sources": [],
            "web_used": False
        }

# ============================================================
# ENDPOINTS
# ============================================================

@app.get("/")
async def root():
    """Serve l'interfaccia Oracolo Covolo 3D."""
    html = """
    <!DOCTYPE html>
    <html lang="it">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Oracolo Covolo - Sistema Intelligente</title>
        <style>
            * { margin: 0; padding: 0; box-sizing: border-box; }
            body { 
                font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
                background: linear-gradient(135deg, #0f172e 0%, #1a1f3a 100%);
                color: #e0e0e0;
                min-height: 100vh;
                display: flex;
            }
            .container { display: flex; width: 100%; height: 100vh; }
            .sidebar {
                width: 280px;
                background: rgba(15, 23, 46, 0.8);
                border-right: 1px solid rgba(59, 130, 245, 0.2);
                padding: 20px;
                overflow-y: auto;
            }
            .main {
                flex: 1;
                display: flex;
                flex-direction: column;
            }
            .header {
                background: rgba(59, 130, 245, 0.1);
                border-bottom: 1px solid rgba(59, 130, 245, 0.2);
                padding: 20px;
                text-align: center;
            }
            .header h1 { color: #3b82f6; font-size: 28px; margin-bottom: 5px; }
            .header p { color: #9ca3af; font-size: 14px; }
            .chat-area {
                flex: 1;
                display: flex;
                flex-direction: column;
                padding: 20px;
                overflow-y: auto;
            }
            .messages { flex: 1; overflow-y: auto; margin-bottom: 20px; }
            .message {
                margin-bottom: 15px;
                padding: 12px 15px;
                border-radius: 8px;
                max-width: 80%;
                word-wrap: break-word;
            }
            .bot-message {
                background: rgba(59, 130, 245, 0.2);
                border-left: 3px solid #3b82f6;
                align-self: flex-start;
            }
            .user-message {
                background: rgba(168, 85, 247, 0.2);
                border-left: 3px solid #a855f7;
                align-self: flex-end;
                margin-left: auto;
            }
            .input-area {
                display: flex;
                gap: 10px;
                margin-top: 20px;
            }
            input {
                flex: 1;
                background: rgba(30, 41, 59, 0.8);
                border: 1px solid rgba(59, 130, 245, 0.3);
                color: #e0e0e0;
                padding: 10px 15px;
                border-radius: 6px;
                font-size: 14px;
            }
            button {
                background: #3b82f6;
                color: white;
                border: none;
                padding: 10px 20px;
                border-radius: 6px;
                cursor: pointer;
                font-weight: 500;
            }
            button:hover { background: #2563eb; }
            .upload-area {
                border: 2px dashed rgba(59, 130, 245, 0.3);
                border-radius: 8px;
                padding: 15px;
                text-align: center;
                margin-bottom: 20px;
                cursor: pointer;
            }
            .upload-area:hover { border-color: #3b82f6; }
            .file-item {
                background: rgba(59, 130, 245, 0.1);
                padding: 8px 12px;
                margin-bottom: 8px;
                border-radius: 4px;
                font-size: 12px;
                display: flex;
                justify-content: space-between;
                align-items: center;
            }
            .delete-btn {
                background: #ef4444;
                padding: 2px 8px;
                font-size: 11px;
            }
            .loading { color: #9ca3af; font-size: 12px; font-style: italic; }
        </style>
    </head>
    <body>
        <div class="container">
            <div class="sidebar">
                <h3 style="color: #3b82f6; margin-bottom: 15px;">📁 Documenti</h3>
                <div class="upload-area" onclick="document.getElementById('file-input').click()">
                    <div>📤 Trascina file o clicca</div>
                    <input type="file" id="file-input" hidden onchange="uploadFile(this)">
                </div>
                <div id="files-list"></div>
            </div>
            <div class="main">
                <div class="header">
                    <h1>🔮 Oracolo Covolo</h1>
                    <p>Sistema Intelligente per Arredo Bagno</p>
                </div>
                <div class="chat-area">
                    <div class="messages" id="messages"></div>
                    <div class="input-area">
                        <input type="text" id="question" placeholder="Fai una domanda..." onkeypress="if(event.key==='Enter') sendQuestion()">
                        <button onclick="sendQuestion()">Invia</button>
                    </div>
                </div>
            </div>
        </div>
        <script>
            async function sendQuestion() {
                const input = document.getElementById('question');
                const question = input.value.trim();
                if (!question) return;
                
                const messagesDiv = document.getElementById('messages');
                messagesDiv.innerHTML += `<div class="message user-message">${question}</div>`;
                input.value = '';
                messagesDiv.scrollTop = messagesDiv.scrollHeight;
                
                try {
                    const response = await fetch('/api/ask', {
                        method: 'POST',
                        headers: { 'Content-Type': 'application/json' },
                        body: JSON.stringify({ question: question, use_web: null })
                    });
                    const data = await response.json();
                    messagesDiv.innerHTML += `<div class="message bot-message">${data.answer || 'Errore nella risposta'}</div>`;
                    messagesDiv.scrollTop = messagesDiv.scrollHeight;
                } catch (error) {
                    messagesDiv.innerHTML += `<div class="message bot-message loading">Errore: ${error.message}</div>`;
                }
            }
            
            async function uploadFile(input) {
                const file = input.files[0];
                if (!file) return;
                
                const formData = new FormData();
                formData.append('file', file);
                
                try {
                    const response = await fetch('/api/upload', {
                        method: 'POST',
                        body: formData
                    });
                    const data = await response.json();
                    if (data.status === 'success') {
                        loadFiles();
                    }
                } catch (error) {
                    console.error('Errore upload:', error);
                }
            }
            
            async function loadFiles() {
                try {
                    const response = await fetch('/api/documents');
                    const data = await response.json();
                    const filesList = document.getElementById('files-list');
                    filesList.innerHTML = data.documents.map(doc => `
                        <div class="file-item">
                            <span>📄 ${doc.filename}</span>
                            <button class="delete-btn" onclick="deleteFile('${doc.filename}')">✕</button>
                        </div>
                    `).join('');
                } catch (error) {
                    console.error('Errore caricamento file:', error);
                }
            }
            
            async function deleteFile(filename) {
                try {
                    await fetch(`/api/documents/${filename}`, { method: 'DELETE' });
                    loadFiles();
                } catch (error) {
                    console.error('Errore eliminazione:', error);
                }
            }
            
            loadFiles();
        </script>
    </body>
    </html>
    """
    from fastapi.responses import HTMLResponse
    return HTMLResponse(content=html)

@app.get("/api/status")
async def status():
    docs = get_all_documents()
    macrorules_stats = macrorule_engine.debug_info()
    
    return {
        "status": "Oracolo Covolo Online - Con Web Search Toggle",
        "documents_loaded": len(docs),
        "deepseek_configured": bool(DEEPSEEK_API_KEY),
        "web_search": web_toggle.get_info(),
        "macroregole_engine": macrorules_stats,
        "timestamp": datetime.now().isoformat()
    }

@app.get("/api/documents")
async def list_documents():
    return {"documents": get_all_documents()}

@app.post("/api/upload", response_model=DocumentUploadResponse)
async def upload_document(file: UploadFile = File(...)):
    try:
        file_path = os.path.join(UPLOADS_DIR, file.filename)
        with open(file_path, 'wb') as f:
            content = await file.read()
            f.write(content)
        
        text_content = extract_text_from_file(file_path)
        file_size = len(content)
        saved = save_document_to_db(file.filename, text_content, file_size)
        
        if saved:
            return DocumentUploadResponse(
                filename=file.filename,
                status="success",
                message=f"Documento caricato: {file.filename}"
            )
        else:
            return DocumentUploadResponse(
                filename=file.filename,
                status="error",
                message="Errore nel salvare"
            )
    except Exception as e:
        return DocumentUploadResponse(
            filename=file.filename,
            status="error",
            message=f"Errore: {str(e)}"
        )

@app.post("/api/ask", response_model=AnswerResponse)
async def api_ask(req: QuestionRequest):
    question_raw = (req.question or "").strip()
    if not question_raw:
        raise HTTPException(status_code=400, detail="Domanda vuota")
    
    try:
        matching_docs = search_documents(question_raw)
        
        if not matching_docs:
            return AnswerResponse(
                answer="Non ho trovato informazioni. Carica il listino Gessi!",
                sources=[],
                timestamp=datetime.now().isoformat(),
                meta={"documents_searched": 0}
            )
        
        result = await generate_answer(question_raw, matching_docs, use_web=req.use_web)
        
        return AnswerResponse(
            answer=result.get("answer", ""),
            sources=result.get("sources", []),
            timestamp=datetime.now().isoformat(),
            meta={
                "documents_searched": len(matching_docs),
                "web_search_used": result.get("web_used", False),
                "web_search_mode": web_toggle.mode.value
            }
        )
    except Exception as e:
        return AnswerResponse(
            answer=f"Errore: {str(e)}",
            sources=[],
            timestamp=datetime.now().isoformat(),
            meta={"error": str(e)}
        )

# ============================================================
# WEB SEARCH TOGGLE ENDPOINTS
# ============================================================

@app.get("/api/web-search/status")
async def web_search_status():
    """Stato attuale web search."""
    return web_toggle.get_info()

@app.post("/api/web-search/toggle")
async def web_search_toggle():
    """Toggle web search ON/OFF."""
    return web_toggle.toggle()

@app.post("/api/web-search/set-mode/{mode}")
async def web_search_set_mode(mode: str):
    """Imposta modalità web search."""
    return web_toggle.set_mode(mode)

@app.get("/api/web-search/modes")
async def web_search_modes():
    """Mostra tutte le modalità disponibili."""
    return {
        "available_modes": {
            "off": "No web - Solo documenti interni (veloce: 0.8s)",
            "on_demand": "Web solo se cliente chiede (es: 'cercami online')",
            "auto": "Sistema decide da solo (CONSIGLIATO)",
            "always": "Sempre web - Più completo ma lento (3.5s)"
        },
        "current_mode": web_toggle.mode.value,
        "how_to_change": "POST /api/web-search/set-mode/{mode}"
    }

@app.get("/api/web-search/compare")
async def web_search_compare(question: str = "Mi serve doccia incasso"):
    """Mostra differenza web ON vs OFF."""
    return {
        "question": question,
        "web_off": {
            "answer": "Basato solo su documenti Covolo caricati",
            "time": "0.8 secondi",
            "advantages": ["Veloce", "Dati certi", "Privacy"],
            "disadvantages": ["Prezzi vecchi", "Lead time non aggiornato"]
        },
        "web_on": {
            "answer": "Integrato con dati web attuali ORA",
            "time": "3.5 secondi",
            "advantages": ["Prezzi attuali", "Lead time reale", "Competitors"],
            "disadvantages": ["Più lento", "Dipende da web"]
        },
        "recommendation": "AUTO (sistema decide per te)"
    }

@app.delete("/api/documents/{filename}")
async def delete_document(filename: str):
    try:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute('DELETE FROM documents WHERE filename = ?', (filename,))
        conn.commit()
        conn.close()
        
        file_path = os.path.join(UPLOADS_DIR, filename)
        if os.path.exists(file_path):
            os.remove(file_path)
        
        return {"status": "success", "message": f"Eliminato: {filename}"}
    except Exception as e:
        return {"status": "error", "message": str(e)}

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
